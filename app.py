"""
Line Balancing Tool - Modern Web Application

A professional Flask web app for:
- IE departments (planning view): full features on desktop
- Real-time balancing calculations and manual overrides
"""

import json
import io
import os
import tempfile
import uuid
from pathlib import Path
from typing import Dict, List, Optional

import pandas as pd
import matplotlib.pyplot as plt
import matplotlib
from flask import Flask, jsonify, render_template_string, request, send_file
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.line_balancer.models import Operation, Workstation
from src.line_balancer.io_utils import read_operations
from src.line_balancer.sequencing import sort_by_id
from src.line_balancer.metrics import calculate_pitch_time, calculate_pitch_time_from_target, calculate_tolerance_bands, calculate_line_balancing_rate, calculate_balance_delay, calculate_line_efficiency, calculate_smoothing_index, calculate_throughput_rate, calculate_required_minutes
from src.line_balancer.balancing import group_and_balance
from src.line_balancer.report import build_report_dataframe, determine_status
from src.line_balancer.before_balancing_metrics import calculate_all_before_metrics
from src.line_balancer.takt_pitch_comparison import calculate_takt_vs_pitch_comparison
from src.line_balancer.comparison_export import generate_comparison_excel

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50 MB max

# In-memory session storage (use Redis/DB for production)
SESSIONS = {}


def generate_session_id():
    """Generate a unique session ID."""
    return str(uuid.uuid4())[:8]


def store_calculation(session_id: str, data: Dict):
    """Store calculation results in session."""
    SESSIONS[session_id] = data
    print(
        f"Stored calculation for session {session_id}. Total sessions: {len(SESSIONS)}"
    )


def get_calculation(session_id: str) -> Optional[Dict]:
    """Retrieve calculation results from session."""
    result = SESSIONS.get(session_id)
    print(
        f"Retrieved calculation for session {session_id}: {result is not None}"
    )
    return result


def calculate_balance(operations: List[Operation],
                      tolerance: Optional[float] = 0.15,
                      manual_pitch_time: Optional[float] = None,
                      production_target: Optional[int] = None,
                      shift_time_minutes: Optional[float] = None,
                      pitch_time_method: str = "auto",
                      efficiency_percentage: Optional[float] = None,
                      available_time_minutes: Optional[float] = None) -> Dict:
    """
    Run the complete balancing calculation and return all results.
    
    Args:
        operations: List of Operation objects from CSV/Excel
        tolerance: UCL/LCL tolerance (default 15%, only used for auto method)
        manual_pitch_time: Optional manual pitch time override
        production_target: Optional production target for line efficiency calculation and pitch time calculation
        shift_time_minutes: Optional shift time in minutes for line efficiency calculation and pitch time calculation
        pitch_time_method: Method for calculating pitch time ("auto", "manual", "target")
        efficiency_percentage: Optional efficiency percentage for Target calculation (0-100)
        available_time_minutes: Optional available time in minutes for Target calculation
    
    Returns:
        Dictionary with all calculation results
    """
    # Step 1: Sort operations by ID
    sorted_ops = sort_by_id(operations)

    # Step 2: Calculate Pitch Time / Takt Time and balance based on method
    demand_met = None
    target_validation_message = None
    target_recheck_messages = []
    target_recheck_summary = None

    if pitch_time_method == "manual":
        if manual_pitch_time is None or manual_pitch_time <= 0:
            raise ValueError(
                "Manual pitch time must be provided and positive when method is 'manual'."
            )
        pitch_time = manual_pitch_time
        pitch_time_source = "manual"
        # Clear target-related parameters for manual method
        production_target = None
        shift_time_minutes = None
        # No tolerance bands for manual method
        ucl = None
        lcl = None
        # Step 3: Balance operations into workstations
        workstations = group_and_balance(sorted_ops,
                                         pitch_time,
                                         pitch_time,
                                         strict=True)
    elif pitch_time_method == "target_direct":
        if production_target is None or production_target <= 0:
            raise ValueError(
                "Production target must be provided and positive when method is 'target_direct'."
            )
        if shift_time_minutes is None or shift_time_minutes <= 0:
            raise ValueError(
                "Shift time must be provided and positive when method is 'target_direct'."
            )

        takt_time = calculate_pitch_time_from_target(production_target,
                                                     shift_time_minutes)
        pitch_time = takt_time
        pitch_time_source = "By Target Direct"
        ucl = None
        lcl = None
        workstations = group_and_balance(sorted_ops,
                                         pitch_time,
                                         pitch_time,
                                         strict=True)
    elif pitch_time_method == "target":
        if production_target is None or production_target <= 0:
            raise ValueError(
                "Production target must be provided and positive when method is 'target'."
            )
        if shift_time_minutes is None or shift_time_minutes <= 0:
            raise ValueError(
                "Shift time must be provided and positive when method is 'target'."
            )

        # Step 1 — Derive Takt Time from By-Target input
        takt_time = calculate_pitch_time_from_target(production_target,
                                                     shift_time_minutes)
        pitch_time = takt_time
        pitch_time_source = "By Target"

        # Step 2 — Validate BEFORE balancing
        max_sam = max(op.basic_time
                      for op in sorted_ops) if sorted_ops else 0.0
        if max_sam > takt_time:
            demand_met = False
            target_validation_message = "Max SAM (Seconds) exceeds Takt Time. Cannot fulfill customer demand."
        else:
            demand_met = True
            target_validation_message = "Max SAM (Seconds) is within Takt Time. Can fulfill customer demand."

        # Step 3 — Balance the line using Pitch Time Auto logic
        auto_pitch_time = calculate_pitch_time(sorted_ops)
        if tolerance is None:
            tolerance = 0.15
        auto_ucl, auto_lcl = calculate_tolerance_bands(auto_pitch_time,
                                                       tolerance)
        # Use auto-computed values for display and status determination
        ucl = auto_ucl
        lcl = auto_lcl
        # Use strict=True to remove 0.5s relaxation for By Target workflow
        workstations = group_and_balance(sorted_ops,
                                         auto_ucl,
                                         auto_lcl,
                                         strict=True)

        # Step 4 — Recheck balanced result against Takt Time, loop until it passes or safety cap is hit
        attempt = 1
        MAX_ATTEMPTS = 5
        target_recheck_messages = []

        while attempt <= MAX_ATTEMPTS:
            recheck_max_sam = max(
                ws.balancing_sam
                for ws in workstations) if workstations else 0.0
            if recheck_max_sam <= takt_time:
                target_recheck_messages.append(
                    "OK — Balancing result satisfies Takt Time.")
                target_recheck_summary = f"OK — Balancing result satisfies Takt Time (Attempt {attempt})."
                break
            else:
                target_recheck_messages.append(
                    f"NOT OK (attempt {attempt}) — re-balancing required.")
                workstations = group_and_balance(sorted_ops,
                                                 auto_ucl,
                                                 auto_lcl,
                                                 strict=True)
                attempt += 1

        if attempt > MAX_ATTEMPTS:
            target_recheck_messages.append(
                f"Unable to fully satisfy Takt Time after {MAX_ATTEMPTS} balancing attempts — showing best achieved result."
            )
            target_recheck_summary = f"Unable to fully satisfy Takt Time after {MAX_ATTEMPTS} balancing attempts — showing best achieved result."
    else:  # auto (default)
        if tolerance is None:
            tolerance = 0.15  # Default tolerance for auto method
        pitch_time = calculate_pitch_time(sorted_ops)
        pitch_time_source = "calculated"
        # Clear target-related parameters for auto method
        production_target = None
        shift_time_minutes = None
        # Calculate tolerance bands for auto method
        ucl, lcl = calculate_tolerance_bands(pitch_time, tolerance)
        # Step 3: Balance operations into workstations
        workstations = group_and_balance(sorted_ops, ucl, lcl)

    # Step 4: Calculate line balancing rate
    line_balancing_rate = calculate_line_balancing_rate(workstations)

    # Step 4.5: Calculate balance delay
    balance_delay = calculate_balance_delay(workstations, sorted_ops)

    # Step 4.6: Calculate line efficiency (if production target and shift time provided and method is target or target_direct)
    line_efficiency = None
    if (pitch_time_method == "target" or pitch_time_method == "target_direct"
        ) and production_target is not None and shift_time_minutes is not None:
        if production_target <= 0 or shift_time_minutes <= 0:
            raise ValueError(
                "Production target and shift time must be positive numbers.")
        line_efficiency = calculate_line_efficiency(workstations, sorted_ops,
                                                    production_target,
                                                    shift_time_minutes)

    # Step 4.7: Calculate smoothing index
    smoothing_index = calculate_smoothing_index(workstations)

    # Step 4.8: Calculate total basic time (SAM) in minutes
    total_basic_time = sum(
        op.basic_time for op in sorted_ops) / 60  # Convert seconds to minutes

    # Step 4.9: Calculate before-balancing metrics
    # Only pass tolerance for auto method
    before_tolerance = tolerance if pitch_time_method == "auto" else None
    before_metrics = calculate_all_before_metrics(
        sorted_ops, production_target, shift_time_minutes, before_tolerance,
        pitch_time_method, manual_pitch_time, efficiency_percentage,
        available_time_minutes)

    # Step 4.10: Calculate Target (if both efficiency and available time are provided)
    target_before = None
    target_after = None
    if efficiency_percentage is not None and available_time_minutes is not None:
        # Calculate Target for before balancing
        # Target = (Efficiency% × Total Manpower × Available Time(minutes)) / Total Basic Time(SAM in Minutes)
        total_manpower_before = before_metrics["total_manpower"]
        total_basic_time_minutes_before = before_metrics[
            "total_basic_time_minutes"]
        if total_basic_time_minutes_before > 0:
            target_before = (
                efficiency_percentage / 100 * total_manpower_before *
                available_time_minutes) / total_basic_time_minutes_before

        # Calculate Target for after balancing
        total_manpower_after = sum(ws.manpower for ws in workstations)
        total_basic_time_minutes_after = total_basic_time  # Already calculated in minutes
        if total_basic_time_minutes_after > 0:
            target_after = (
                efficiency_percentage / 100 * total_manpower_after *
                available_time_minutes) / total_basic_time_minutes_after

    # Step 4.11: Calculate Labour Productivity for before and after balancing
    # Labour Productivity = Production Target (Customer Demand) / Total Manpower
    # Use production_target if provided (target/target_direct method), otherwise use calculated target
    labour_productivity_before = before_metrics.get("labour_productivity")

    labour_productivity_after = None
    target_for_productivity_after = production_target if production_target is not None else target_after
    total_manpower_after = sum(ws.manpower for ws in workstations)
    if target_for_productivity_after is not None and total_manpower_after > 0:
        labour_productivity_after = target_for_productivity_after / total_manpower_after

    # Step 4.12: Calculate Throughput Rate and Required Minutes for By Target / By Target Direct methods
    throughput_rate = None
    required_minutes = None
    if (pitch_time_method == "target" or pitch_time_method
            == "target_direct") and production_target is not None:
        throughput_rate = calculate_throughput_rate(workstations)
        if line_efficiency is not None:
            required_minutes = calculate_required_minutes(
                production_target, workstations, line_efficiency)

    # Step 5: Build report
    # For By Target method, use auto_pitch_time for display in the table alongside takt_time
    display_pitch_time = auto_pitch_time if pitch_time_method == "target" else pitch_time
    report_df = build_report_dataframe(workstations, ucl, lcl,
                                       display_pitch_time, pitch_time_source)

    # Build return dictionary
    result = {
        "operations": operations,
        "sorted_operations": sorted_ops,
        "pitch_time": pitch_time,
        "pitch_time_source": pitch_time_source,
        "ucl": ucl,
        "lcl": lcl,
        "workstations": workstations,
        "line_balancing_rate": line_balancing_rate,
        "balance_delay": balance_delay,
        "line_efficiency": line_efficiency,
        "smoothing_index": smoothing_index,
        "total_basic_time": total_basic_time,
        "production_target": production_target,
        "shift_time_minutes": shift_time_minutes,
        "report_df": report_df,
        "before_metrics": before_metrics,
        "efficiency_percentage": efficiency_percentage,
        "available_time_minutes": available_time_minutes,
        "target_before": target_before,
        "target_after": target_after,
        "labour_productivity_before": labour_productivity_before,
        "labour_productivity_after": labour_productivity_after,
        "throughput_rate": throughput_rate,
        "required_minutes": required_minutes,
        "demand_met": demand_met,
        "target_validation_message": target_validation_message,
        "target_recheck_messages": target_recheck_messages,
        "target_recheck_summary": target_recheck_summary,
    }

    # Include tolerance for auto and By Target methods
    if pitch_time_source == "calculated" or pitch_time_source == "By Target":
        result["tolerance"] = tolerance

    # For By Target method, also include auto-computed values for display
    if pitch_time_method == "target":
        result["auto_pitch_time"] = auto_pitch_time
        result["auto_ucl"] = auto_ucl
        result["auto_lcl"] = auto_lcl

    return result


def generate_chart_image(workstations,
                         pitch_time,
                         ucl,
                         lcl,
                         pitch_time_source="calculated",
                         y_max=None,
                         auto_pitch_time=None) -> io.BytesIO:
    """
    Generate a bar chart image using matplotlib that matches the client-side Chart.js styling.
    
    Args:
        workstations: List of Workstation objects
        pitch_time: Target pitch time
        ucl: Upper control limit
        lcl: Lower control limit
        pitch_time_source: Source of pitch time calculation ("manual", "By Target", or "calculated")
        y_max: Optional Y-axis maximum for shared scaling
        auto_pitch_time: Auto-calculated pitch time (for By Target method)
    
    Returns:
        BytesIO object containing the PNG image
    """
    # Use non-interactive backend to avoid display issues
    matplotlib.use('Agg')

    # Prepare data
    workstation_names = []
    balancing_sam = []

    for ws in workstations:
        # Join operation names with " + " for each workstation
        op_names = " + ".join(op.name for op in ws.operations)
        workstation_names.append(op_names)
        balancing_sam.append(ws.balancing_sam)

    # Determine pitch time label based on source
    if pitch_time_source == "manual" or pitch_time_source == "By Target Direct":
        pitch_time_label = "Takt Time"
    elif pitch_time_source == "By Target":
        pitch_time_label = "Pitch Time (Auto)"
    else:
        pitch_time_label = "Pitch Time"

    # For By Target method, use auto_pitch_time for display
    display_pitch_time = auto_pitch_time if pitch_time_source == "By Target" and auto_pitch_time is not None else pitch_time

    # Create figure with appropriate size
    fig, ax = plt.subplots(figsize=(18, 9))

    # Create bar chart with blue color matching Chart.js
    bars = ax.bar(range(len(workstation_names)),
                  balancing_sam,
                  color=(59 / 255, 130 / 255, 246 / 255, 0.8),
                  edgecolor=(59 / 255, 130 / 255, 246 / 255, 1.0),
                  linewidth=1,
                  width=0.6,
                  label='Balancing SAM')

    # Add reference lines without labels (labels will be in legend)
    # Show UCL/LCL for auto and By Target methods
    if pitch_time_source == "calculated" or (pitch_time_source == "By Target"
                                             and ucl is not None
                                             and lcl is not None):
        ax.axhline(y=ucl,
                   color=(239 / 255, 68 / 255, 68 / 255),
                   linestyle='--',
                   linewidth=2)
        ax.axhline(y=lcl,
                   color=(249 / 255, 115 / 255, 22 / 255),
                   linestyle='--',
                   linewidth=2)
    # Show pitch time/takt time line for manual and By Target Direct methods
    if pitch_time_source == "manual" or pitch_time_source == "By Target Direct":
        ax.axhline(y=pitch_time,
                   color=(34 / 255, 197 / 255, 94 / 255),
                   linestyle='--',
                   linewidth=2)
    # Show auto pitch time line for auto and By Target methods
    if pitch_time_source == "calculated" or (pitch_time_source == "By Target"
                                             and auto_pitch_time is not None):
        ax.axhline(y=display_pitch_time,
                   color=(34 / 255, 197 / 255, 94 / 255),
                   linestyle='--',
                   linewidth=2)

    # Set Y-axis maximum if provided
    if y_max is not None:
        ax.set_ylim(0, y_max)

    # Set labels and title
    ax.set_xlabel('Workstations',
                  fontsize=12,
                  fontweight='500',
                  color='#333333')
    ax.set_ylabel('Time (seconds)',
                  fontsize=12,
                  fontweight='500',
                  color='#333333')
    ax.set_title('After Balancing Chart',
                 fontsize=14,
                 fontweight='600',
                 color='#3b82f6')

    # Set x-axis ticks to workstation names
    ax.set_xticks(range(len(workstation_names)))
    ax.set_xticklabels(workstation_names,
                       rotation=45,
                       ha='right',
                       fontsize=9,
                       color='#333333')

    # Format y-axis labels (lock ticks first to avoid matplotlib UserWarning)
    from matplotlib.ticker import FixedLocator
    yticks = ax.get_yticks()
    ax.yaxis.set_major_locator(FixedLocator(yticks))
    ax.set_yticklabels([f'{x:.1f}s' for x in yticks],
                       fontsize=10,
                       color='#333333')

    # Set tick colors
    ax.tick_params(axis='x', colors='#333333')
    ax.tick_params(axis='y', colors='#333333')

    # Create custom legend with reference lines
    from matplotlib.lines import Line2D
    legend_elements = [
        Line2D([0], [0],
               color=(59 / 255, 130 / 255, 246 / 255, 0.8),
               lw=4,
               label='Balancing SAM'),
    ]

    # Add pitch time/takt time line for manual and By Target Direct methods
    if pitch_time_source == "manual" or pitch_time_source == "By Target Direct":
        legend_elements.append(
            Line2D([0], [0],
                   color=(34 / 255, 197 / 255, 94 / 255),
                   lw=2,
                   linestyle='--',
                   label=f'{pitch_time_label} {display_pitch_time:.1f}s'))

    # Add UCL/LCL and pitch time line for auto and By Target methods
    if pitch_time_source == "calculated" or (pitch_time_source == "By Target"
                                             and ucl is not None
                                             and lcl is not None):
        legend_elements.extend([
            Line2D([0], [0],
                   color=(34 / 255, 197 / 255, 94 / 255),
                   lw=2,
                   linestyle='--',
                   label=f'{pitch_time_label} {display_pitch_time:.1f}s'),
            Line2D([0], [0],
                   color=(239 / 255, 68 / 255, 68 / 255),
                   lw=2,
                   linestyle='--',
                   label=f'UCL {ucl:.1f}s'),
            Line2D([0], [0],
                   color=(249 / 255, 115 / 255, 22 / 255),
                   lw=2,
                   linestyle='--',
                   label=f'LCL {lcl:.1f}s')
        ])

    ax.legend(handles=legend_elements,
              loc='upper right',
              fontsize=10,
              labelcolor='#333333')

    # Add grid for better readability
    ax.grid(axis='y', alpha=0.1, linestyle='-')
    ax.grid(axis='x', alpha=0.1, linestyle='-')

    # Set background color to white for Excel export
    ax.set_facecolor('white')
    fig.patch.set_facecolor('white')

    # Set axis colors for light theme visibility
    ax.spines['bottom'].set_color('#333333')
    ax.spines['top'].set_color('#333333')
    ax.spines['left'].set_color('#333333')
    ax.spines['right'].set_color('#333333')

    # Adjust layout to prevent label cutoff
    plt.tight_layout()

    # Save to BytesIO
    img_buffer = io.BytesIO()
    plt.savefig(img_buffer, format='png', dpi=150, bbox_inches='tight')
    img_buffer.seek(0)

    # Close the figure to free memory
    plt.close(fig)

    return img_buffer


def generate_before_chart_image(operations,
                                pitch_time,
                                ucl,
                                lcl,
                                pitch_time_source="calculated",
                                y_max=None,
                                auto_pitch_time=None) -> io.BytesIO:
    """
    Generate a before balancing bar chart image using matplotlib.
    
    Args:
        operations: List of Operation objects
        pitch_time: Target pitch time
        ucl: Upper control limit
        lcl: Lower control limit
        pitch_time_source: Source of pitch time calculation ("manual", "By Target", or "calculated")
        y_max: Optional Y-axis maximum for shared scaling
        auto_pitch_time: Auto-calculated pitch time (for By Target method)
    
    Returns:
        BytesIO object containing the PNG image
    """
    # Use non-interactive backend to avoid display issues
    matplotlib.use('Agg')

    # Prepare data
    operation_names = [op.name for op in operations]
    basic_times = [op.basic_time for op in operations]

    # Determine pitch time label based on source
    if pitch_time_source == "manual" or pitch_time_source == "By Target Direct":
        pitch_time_label = "Takt Time"
    elif pitch_time_source == "By Target":
        pitch_time_label = "Pitch Time (Auto)"
    else:
        pitch_time_label = "Pitch Time"

    # For By Target method, use auto_pitch_time for display
    display_pitch_time = auto_pitch_time if pitch_time_source == "By Target" and auto_pitch_time is not None else pitch_time

    # Create figure with appropriate size
    fig, ax = plt.subplots(figsize=(18, 9))

    # Create bar chart with blue color matching Chart.js
    bars = ax.bar(range(len(operation_names)),
                  basic_times,
                  color=(59 / 255, 130 / 255, 246 / 255, 0.8),
                  edgecolor=(59 / 255, 130 / 255, 246 / 255, 1.0),
                  linewidth=1,
                  width=0.6,
                  label='Basic Time (SAM)')

    # Add reference lines without labels (labels will be in legend)
    # Show UCL/LCL for auto and By Target methods
    if pitch_time_source == "calculated" or (pitch_time_source == "By Target"
                                             and ucl is not None
                                             and lcl is not None):
        ax.axhline(y=ucl,
                   color=(239 / 255, 68 / 255, 68 / 255),
                   linestyle='--',
                   linewidth=2)
        ax.axhline(y=lcl,
                   color=(249 / 255, 115 / 255, 22 / 255),
                   linestyle='--',
                   linewidth=2)
    # Show pitch time/takt time line for manual and By Target Direct methods
    if pitch_time_source == "manual" or pitch_time_source == "By Target Direct":
        ax.axhline(y=pitch_time,
                   color=(34 / 255, 197 / 255, 94 / 255),
                   linestyle='--',
                   linewidth=2)
    # Show auto pitch time line for auto and By Target methods
    if pitch_time_source == "calculated" or (pitch_time_source == "By Target"
                                             and auto_pitch_time is not None):
        ax.axhline(y=display_pitch_time,
                   color=(34 / 255, 197 / 255, 94 / 255),
                   linestyle='--',
                   linewidth=2)

    # Set Y-axis maximum if provided
    if y_max is not None:
        ax.set_ylim(0, y_max)

    # Set labels and title
    ax.set_xlabel('Operations', fontsize=12, fontweight='500', color='#333333')
    ax.set_ylabel('Time (seconds)',
                  fontsize=12,
                  fontweight='500',
                  color='#333333')
    ax.set_title('Before Balancing Chart',
                 fontsize=14,
                 fontweight='600',
                 color='#3b82f6')

    # Set x-axis ticks to operation names
    ax.set_xticks(range(len(operation_names)))
    ax.set_xticklabels(operation_names,
                       rotation=45,
                       ha='right',
                       fontsize=9,
                       color='#333333')

    # Format y-axis labels (lock ticks first to avoid matplotlib UserWarning)
    from matplotlib.ticker import FixedLocator
    yticks = ax.get_yticks()
    ax.yaxis.set_major_locator(FixedLocator(yticks))
    ax.set_yticklabels([f'{x:.1f}s' for x in yticks],
                       fontsize=10,
                       color='#333333')

    # Set tick colors
    ax.tick_params(axis='x', colors='#333333')
    ax.tick_params(axis='y', colors='#333333')

    # Create custom legend with reference lines
    from matplotlib.lines import Line2D
    legend_elements = [
        Line2D([0], [0],
               color=(59 / 255, 130 / 255, 246 / 255, 0.8),
               lw=4,
               label='Basic Time (SAM)'),
    ]

    # Add pitch time/takt time line for manual and By Target Direct methods
    if pitch_time_source == "manual" or pitch_time_source == "By Target Direct":
        legend_elements.append(
            Line2D([0], [0],
                   color=(34 / 255, 197 / 255, 94 / 255),
                   lw=2,
                   linestyle='--',
                   label=f'{pitch_time_label} {display_pitch_time:.1f}s'))

    # Add UCL/LCL and pitch time line for auto and By Target methods
    if pitch_time_source == "calculated" or (pitch_time_source == "By Target"
                                             and ucl is not None
                                             and lcl is not None):
        legend_elements.extend([
            Line2D([0], [0],
                   color=(34 / 255, 197 / 255, 94 / 255),
                   lw=2,
                   linestyle='--',
                   label=f'{pitch_time_label} {display_pitch_time:.1f}s'),
            Line2D([0], [0],
                   color=(239 / 255, 68 / 255, 68 / 255),
                   lw=2,
                   linestyle='--',
                   label=f'UCL {ucl:.1f}s'),
            Line2D([0], [0],
                   color=(249 / 255, 115 / 255, 22 / 255),
                   lw=2,
                   linestyle='--',
                   label=f'LCL {lcl:.1f}s')
        ])

    ax.legend(handles=legend_elements,
              loc='upper right',
              fontsize=10,
              labelcolor='#333333')

    # Add grid for better readability
    ax.grid(axis='y', alpha=0.1, linestyle='-')
    ax.grid(axis='x', alpha=0.1, linestyle='-')

    # Set background color to white for Excel export
    ax.set_facecolor('white')
    fig.patch.set_facecolor('white')

    # Set axis colors for light theme visibility
    ax.spines['bottom'].set_color('#333333')
    ax.spines['top'].set_color('#333333')
    ax.spines['left'].set_color('#333333')
    ax.spines['right'].set_color('#333333')

    # Adjust layout to prevent label cutoff
    plt.tight_layout()

    # Save to BytesIO
    img_buffer = io.BytesIO()
    plt.savefig(img_buffer, format='png', dpi=150, bbox_inches='tight')
    img_buffer.seek(0)

    # Close the figure to free memory
    plt.close(fig)

    return img_buffer


# ============== ROUTES ==============


@app.route("/", methods=["GET", "POST"])
def index():
    """Main page: upload file, configure parameters, view results."""
    error = None
    result = None
    session_id = None
    rows = []

    if request.method == "POST":
        try:
            # Get file and parameters
            file = request.files.get("file")
            pitch_time_method = request.form.get("pitch_time_method", "auto")
            pitch_time_str = request.form.get("pitch_time", "")
            tolerance_str = request.form.get("tolerance", "15")
            production_target_str = request.form.get("production_target", "")
            shift_time_str = request.form.get("shift_time", "")
            efficiency_str = request.form.get("efficiency", "")
            available_time_str = request.form.get("available_time", "")

            if not file or file.filename == "":
                error = "Please select a file to upload."
            else:
                # Parse parameters
                manual_pitch_time = float(
                    pitch_time_str) if pitch_time_str else None
                tolerance_percentage = float(
                    tolerance_str) if tolerance_str else None
                production_target = int(
                    production_target_str) if production_target_str else None
                shift_time_minutes = float(
                    shift_time_str) if shift_time_str else None
                efficiency_percentage = float(
                    efficiency_str) if efficiency_str else None
                available_time_minutes = float(
                    available_time_str) if available_time_str else None

                # Validate tolerance range (0-100% input) - only for auto method
                tolerance = 0.15  # Default tolerance
                if pitch_time_method == "auto":
                    if tolerance_percentage is not None:
                        if tolerance_percentage < 0 or tolerance_percentage > 100:
                            error = "Tolerance must be between 0 and 100%."
                        tolerance = tolerance_percentage / 100  # Convert percentage to decimal
                else:
                    # For manual and target methods, ignore tolerance
                    tolerance = None

                # Validate efficiency if provided
                if efficiency_percentage is not None:
                    if efficiency_percentage <= 0 or efficiency_percentage > 100:
                        error = "Efficiency must be between 0 and 100%."

                # Validate available time if provided
                if available_time_minutes is not None:
                    if available_time_minutes <= 0:
                        error = "Available time must be a positive number."
                    # Use a sensible upper bound (e.g., 1440 minutes = 24 hours)
                    if available_time_minutes > 1440:
                        error = "Available time must be less than 1440 minutes (24 hours)."

                # Validate based on pitch time method
                if pitch_time_method == "manual":
                    if manual_pitch_time is None or manual_pitch_time <= 0:
                        error = "Manual Takt time must be provided and positive when method is 'manual'."
                    # Clear shift time and production target for manual method
                    shift_time_minutes = None
                    production_target = None
                    # Keep efficiency and available time for manual method (they are optional)
                elif pitch_time_method == "target" or pitch_time_method == "target_direct":
                    if production_target is None or production_target <= 0:
                        error = "Production target must be provided and positive when method is selected."
                    elif shift_time_minutes is None or shift_time_minutes <= 0:
                        error = "Shift time must be provided and positive when method is selected."
                    # Clear efficiency and available time for target methods (out of scope)
                    efficiency_percentage = None
                    available_time_minutes = None
                else:  # auto
                    # No validation needed for auto method
                    # Clear shift time and production target for auto method
                    shift_time_minutes = None
                    production_target = None
                    # Keep efficiency and available time for auto method (they are optional)

                if not error:
                    # Read operations from file
                    filepath = Path(file.filename)
                    if filepath.suffix.lower() not in (".csv", ".xlsx",
                                                       ".xls"):
                        error = "File must be Excel (.xlsx, .xls) or CSV."
                    else:
                        # Save temporarily and read
                        temp_path = None
                        try:
                            with tempfile.NamedTemporaryFile(
                                    mode='wb',
                                    delete=False,
                                    suffix=Path(file.filename).suffix) as tmp:
                                file.save(tmp.name)
                                temp_path = tmp.name

                            operations = read_operations(temp_path)

                            # Check for errors in operations
                            flagged = [op for op in operations if op.flagged]
                            if flagged:
                                error_list = "<br>".join([
                                    f"Op {op.op_id}: {op.flagged}"
                                    for op in flagged
                                ])
                                error = f"File has validation errors:<br>{error_list}"
                            else:
                                # Run calculation
                                result = calculate_balance(
                                    operations, tolerance, manual_pitch_time,
                                    production_target, shift_time_minutes,
                                    pitch_time_method, efficiency_percentage,
                                    available_time_minutes)

                                # Convert dataframe to list of dicts for template
                                df = result["report_df"]
                                rows = df.to_dict("records")

                                # Format numeric values
                                for row in rows:
                                    row["Combined Basic Time"] = f"{row['Combined Basic Time']:.1f}"
                                    row["Balancing SAM"] = f"{row['Balancing SAM']:.1f}"
                                    # Format new columns if they are numeric - handle dynamic column name
                                    if "Pitch Time" in row and row[
                                            "Pitch Time"]:
                                        row["Pitch Time"] = f"{row['Pitch Time']:.1f}"
                                    elif "Takt Time" in row and row[
                                            "Takt Time"]:
                                        row["Takt Time"] = f"{row['Takt Time']:.1f}"
                                    elif "Pitch Time (Auto)" in row and row[
                                            "Pitch Time (Auto)"]:
                                        row["Pitch Time (Auto)"] = f"{row['Pitch Time (Auto)']:.1f}"
                                    if row.get("UCL"):
                                        row["UCL"] = f"{row['UCL']:.1f}"
                                    if row.get("LCL"):
                                        row["LCL"] = f"{row['LCL']:.1f}"

                                # Generate session ID
                                session_id = generate_session_id()
                                store_calculation(session_id, result)

                        finally:
                            # Clean up temp file
                            if temp_path and os.path.exists(temp_path):
                                try:
                                    os.remove(temp_path)
                                except:
                                    pass

        except Exception as e:
            error = f"Error processing file: {str(e)}"

    return render_template_string(HTML_TEMPLATE,
                                  error=error,
                                  result=result,
                                  rows=rows,
                                  session_id=session_id)


@app.route("/api/export/<format>/<session_id>")
def export(format: str, session_id: str):
    """Export results to Excel."""
    calc = get_calculation(session_id)
    if not calc:
        return jsonify({"error": "Session not found"}), 404

    if calc.get("method_a") and calc.get("method_b"):
        if format == "xlsx":
            excel_buf = generate_comparison_excel(calc)
            filename = f"takt_vs_pitch_comparison_{session_id}.xlsx"
            return send_file(
                excel_buf,
                mimetype=
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                as_attachment=True,
                download_name=filename,
            )
        return jsonify(
            {"error":
             "Only xlsx format supported for comparison export."}), 400

    df = calc["report_df"]

    if format == "xlsx":
        # Calculate shared Y-axis maximum for both charts
        before_basic_times = [
            op.basic_time for op in calc["sorted_operations"]
        ]
        after_balancing_sam = [ws.balancing_sam for ws in calc["workstations"]]

        max_before = max(before_basic_times) if before_basic_times else 0
        max_after = max(after_balancing_sam) if after_balancing_sam else 0
        shared_y_max = max(max_before, max_after) * 1.1  # Add 10% padding

        # Generate before chart image
        before_chart_img = generate_before_chart_image(
            calc["sorted_operations"], calc["pitch_time"],
            calc["ucl"], calc["lcl"],
            calc.get("pitch_time_source", "calculated"), shared_y_max,
            calc.get("auto_pitch_time"))

        # Generate after chart image
        after_chart_img = generate_chart_image(
            calc["workstations"], calc["pitch_time"], calc["ucl"], calc["lcl"],
            calc.get("pitch_time_source", "calculated"), shared_y_max,
            calc.get("auto_pitch_time"))

        # Create Excel workbook with openpyxl
        wb = Workbook()
        worksheet = wb.active
        worksheet.title = "Line Balance Report"

        # Add title and metrics
        worksheet['A1'] = "Line Balancing Report"
        worksheet['A1'].font = Font(size=16, bold=True, color="3B82F6")
        worksheet['A1'].alignment = Alignment(horizontal="center",
                                              vertical="center",
                                              wrap_text=True)
        worksheet.merge_cells('A1:M1')

        # Add metrics below title in the specified order
        current_row = 2

        # 1. Production Target (If available)
        production_target = calc.get('production_target')
        if production_target is not None:
            worksheet[
                f'A{current_row}'] = f"Customer Demand: {production_target} units"
            worksheet[f'A{current_row}'].font = Font(bold=True)
            current_row += 1

        # 2. Shift Time (If available)
        shift_time = calc.get('shift_time_minutes')
        if shift_time is not None:
            # Format to show at most 1 decimal place, truncating (not rounding), removing trailing zeros
            if shift_time == int(shift_time):
                formatted_shift = int(shift_time)
            else:
                truncated = int(shift_time * 10) / 10
                formatted_shift = int(truncated) if truncated == int(
                    truncated) else truncated
            worksheet[
                f'A{current_row}'] = f"Available Time: {formatted_shift} minutes"
            worksheet[f'A{current_row}'].font = Font(bold=True)
            current_row += 1

        # 2.5. Required Efficiency (If available)
        efficiency_percentage = calc.get('efficiency_percentage')
        available_time_for_target = calc.get('available_time_minutes')
        if efficiency_percentage is not None and available_time_for_target is not None:
            # Format to show at most 1 decimal place, truncating (not rounding), removing trailing zeros
            if efficiency_percentage == int(efficiency_percentage):
                formatted_efficiency = int(efficiency_percentage)
            else:
                truncated = int(efficiency_percentage * 10) / 10
                formatted_efficiency = int(truncated) if truncated == int(
                    truncated) else truncated
            worksheet[
                f'A{current_row}'] = f"Required Efficiency: {formatted_efficiency}%"
            worksheet[f'A{current_row}'].font = Font(bold=True)
            current_row += 1

            # Also include the available time used for efficiency calculation
            if available_time_for_target == int(available_time_for_target):
                formatted_available_time = int(available_time_for_target)
            else:
                truncated = int(available_time_for_target * 10) / 10
                formatted_available_time = int(truncated) if truncated == int(
                    truncated) else truncated
            worksheet[
                f'A{current_row}'] = f"Available Time (for Target): {formatted_available_time} minutes"
            worksheet[f'A{current_row}'].font = Font(bold=True)
            current_row += 1

        # 3. No. of Composite operations
        total_composite_operations = len(calc['workstations'])
        worksheet[
            f'A{current_row}'] = f"Composite operations: {total_composite_operations}"
        worksheet[f'A{current_row}'].font = Font(bold=True)
        current_row += 1

        # 4. Total Basic Time (SAM)
        total_basic_time = calc[
            'total_basic_time']  # Already calculated in minutes
        # Format to show at most 1 decimal place, truncating (not rounding), removing trailing zeros
        if total_basic_time == int(total_basic_time):
            formatted_sam = int(total_basic_time)
        else:
            truncated = int(total_basic_time * 10) / 10
            formatted_sam = int(truncated) if truncated == int(
                truncated) else truncated
        worksheet[f'A{current_row}'] = f"SAM: {formatted_sam} min"
        worksheet[f'A{current_row}'].font = Font(bold=True)
        current_row += 1

        # 5. Line Efficiency% (If available)
        line_efficiency = calc.get('line_efficiency')
        if line_efficiency is not None:
            # Format to show at most 1 decimal place, truncating (not rounding), removing trailing zeros
            if line_efficiency == int(line_efficiency):
                formatted_efficiency = int(line_efficiency)
            else:
                truncated = int(line_efficiency * 10) / 10
                formatted_efficiency = int(truncated) if truncated == int(
                    truncated) else truncated
            worksheet[
                f'A{current_row}'] = f"Required Efficiency: {formatted_efficiency}%"
            worksheet[f'A{current_row}'].font = Font(bold=True)
            current_row += 1

        # 6. Total ManPower
        manpower_sum = 0
        for each_ws in calc['workstations']:
            manpower_sum += each_ws.manpower
        total_manpower = manpower_sum
        worksheet[f'A{current_row}'] = f"ManPower: {total_manpower}"
        worksheet[f'A{current_row}'].font = Font(bold=True)
        current_row += 1

        # 7. Pitch Time / Takt Time
        pitch_time_source_tag = calc.get('pitch_time_source', 'calculated')
        if pitch_time_source_tag == "manual":
            source_display = "(Manual)"
            time_display_name = "Takt Time"
            display_time = calc['pitch_time']
        elif pitch_time_source_tag == "By Target":
            source_display = "(By Target)"
            time_display_name = "Takt Time"
            display_time = calc['pitch_time']
        elif pitch_time_source_tag == "By Target Direct":
            source_display = "(By Target Direct)"
            time_display_name = "Takt Time"
            display_time = calc['pitch_time']
        else:  # "calculated" or any other value
            source_display = "(Auto)"
            time_display_name = "Pitch Time"
            display_time = calc['pitch_time']

        # Format to show at most 1 decimal place, truncating (not rounding), removing trailing zeros
        if display_time == int(display_time):
            formatted_pitch = int(display_time)
        else:
            truncated = int(display_time * 10) / 10
            formatted_pitch = int(truncated) if truncated == int(
                truncated) else truncated
        worksheet[
            f'A{current_row}'] = f"{time_display_name}: {formatted_pitch}s {source_display}"
        worksheet[f'A{current_row}'].font = Font(bold=True)
        current_row += 1

        # 7.5. Pitch Time (Auto) for By Target method
        if pitch_time_source_tag == "By Target" and calc.get(
                'auto_pitch_time') is not None:
            auto_pitch = calc['auto_pitch_time']
            if auto_pitch == int(auto_pitch):
                formatted_auto_pitch = int(auto_pitch)
            else:
                truncated = int(auto_pitch * 10) / 10
                formatted_auto_pitch = int(truncated) if truncated == int(
                    truncated) else truncated
            worksheet[
                f'A{current_row}'] = f"Pitch Time (Auto): {formatted_auto_pitch}s"
            worksheet[f'A{current_row}'].font = Font(bold=True)
            current_row += 1

        # 8. Tolerance (for auto and By Target methods)
        tolerance_value = calc.get('tolerance')
        if tolerance_value is not None:
            tolerance_percentage = tolerance_value * 100
            # Format to show at most 1 decimal place, truncating (not rounding), removing trailing zeros
            if tolerance_percentage == int(tolerance_percentage):
                formatted_tolerance = int(tolerance_percentage)
            else:
                truncated = int(tolerance_percentage * 10) / 10
                formatted_tolerance = int(truncated) if truncated == int(
                    truncated) else truncated
            if tolerance_percentage != 15.0:
                tolerance_label = f"Tolerance (Manual): {formatted_tolerance}%"
            else:
                tolerance_label = f"Tolerance: {formatted_tolerance}%"
            worksheet[f'A{current_row}'] = tolerance_label
            worksheet[f'A{current_row}'].font = Font(bold=True)
            current_row += 1

        # 9. UCL (for auto and By Target methods)
        if calc.get('ucl') is not None:
            # Format to show at most 1 decimal place, truncating (not rounding), removing trailing zeros
            if calc['ucl'] == int(calc['ucl']):
                formatted_ucl = int(calc['ucl'])
            else:
                truncated = int(calc['ucl'] * 10) / 10
                formatted_ucl = int(truncated) if truncated == int(
                    truncated) else truncated
            worksheet[f'A{current_row}'] = f"UCL: {formatted_ucl}s"
            worksheet[f'A{current_row}'].font = Font(bold=True)
            current_row += 1

        # 10. LCL (for auto and By Target methods)
        if calc.get('lcl') is not None:
            # Format to show at most 1 decimal place, truncating (not rounding), removing trailing zeros
            if calc['lcl'] == int(calc['lcl']):
                formatted_lcl = int(calc['lcl'])
            else:
                truncated = int(calc['lcl'] * 10) / 10
                formatted_lcl = int(truncated) if truncated == int(
                    truncated) else truncated
            worksheet[f'A{current_row}'] = f"LCL: {formatted_lcl}s"
            worksheet[f'A{current_row}'].font = Font(bold=True)
            current_row += 1

        # 11. Balancing Rate
        # Format to show at most 1 decimal place, truncating (not rounding), removing trailing zeros
        if calc['line_balancing_rate'] == int(calc['line_balancing_rate']):
            formatted_rate = int(calc['line_balancing_rate'])
        else:
            truncated = int(calc['line_balancing_rate'] * 10) / 10
            formatted_rate = int(truncated) if truncated == int(
                truncated) else truncated
        worksheet[f'A{current_row}'] = f"Balancing Rate: {formatted_rate}%"
        worksheet[f'A{current_row}'].font = Font(bold=True)
        current_row += 1

        # 12. Balance Delay
        # Format to show at most 1 decimal place, truncating (not rounding), removing trailing zeros
        if calc['balance_delay'] == int(calc['balance_delay']):
            formatted_delay = int(calc['balance_delay'])
        else:
            truncated = int(calc['balance_delay'] * 10) / 10
            formatted_delay = int(truncated) if truncated == int(
                truncated) else truncated
        worksheet[f'A{current_row}'] = f"Balance Delay: {formatted_delay}%"
        worksheet[f'A{current_row}'].font = Font(bold=True)
        current_row += 1

        # 13. Smoothing Index
        worksheet[
            f'A{current_row}'] = f"Smoothing Index: {calc['smoothing_index']:.2f} min"
        worksheet[f'A{current_row}'].font = Font(bold=True)
        current_row += 1

        # 13.5. Throughput Rate (for By Target and By Target Direct methods)
        throughput_before = calc.get('before_metrics',
                                     {}).get('throughput_rate')
        throughput_after = calc.get('throughput_rate')
        if (pitch_time_source_tag == "By Target" or pitch_time_source_tag
                == "By Target Direct") and (throughput_before is not None
                                            or throughput_after is not None):
            if throughput_before is not None:
                worksheet[
                    f'A{current_row}'] = f"Throughput Rate (Before): {throughput_before:.1f}s"
                worksheet[f'A{current_row}'].font = Font(bold=True)
                current_row += 1
            if throughput_after is not None:
                worksheet[
                    f'A{current_row}'] = f"Throughput Rate (After): {throughput_after:.1f}s"
                worksheet[f'A{current_row}'].font = Font(bold=True)
                current_row += 1

        # 13.6. Required Minutes (for By Target and By Target Direct methods)
        req_min_before = calc.get('before_metrics', {}).get('required_minutes')
        req_min_after = calc.get('required_minutes')
        if (pitch_time_source_tag == "By Target" or pitch_time_source_tag
                == "By Target Direct") and (req_min_before is not None
                                            or req_min_after is not None):
            if req_min_before is not None:
                worksheet[
                    f'A{current_row}'] = f"Required Minutes (Before): {req_min_before:.1f} min"
                worksheet[f'A{current_row}'].font = Font(bold=True)
                current_row += 1
            if req_min_after is not None:
                worksheet[
                    f'A{current_row}'] = f"Required Minutes (After): {req_min_after:.1f} min"
                worksheet[f'A{current_row}'].font = Font(bold=True)
                current_row += 1

        # 14. Target (If available)
        target_before = calc.get('target_before')
        target_after = calc.get('target_after')
        if target_before is not None and target_after is not None:
            # Format to show at most 2 decimal places, truncating (not rounding), removing trailing zeros
            if target_before == int(target_before):
                formatted_target_before = int(target_before)
            else:
                truncated = int(target_before * 100) / 100
                formatted_target_before = int(truncated) if truncated == int(
                    truncated) else truncated

            if target_after == int(target_after):
                formatted_target_after = int(target_after)
            else:
                truncated = int(target_after * 100) / 100
                formatted_target_after = int(truncated) if truncated == int(
                    truncated) else truncated

            worksheet[
                f'A{current_row}'] = f"Target (Before): {formatted_target_before} units"
            worksheet[f'A{current_row}'].font = Font(bold=True)
            current_row += 1

            worksheet[
                f'A{current_row}'] = f"Target (After): {formatted_target_after} units"
            worksheet[f'A{current_row}'].font = Font(bold=True)
            current_row += 1

        # Set start_row for data table (add one row spacing after metrics)
        start_row = current_row + 1

        # Write headers
        headers = list(df.columns)
        for col_idx, header in enumerate(headers, 1):
            cell = worksheet.cell(row=start_row, column=col_idx, value=header)
            cell.font = Font(bold=True, size=11)
            cell.fill = PatternFill(start_color="E8EDF4",
                                    end_color="E8EDF4",
                                    fill_type="solid")
            cell.alignment = Alignment(horizontal="center",
                                       vertical="center",
                                       wrap_text=True)

        # Write data
        for row_idx, row in enumerate(df.itertuples(index=False),
                                      start_row + 1):
            for col_idx, value in enumerate(row, 1):
                cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(horizontal="center",
                                           vertical="center",
                                           wrap_text=True)

                # Format numeric values - but keep workstation identifiers as whole numbers
                if isinstance(value, (int, float)):
                    header_name = headers[col_idx - 1]
                    # For specified columns, truncate to 1 decimal place (not rounding) and remove trailing zeros
                    if header_name in [
                            'Combined Basic Time', 'Balancing SAM',
                            'Pitch Time', 'Takt Time', 'Pitch Time (Auto)',
                            'UCL', 'LCL'
                    ]:
                        if value == int(value):
                            cell.value = int(value)
                            cell.number_format = '0'  # No decimal for whole numbers
                        else:
                            truncated_value = int(value * 10) / 10
                            if truncated_value == int(truncated_value):
                                cell.value = int(truncated_value)
                                cell.number_format = '0'  # No decimal for whole numbers after truncation
                            else:
                                cell.value = truncated_value
                                cell.number_format = '0.0'  # Show 1 decimal place
                    elif header_name == 'Workstation' or isinstance(
                            value, int):
                        cell.number_format = '0'  # No decimal for workstation identifiers and integers
                    else:
                        cell.number_format = '0.1'  # One decimal for other numeric values

        # Auto-adjust column widths for data columns only
        for col_idx in range(1, len(headers) + 1):
            max_length = 0
            for row_idx in range(start_row, start_row + len(df) + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(
                max_length + 2, 30)  # Cap at 30 to prevent overly wide columns
            worksheet.column_dimensions[get_column_letter(
                col_idx)].width = adjusted_width

        # Insert chart images after the data table
        chart_row = start_row + len(df) + 3  # 3 rows gap after data table

        # Add Before Balancing Chart title
        worksheet[f'A{chart_row}'] = "Before Balancing Chart"
        worksheet[f'A{chart_row}'].font = Font(size=12,
                                               bold=True,
                                               color="3B82F6")
        chart_row += 1

        # Add Before Balancing Chart
        before_img = Image(before_chart_img)
        before_img.width = 800
        before_img.height = 400
        worksheet.add_image(before_img, f'A{chart_row}')

        # Move down for After Balancing Chart (approximately 25 rows for the chart image)
        chart_row += 25

        # Add After Balancing Chart title
        worksheet[f'A{chart_row}'] = "After Balancing Chart"
        worksheet[f'A{chart_row}'].font = Font(size=12,
                                               bold=True,
                                               color="3B82F6")
        chart_row += 1

        # Add After Balancing Chart
        after_img = Image(after_chart_img)
        after_img.width = 800
        after_img.height = 400
        worksheet.add_image(after_img, f'A{chart_row}')

        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return send_file(
            buffer,
            mimetype=
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=f"line_balance_{session_id}.xlsx")


@app.route("/api/recalculate", methods=["POST"])
def recalculate():
    """Recalculate with manual overrides including pitch time."""
    data = request.json
    session_id = data.get("session_id")
    manual_pitch_time = data.get("pitch_time")
    production_target = data.get("production_target")
    shift_time_minutes = data.get("shift_time_minutes")
    tolerance = data.get("tolerance")
    efficiency_percentage = data.get("efficiency_percentage")
    available_time_minutes = data.get("available_time_minutes")

    calc = get_calculation(session_id)
    if not calc:
        return jsonify({"error": "Session not found"}), 404

    # Validate manual pitch time if provided
    if manual_pitch_time is not None:
        try:
            manual_pitch_time = float(manual_pitch_time)
            if manual_pitch_time <= 0:
                return jsonify(
                    {"error":
                     "Manual pitch time must be a positive number."}), 400
        except (ValueError, TypeError):
            return jsonify({"error": "Invalid pitch time value."}), 400

    # Validate production target and shift time if provided
    if production_target is not None:
        try:
            production_target = int(production_target)
            if production_target <= 0:
                return jsonify(
                    {"error":
                     "Production target must be a positive number."}), 400
        except (ValueError, TypeError):
            return jsonify({"error": "Invalid production target value."}), 400

    if shift_time_minutes is not None:
        try:
            shift_time_minutes = float(shift_time_minutes)
            if shift_time_minutes <= 0:
                return jsonify(
                    {"error": "Shift time must be a positive number."}), 400
        except (ValueError, TypeError):
            return jsonify({"error": "Invalid shift time value."}), 400

    # Validate efficiency if provided
    if efficiency_percentage is not None:
        try:
            efficiency_percentage = float(efficiency_percentage)
            if efficiency_percentage <= 0 or efficiency_percentage > 100:
                return jsonify(
                    {"error": "Efficiency must be between 0 and 100%."}), 400
        except (ValueError, TypeError):
            return jsonify({"error": "Invalid efficiency value."}), 400

    # Validate available time if provided
    if available_time_minutes is not None:
        try:
            available_time_minutes = float(available_time_minutes)
            if available_time_minutes <= 0:
                return jsonify(
                    {"error":
                     "Available time must be a positive number."}), 400
            if available_time_minutes > 1440:
                return jsonify({
                    "error":
                    "Available time must be less than 1440 minutes (24 hours)."
                }), 400
        except (ValueError, TypeError):
            return jsonify({"error": "Invalid available time value."}), 400

    # Validate tolerance if provided (only for auto method)
    if tolerance is not None:
        try:
            tolerance = float(tolerance)
            if tolerance < 0 or tolerance > 100:
                return jsonify(
                    {"error": "Tolerance must be between 0 and 100."}), 400
            tolerance = tolerance / 100  # Convert percentage to decimal
        except (ValueError, TypeError):
            return jsonify({"error": "Invalid tolerance value."}), 400

    # Recalculate with new parameters if provided
    try:
        operations = calc["operations"]

        # Determine pitch time method based on what parameters are provided
        pitch_time_method = data.get("pitch_time_method")
        if not pitch_time_method:
            if manual_pitch_time is not None:
                pitch_time_method = "manual"
            elif production_target is not None and shift_time_minutes is not None:
                if calc.get("pitch_time_source") == "By Target Direct":
                    pitch_time_method = "target_direct"
                else:
                    pitch_time_method = "target"
            else:
                pitch_time_method = "auto"

        if pitch_time_method == "manual":
            # Clear other parameters for manual method
            production_target = None
            shift_time_minutes = None
            tolerance = None  # No tolerance for manual method
            # Keep efficiency and available time for manual method
        elif pitch_time_method in ("target", "target_direct"):
            # Clear efficiency and available time for target method
            efficiency_percentage = None
            available_time_minutes = None
            tolerance = None  # No tolerance for target method
        else:
            pitch_time_method = "auto"
            # Clear other parameters for auto method
            production_target = None
            shift_time_minutes = None
            # Keep efficiency and available time for auto method
            # Use default tolerance if not provided
            if tolerance is None:
                tolerance = calc.get("tolerance", 0.15)

        result = calculate_balance(operations, tolerance, manual_pitch_time,
                                   production_target, shift_time_minutes,
                                   pitch_time_method, efficiency_percentage,
                                   available_time_minutes)

        # Update session with new calculation
        store_calculation(session_id, result)

        response_data = {
            "status": "ok",
            "result": {
                "pitch_time": result["pitch_time"],
                "pitch_time_source": result["pitch_time_source"],
                "ucl": result["ucl"],
                "lcl": result["lcl"],
                "line_balancing_rate": result["line_balancing_rate"],
                "balance_delay": result["balance_delay"],
                "smoothing_index": result["smoothing_index"],
                "total_basic_time": result["total_basic_time"],
                "production_target": result["production_target"],
                "shift_time_minutes": result["shift_time_minutes"],
                "workstations_count": len(result["workstations"]),
                "before_metrics": result["before_metrics"],
                "efficiency_percentage": result["efficiency_percentage"],
                "available_time_minutes": result["available_time_minutes"],
                "target_before": result["target_before"],
                "target_after": result["target_after"],
                "throughput_rate": result.get("throughput_rate"),
                "required_minutes": result.get("required_minutes"),
            },
            "before_metrics": result["before_metrics"]
        }

        # Only add tolerance and line_efficiency for auto method
        if result["line_efficiency"] is not None:
            response_data["result"]["line_efficiency"] = result[
                "line_efficiency"]

        if result.get("pitch_time_source") == "By Target":
            response_data["result"]["demand_met"] = result.get("demand_met")
            response_data["result"]["target_validation_message"] = result.get(
                "target_validation_message")
            response_data["result"]["target_recheck_messages"] = result.get(
                "target_recheck_messages", [])
            response_data["result"]["target_recheck_summary"] = result.get(
                "target_recheck_summary")
            response_data["result"]["throughput_rate"] = result.get(
                "throughput_rate")
            response_data["result"]["required_minutes"] = result.get(
                "required_minutes")

        return jsonify(response_data)
    except Exception as e:
        return jsonify({"error": f"Recalculation failed: {str(e)}"}), 500


@app.route("/api/chart-data/<session_id>")
def get_chart_data(session_id: str):
    """Get chart data for the home view."""
    print(f"API request for session: {session_id}")
    print(f"Available sessions: {list(SESSIONS.keys())}")

    try:
        calc = get_calculation(session_id)
        if not calc:
            return jsonify({
                "error":
                "Session not found",
                "message":
                "The calculation session has expired. Please reload the data from the home page."
            }), 404

        # Extract workstation data
        workstations = calc["workstations"]
        pitch_time = calc["pitch_time"]
        ucl = calc["ucl"]
        lcl = calc["lcl"]

        # Prepare data for chart with operation names for each workstation
        workstation_names = []
        for ws in workstations:
            # Join operation names with " + " for each workstation
            op_names = " + ".join(op.name for op in ws.operations)
            workstation_names.append(op_names)

        chart_data = {
            "workstations": workstation_names,
            "balancing_sam": [ws.balancing_sam for ws in workstations],
            "pitch_time": pitch_time,
            "pitch_time_source": calc.get("pitch_time_source", "calculated"),
            "line_balancing_rate": calc["line_balancing_rate"],
            "balance_delay": calc["balance_delay"],
            "smoothing_index": calc["smoothing_index"],
            "total_basic_time": calc["total_basic_time"],
            "production_target": calc.get("production_target"),
            "shift_time_minutes": calc.get("shift_time_minutes"),
        }

        # Only include UCL/LCL and tolerance for auto method
        if calc.get("pitch_time_source") == "calculated":
            chart_data["ucl"] = ucl
            chart_data["lcl"] = lcl
            tolerance = calc.get("tolerance", 0.15)
            if tolerance is not None:
                chart_data[
                    "tolerance"] = tolerance * 100  # Convert to percentage for display
        elif calc.get("pitch_time_source") == "By Target":
            # For By Target method, include auto-computed values
            chart_data["ucl"] = ucl
            chart_data["lcl"] = lcl
            tolerance = calc.get("tolerance", 0.15)
            if tolerance is not None:
                chart_data[
                    "tolerance"] = tolerance * 100  # Convert to percentage for display
            chart_data["auto_pitch_time"] = calc.get("auto_pitch_time")

        if calc.get("line_efficiency") is not None:
            chart_data["line_efficiency"] = calc["line_efficiency"]

        return jsonify(chart_data)
    except Exception as e:
        print(f"Error in get_chart_data: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"Failed to load chart data: {str(e)}"}), 500


@app.route("/api/before-chart-data/<session_id>")
def get_before_chart_data(session_id: str):
    """Get before balancing chart data for the home view."""
    print(f"API request for before chart session: {session_id}")

    try:
        calc = get_calculation(session_id)
        if not calc:
            return jsonify({
                "error":
                "Session not found",
                "message":
                "The calculation session has expired. Please reload the data from the home page."
            }), 404

        # Extract before balancing metrics
        before_metrics = calc["before_metrics"]
        operations = calc["sorted_operations"]

        # Prepare data for chart with operation names and basic times
        operation_names = [op.name for op in operations]
        basic_times = [op.basic_time for op in operations]

        chart_data = {
            "operations":
            operation_names,
            "basic_times":
            basic_times,
            "pitch_time":
            before_metrics["pitch_time"],
            "pitch_time_source":
            before_metrics.get("pitch_time_source", "calculated"),
            "balancing_rate":
            before_metrics["balancing_rate"],
            "balance_delay":
            before_metrics["balance_delay"],
            "smoothing_index":
            before_metrics["smoothing_index"],
            "total_basic_time":
            before_metrics["total_basic_time"],
        }

        # Only include UCL/LCL and tolerance for auto method
        if before_metrics.get("pitch_time_source") == "calculated":
            chart_data["ucl"] = before_metrics["ucl"]
            chart_data["lcl"] = before_metrics["lcl"]
            tolerance = before_metrics.get("tolerance", 0.15)
            if tolerance is not None:
                chart_data[
                    "tolerance"] = tolerance * 100  # Convert to percentage for display
        elif before_metrics.get("pitch_time_source") == "By Target":
            # For By Target method, include auto-computed values
            chart_data["ucl"] = before_metrics["ucl"]
            chart_data["lcl"] = before_metrics["lcl"]
            tolerance = before_metrics.get("tolerance", 0.15)
            if tolerance is not None:
                chart_data[
                    "tolerance"] = tolerance * 100  # Convert to percentage for display
            chart_data["auto_pitch_time"] = before_metrics.get(
                "auto_pitch_time")

        if before_metrics.get("line_efficiency") is not None:
            chart_data["line_efficiency"] = before_metrics["line_efficiency"]

        return jsonify(chart_data)
    except Exception as e:
        print(f"Error in get_before_chart_data: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify(
            {"error": f"Failed to load before chart data: {str(e)}"}), 500


@app.route("/layout")
@app.route("/layout/<session_id>")
def layout(session_id: str = None):
    """Layout view with line balancing report table and metrics."""
    calc = None
    rows = []
    if session_id:
        try:
            calc = get_calculation(session_id)
            if calc:
                # Convert dataframe to list of dicts for template
                df = calc["report_df"]
                rows = df.to_dict("records")

                # Format numeric values
                for row in rows:
                    row["Combined Basic Time"] = f"{row['Combined Basic Time']:.1f}"
                    row["Balancing SAM"] = f"{row['Balancing SAM']:.1f}"
                    # Format new columns if they are numeric - handle dynamic column name
                    if "Pitch Time" in row and row["Pitch Time"]:
                        row["Pitch Time"] = f"{row['Pitch Time']:.1f}"
                    elif "Takt Time" in row and row["Takt Time"]:
                        row["Takt Time"] = f"{row['Takt Time']:.1f}"
                    elif "Pitch Time (Auto)" in row and row[
                            "Pitch Time (Auto)"]:
                        row["Pitch Time (Auto)"] = f"{row['Pitch Time (Auto)']:.1f}"
                    if row.get("UCL"):
                        row["UCL"] = f"{row['UCL']:.1f}"
                    if row.get("LCL"):
                        row["LCL"] = f"{row['LCL']:.1f}"
        except Exception as e:
            print(f"Error in layout route: {str(e)}")
            import traceback
            traceback.print_exc()
            calc = None
            rows = []

    return render_template_string(LAYOUT_TEMPLATE,
                                  session_id=session_id,
                                  has_data=calc is not None,
                                  result=calc,
                                  rows=rows)


@app.route("/compare", methods=["GET", "POST"])
@app.route("/takt-vs-pitch", methods=["GET", "POST"])
def takt_vs_pitch():
    """Standalone Takt vs Pitch Comparison view."""
    error = None
    result = None
    session_id = None

    if request.method == "POST":
        try:
            file = request.files.get("file")
            shift_time_str = request.form.get("shift_time", "")
            production_target_str = request.form.get("production_target", "")

            if not file or file.filename == "":
                error = "Please select an Excel or CSV file to upload."
            elif not shift_time_str or not production_target_str:
                error = "Both Shift Time and Production Target are required."
            else:
                shift_time_minutes = float(shift_time_str)
                production_target = int(production_target_str)

                if shift_time_minutes <= 0:
                    error = "Shift time must be a positive number."
                elif production_target <= 0:
                    error = "Production target must be a positive number."
                else:
                    filepath = Path(file.filename)
                    if filepath.suffix.lower() not in (".csv", ".xlsx",
                                                       ".xls"):
                        error = "File must be Excel (.xlsx, .xls) or CSV."
                    else:
                        temp_path = None
                        try:
                            with tempfile.NamedTemporaryFile(
                                    mode='wb',
                                    delete=False,
                                    suffix=Path(file.filename).suffix) as tmp:
                                file.save(tmp.name)
                                temp_path = tmp.name

                            operations = read_operations(temp_path)
                            flagged = [op for op in operations if op.flagged]
                            if flagged:
                                error_list = "<br>".join([
                                    f"Op {op.op_id}: {op.flagged}"
                                    for op in flagged
                                ])
                                error = f"File has validation errors:<br>{error_list}"
                            else:
                                result = calculate_takt_vs_pitch_comparison(
                                    operations,
                                    shift_time_minutes,
                                    production_target,
                                    calc_method=calc_method
                                    if 'calc_method' in locals() else
                                    'comparison') if 'calc_method' in locals(
                                    ) else calculate_takt_vs_pitch_comparison(
                                        operations,
                                        shift_time_minutes,
                                        production_target,
                                    )
                                session_id = generate_session_id()
                                store_calculation(session_id, result)
                        finally:
                            if temp_path and os.path.exists(temp_path):
                                try:
                                    os.unlink(temp_path)
                                except Exception:
                                    pass
        except Exception as e:
            error = f"Error during calculation: {str(e)}"
            import traceback
            traceback.print_exc()

    return render_template_string(
        COMPARISON_TEMPLATE,
        error=error,
        result=result,
        session_id=session_id,
    )


@app.route("/api/takt-vs-pitch", methods=["POST"])
@app.route("/api/compare", methods=["POST"])
def api_takt_vs_pitch():
    """API endpoint for Takt vs Pitch comparison."""
    try:
        file = request.files.get("file")
        shift_time_str = request.form.get("shift_time", "")
        production_target_str = request.form.get("production_target", "")

        if not file or file.filename == "":
            return jsonify({"error": "Please select a file to upload."}), 400
        if not shift_time_str or not production_target_str:
            return jsonify(
                {"error":
                 "Shift time and production target are required."}), 400

        shift_time_minutes = float(shift_time_str)
        production_target = int(production_target_str)

        if shift_time_minutes <= 0 or production_target <= 0:
            return jsonify({
                "error":
                "Shift time and production target must be positive numbers."
            }), 400

        temp_path = None
        try:
            with tempfile.NamedTemporaryFile(mode='wb',
                                             delete=False,
                                             suffix=Path(
                                                 file.filename).suffix) as tmp:
                file.save(tmp.name)
                temp_path = tmp.name

            operations = read_operations(temp_path)
            flagged = [op for op in operations if op.flagged]
            if flagged:
                return jsonify({
                    "error":
                    "Validation errors in file",
                    "flagged_operations": [{
                        "op_id": op.op_id,
                        "error": op.flagged
                    } for op in flagged]
                }), 400

            result = calculate_takt_vs_pitch_comparison(
                operations, shift_time_minutes, production_target)
            session_id = generate_session_id()
            store_calculation(session_id, result)

            # Prepare JSON response (exclude raw dataframe and workstation objects)
            json_result = {
                "session_id": session_id,
                "shift_time_minutes": result["shift_time_minutes"],
                "production_target": result["production_target"],
                "takt_time": result["takt_time"],
                "pitch_time": result["pitch_time"],
                "ucl": result["ucl"],
                "lcl": result["lcl"],
                "total_sam": result["total_sam"],
                "before": result["before"],
                "method_a": {
                    k: v
                    for k, v in result["method_a"].items()
                    if k not in ("report_df", "workstations")
                },
                "method_b": {
                    k: v
                    for k, v in result["method_b"].items()
                    if k not in ("report_df", "workstations")
                },
                "comparison": result["comparison"],
                "recommendations": result.get("recommendations", []),
            }
            return jsonify(json_result)
        finally:
            if temp_path and os.path.exists(temp_path):
                try:
                    os.unlink(temp_path)
                except Exception:
                    pass
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/takt-vs-pitch-chart-data/<session_id>")
def api_takt_vs_pitch_chart_data(session_id: str):
    """Return chart data for Takt vs Pitch Comparison."""
    calc = get_calculation(session_id)
    if not calc or "method_a" not in calc or "method_b" not in calc:
        return jsonify({"error": "Calculation not found."}), 404

    try:
        ws_a = calc["method_a"]["workstations"]
        ws_b = calc["method_b"]["workstations"]
        max_len = max(len(ws_a), len(ws_b))
        labels = [f"WS {i+1}" for i in range(max_len)]

        times_a = [ws.balancing_sam for ws in ws_a]
        times_b = [ws.balancing_sam for ws in ws_b]

        # Before-balancing operation times (each op is its own station)
        sorted_ops = calc.get("sorted_operations", [])
        before_times = [op.basic_time for op in sorted_ops]
        before_labels = [f"Op {i+1}" for i in range(len(sorted_ops))]

        comp = calc.get("comparison", [])
        kpi_labels = [row["metric"] for row in comp]
        kpi_before = [row["before"] for row in comp]
        kpi_method_a = [row["method_a"] for row in comp]
        kpi_method_b = [row["method_b"] for row in comp]
        kpi_units = [row["unit"] for row in comp]

        return jsonify({
            "labels": labels,
            "method_a_times": times_a,
            "method_b_times": times_b,
            "before_times": before_times,
            "before_labels": before_labels,
            "takt_time": calc["takt_time"],
            "ucl": calc["ucl"],
            "lcl": calc["lcl"],
            "pitch_time": calc["pitch_time"],
            "kpi_labels": kpi_labels,
            "kpi_units": kpi_units,
            "kpi_before": kpi_before,
            "kpi_method_a": kpi_method_a,
            "kpi_method_b": kpi_method_b,
        })
    except Exception as e:
        return jsonify({"error": f"Failed to build chart data: {str(e)}"}), 500


@app.route("/api/export/compare/xlsx/<session_id>")
@app.route("/api/export/compare/<format>/<session_id>")
def export_comparison(session_id: str, format: str = "xlsx"):
    """Export comparison results to Excel."""
    calc = get_calculation(session_id)
    if not calc or "method_a" not in calc or "method_b" not in calc:
        return jsonify({"error": "Comparison session not found"}), 404

    excel_buf = generate_comparison_excel(calc)
    filename = f"takt_vs_pitch_comparison_{session_id}.xlsx"
    return send_file(
        excel_buf,
        mimetype=
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


# ============== HTML TEMPLATES ==============

HTML_TEMPLATE = """
<!DOCTYPE html>
<html lang="en" data-theme="dark">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Line Balancing Optimizer</title>
    <script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
    <script src="https://cdn.jsdelivr.net/npm/chartjs-plugin-annotation@3.0.1/dist/chartjs-plugin-annotation.min.js"></script>
    <style>
        :root {
            --bg: #0f1419;
            --surface: #1a2332;
            --surface-2: #243044;
            --border: rgba(255, 255, 255, 0.08);
            --text: #e8edf4;
            --text-muted: #8b9cb3;
            --accent: #3b82f6;
            --accent-hover: #2563eb;
            --success: #22c55e;
            --warning: #f59e0b;
            --danger: #ef4444;
            --shadow: 0 4px 24px rgba(0, 0, 0, 0.35);
            --radius: 12px;
            --radius-sm: 8px;
            --transition: 0.2s ease;
        }

        [data-theme="light"] {
            --bg: #f1f5f9;
            --surface: #ffffff;
            --surface-2: #f8fafc;
            --border: rgba(15, 23, 42, 0.1);
            --text: #0f172a;
            --text-muted: #64748b;
            --accent: #2563eb;
            --accent-hover: #1d4ed8;
            --success: #16a34a;
            --warning: #d97706;
            --danger: #dc2626;
            --shadow: 0 4px 24px rgba(15, 23, 42, 0.08);
        }

        * { 
            margin: 0; 
            padding: 0; 
            box-sizing: border-box; 
        }

        body {
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
            background: var(--bg);
            color: var(--text);
            line-height: 1.6;
            transition: background var(--transition), color var(--transition);
        }

        .container {
            max-width: 1400px;
            margin: 0 auto;
            padding: 32px 24px;
        }

        /* Header */
        .header {
            display: flex;
            justify-content: space-between;
            align-items: center;
            margin-bottom: 40px;
            flex-wrap: wrap;
            gap: 20px;
        }

        .header-content {
            flex: 1;
        }

        .header h1 {
            font-size: 28px;
            font-weight: 700;
            background: linear-gradient(135deg, #fff 0%, #94a3b8 100%);
            -webkit-background-clip: text;
            -webkit-text-fill-color: transparent;
            background-clip: text;
        }

        [data-theme="light"] .header h1 {
            background: linear-gradient(135deg, #1e40af 0%, #3b82f6 100%);
            -webkit-background-clip: text;
            -webkit-text-fill-color: transparent;
            background-clip: text;
        }

        .header p {
            color: var(--text-muted);
            font-size: 14px;
        }

        .theme-toggle {
            background: var(--surface-2);
            border: 1px solid var(--border);
            border-radius: 999px;
            padding: 8px 16px;
            cursor: pointer;
            font-size: 14px;
            font-weight: 500;
            color: var(--text-muted);
            transition: all var(--transition);
        }

        .theme-toggle:hover {
            border-color: var(--accent);
            box-shadow: 0 0 0 3px rgba(59, 130, 246, 0.25);
        }

        .header-actions {
            display: flex;
            align-items: center;
            gap: 12px;
            flex-wrap: wrap;
        }

        .nav-tabs {
            display: flex;
            gap: 6px;
            background: var(--surface-2);
            padding: 4px;
            border-radius: 999px;
            border: 1px solid var(--border);
        }

        .nav-tab {
            padding: 6px 14px;
            border-radius: 999px;
            text-decoration: none;
            font-size: 13px;
            font-weight: 500;
            color: var(--text-muted);
            transition: all var(--transition);
            display: inline-flex;
            align-items: center;
            gap: 6px;
        }

        .nav-tab:hover {
            color: var(--text);
        }

        .nav-tab.active {
            background: var(--accent);
            color: #ffffff;
            box-shadow: 0 2px 8px rgba(37, 99, 235, 0.4);
        }

        /* Metrics Grid */
        .form-card {
            background: var(--surface);
            border: 1px solid var(--border);
            border-radius: var(--radius);
            padding: 28px;
            margin-bottom: 32px;
            box-shadow: var(--shadow);
        }

        .form-card h2 {
            font-size: 12px;
            font-weight: 600;
            text-transform: uppercase;
            letter-spacing: 0.08em;
            color: var(--text-muted);
            margin-bottom: 20px;
        }

        .form-grid {
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 20px;
            align-items: end;
        }

        .field {
            display: flex;
            flex-direction: column;
            gap: 8px;
        }

        .field.hidden {
            display: none;
        }

        label {
            font-size: 13px;
            font-weight: 500;
            color: var(--text-muted);
        }

        input[type="file"],
        input[type="number"],
        select {
            background: var(--surface-2);
            border: 1px solid var(--border);
            border-radius: var(--radius-sm);
            color: var(--text);
            padding: 12px 14px;
            font-size: 14px;
            transition: all var(--transition);
        }

        input:focus,
        select:focus {
            outline: none;
            border-color: var(--accent);
            box-shadow: 0 0 0 3px rgba(59, 130, 246, 0.25);
        }

        input[type="file"]::file-selector-button {
            background: var(--accent);
            color: white;
            border: none;
            border-radius: 6px;
            padding: 8px 16px;
            margin-right: 12px;
            font-weight: 500;
            cursor: pointer;
            transition: background var(--transition);
        }

        input[type="file"]::file-selector-button:hover {
            background: var(--accent-hover);
        }

        button[type="submit"] {
            background: linear-gradient(135deg, var(--accent) 0%, #6366f1 100%);
            color: white;
            border: none;
            border-radius: var(--radius-sm);
            padding: 12px 28px;
            font-size: 14px;
            font-weight: 600;
            cursor: pointer;
            box-shadow: 0 4px 14px rgba(59, 130, 246, 0.4);
            transition: all var(--transition);
        }

        button[type="submit"]:hover {
            transform: translateY(-2px);
            box-shadow: 0 6px 20px rgba(59, 130, 246, 0.5);
        }

        button[type="submit"]:active {
            transform: translateY(0);
        }

        /* Error Message */
        .error-box {
            background: rgba(239, 68, 68, 0.1);
            border: 1px solid var(--danger);
            border-radius: var(--radius);
            color: var(--danger);
            padding: 16px 20px;
            margin-bottom: 24px;
            font-size: 14px;
        }

        /* Target Workflow Notifications */
        .target-workflow-notice-container {
            margin-bottom: 24px;
            display: flex;
            flex-direction: column;
            gap: 12px;
        }

        .target-validation-banner {
            padding: 14px 18px;
            border-radius: var(--radius-sm);
            font-size: 14px;
            font-weight: 500;
            display: flex;
            align-items: center;
            gap: 12px;
            box-shadow: var(--shadow);
            transition: all var(--transition);
        }

        .target-validation-banner.success {
            background: rgba(34, 197, 94, 0.12);
            border: 1px solid var(--success);
            color: var(--success);
        }

        .target-validation-banner.warning {
            background: rgba(245, 158, 11, 0.12);
            border: 1px solid var(--warning);
            color: var(--warning);
        }

        .target-recheck-banner {
            padding: 12px 18px;
            border-radius: var(--radius-sm);
            font-size: 13px;
            display: flex;
            align-items: center;
            justify-content: space-between;
            gap: 12px;
            background: var(--surface);
            border: 1px solid var(--border);
            color: var(--text);
            box-shadow: var(--shadow);
            flex-wrap: wrap;
        }

        /* Metrics Grid */
        .metrics-grid {
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(140px, 1fr));
            gap: 16px;
            margin-bottom: 32px;
        }

        .metric-card {
            background: var(--surface);
            border: 1px solid var(--border);
            border-radius: var(--radius);
            padding: 20px;
            transition: all var(--transition);
        }

        .metric-card:hover {
            border-color: rgba(59, 130, 246, 0.3);
            transform: translateY(-2px);
        }

        .metric-card .label {
            font-size: 12px;
            font-weight: 500;
            text-transform: uppercase;
            letter-spacing: 0.05em;
            color: var(--text-muted);
            margin-bottom: 8px;
        }

        .metric-card .value {
            font-size: 24px;
            font-weight: 600;
            display: flex;
            align-items: center;
            gap: 8px;
            flex-wrap: wrap;
        }

        .pitch-source-badge {
            font-size: 10px;
            font-weight: 600;
            padding: 2px 8px;
            border-radius: 4px;
            text-transform: uppercase;
            letter-spacing: 0.05em;
        }

        .pitch-source-badge.manual {
            background: rgba(59, 130, 246, 0.2);
            color: #3b82f6;
            border: 1px solid rgba(59, 130, 246, 0.3);
        }

        .pitch-source-badge.auto {
            background: rgba(34, 197, 94, 0.2);
            color: #22c55e;
            border: 1px solid rgba(34, 197, 94, 0.3);
        }

        .pitch-source-badge.target {
            background: rgba(245, 158, 11, 0.2);
            color: #f59e0b;
            border: 1px solid rgba(245, 158, 11, 0.3);
        }

        .tolerance-badge {
            font-size: 10px;
            font-weight: 600;
            padding: 2px 8px;
            border-radius: 4px;
            text-transform: uppercase;
            letter-spacing: 0.05em;
        }

        .tolerance-badge.manual {
            background: rgba(59, 130, 246, 0.2);
            color: #3b82f6;
            border: 1px solid rgba(59, 130, 246, 0.3);
        }

        .metric-card .value {
            font-weight: 700;
            font-variant-numeric: tabular-nums;
            color: var(--text);
        }

        .metric-card.highlight .value {
            color: var(--accent);
        }

        /* Table Section */
        .table-section {
            position: relative;
        }

        .table-header {
            display: flex;
            justify-content: space-between;
            align-items: center;
            margin-bottom: 16px;
        }

        .table-section h3 {
            font-size: 14px;
            font-weight: 600;
            color: var(--text-muted);
            text-transform: uppercase;
            letter-spacing: 0.05em;
            margin: 0;
        }

        .table-wrapper {
            background: var(--surface);
            border: 1px solid var(--border);
            border-radius: var(--radius);
            overflow: hidden;
            box-shadow: var(--shadow);
        }

        .table-scroll {
            overflow-x: auto;
        }

        table {
            width: 100%;
            border-collapse: collapse;
            font-size: 13px;
            table-layout: fixed;
        }

        th {
            background: var(--surface-2);
            padding: 12px 6px;
            font-size: 10px;
            font-weight: 600;
            text-transform: uppercase;
            letter-spacing: 0.03em;
            color: var(--text-muted);
            text-align: center;
            border-bottom: 1px solid var(--border);
            white-space: normal;
            line-height: 1.3;
        }

        th:nth-child(1) { width: 7%; }  /* Workstation */
        th:nth-child(2) { width: 7%; }  /* Serial/Id */
        th:nth-child(3) { width: 16%; } /* Operations */
        th:nth-child(4) { width: 9%; }  /* Machine */
        th:nth-child(5) { width: 9%; }  /* Predecessor */
        th:nth-child(6) { width: 9%; }  /* Basic Time */
        th:nth-child(7) { width: 11%; } /* Combined Basic Time */
        th:nth-child(8) { width: 9%; }  /* Balancing SAM */
        th:nth-child(9) { width: 6%; }  /* M/P */
        th:nth-child(10) { width: 10%; } /* Pitch Time / Pitch Time (Auto) */
        th:nth-child(11) { width: 8%; } /* UCL */
        th:nth-child(12) { width: 8%; } /* LCL */
        th:nth-child(13) { width: 10%; } /* Status */

        td {
            padding: 12px 8px;
            border-bottom: 1px solid var(--border);
            color: var(--text);
            font-variant-numeric: tabular-nums;
            text-align: center;
            vertical-align: middle;
            word-wrap: break-word;
            overflow-wrap: break-word;
            white-space: normal;
        }

        /* Handle + values on new line */
        td .multiline {
            white-space: pre-wrap;
            line-height: 1.4;
        }

        .smart-break-cell {
            white-space: pre-wrap;
            line-height: 1.4;
            word-break: break-word;
        }

        tbody tr:hover {
            background: rgba(59, 130, 246, 0.06);
        }

        .status-badge {
            display: inline-block;
            padding: 6px 12px;
            border-radius: 20px;
            font-size: 11px;
            font-weight: 600;
            text-transform: uppercase;
            letter-spacing: 0.03em;
            min-width: 60px;
            white-space: nowrap;
        }

        .status-ok {
            background: rgba(34, 197, 94, 0.15);
            color: var(--success);
        }

        .status-ucl {
            background: rgba(239, 68, 68, 0.15);
            color: var(--danger);
        }

        .status-lcl {
            background: rgba(245, 158, 11, 0.15);
            color: var(--warning);
        }

        /* Export Buttons */
        .export-buttons {
            display: flex;
            gap: 8px;
        }

        .export-buttons button {
            background: var(--surface-2);
            color: var(--text);
            border: 1px solid var(--border);
            border-radius: var(--radius-sm);
            padding: 8px 16px;
            font-size: 12px;
            font-weight: 500;
            cursor: pointer;
            transition: all var(--transition);
            display: flex;
            align-items: center;
            gap: 6px;
        }

        .export-buttons button:hover {
            background: var(--accent);
            color: white;
            border-color: var(--accent);
        }

        .export-buttons button svg {
            width: 14px;
            height: 14px;
        }

        /* Results Section */
        .results-section {
            animation: fadeIn 0.4s ease;
        }

        @keyframes fadeIn {
            from { opacity: 0; transform: translateY(10px); }
            to { opacity: 1; transform: translateY(0); }
        }

        /* Chart Section */
        .chart-section {
            background: var(--surface);
            border: 1px solid var(--border);
            border-radius: var(--radius);
            padding: 24px;
            margin-bottom: 24px;
            box-shadow: var(--shadow);
        }

        .chart-header {
            display: flex;
            justify-content: space-between;
            align-items: center;
            margin-bottom: 20px;
            flex-wrap: wrap;
            gap: 16px;
        }

        .chart-title {
            font-size: 24px;
            font-weight: 600;
            color: var(--accent);
            display: flex;
            align-items: center;
            gap: 12px;
        }

        .chart-title svg {
            width: 28px;
            height: 28px;
        }

        .chart-button {
            background: var(--accent);
            color: white;
            border: none;
            padding: 10px 20px;
            border-radius: var(--radius-sm);
            font-size: 14px;
            font-weight: 500;
            cursor: pointer;
            transition: all var(--transition);
            display: flex;
            align-items: center;
            gap: 8px;
            text-decoration: none;
        }

        .chart-button:hover {
            background: var(--accent-hover);
            transform: translateY(-1px);
        }

        .chart-button svg {
            width: 16px;
            height: 16px;
        }

        .chart-buttons {
            display: flex;
            gap: 8px;
        }

        .chart-container {
            position: relative;
            height: 400px;
            width: 100%;
            background: var(--surface);
            border-radius: var(--radius-sm);
            padding: 16px;
        }

        .chart-container canvas {
            background: var(--surface);
            border-radius: var(--radius-sm);
        }

        /* Comparison Section Styles */
        .comparison-section {
            margin-bottom: 32px;
        }

        .comparison-table-wrapper {
            background: var(--surface);
            border: 1px solid var(--border);
            border-radius: var(--radius);
            overflow: hidden;
            box-shadow: var(--shadow);
        }

        .comparison-table {
            width: 100%;
            border-collapse: collapse;
            font-size: 14px;
        }

        .comparison-table thead {
            background: var(--surface-2);
        }

        .comparison-table th {
            padding: 16px;
            font-size: 12px;
            font-weight: 600;
            text-transform: uppercase;
            letter-spacing: 0.05em;
            color: var(--text-muted);
            text-align: center;
            border-bottom: 1px solid var(--border);
        }

        .comparison-table th.metric-name-header {
            width: 30%;
        }

        .comparison-table th.before-header {
            width: 35%;
            color: var(--warning);
            border-left: 1px solid var(--border);
        }

        .comparison-table th.after-header {
            width: 35%;
            color: var(--success);
            border-left: 1px solid var(--border);
        }

        .comparison-table td {
            padding: 16px;
            border-bottom: 1px solid var(--border);
            color: var(--text);
            font-variant-numeric: tabular-nums;
            vertical-align: middle;
        }

        .comparison-table tr:last-child td {
            border-bottom: none;
        }

        .comparison-table tr:hover {
            background: rgba(59, 130, 246, 0.04);
        }

        .comparison-table .metric-name {
            font-weight: 600;
            color: var(--text);
            font-size: 13px;
        }

        .comparison-table .before-cell {
            border-left: 1px solid var(--border);
        }

        .comparison-table .after-cell {
            border-left: 1px solid var(--border);
        }

        .comparison-table .metric-label {
            display: block;
            font-size: 11px;
            font-weight: 500;
            color: var(--text-muted);
            margin-bottom: 4px;
        }

        .comparison-table .metric-value {
            font-size: 18px;
            font-weight: 700;
            color: var(--text);
        }

        .comparison-table .unit {
            font-size: 12px;
            font-weight: 500;
            color: var(--text-muted);
            margin-left: 4px;
        }

        /* Charts Comparison Section */
        .charts-comparison-section {
            margin-bottom: 32px;
        }

        .charts-comparison-grid {
            display: grid;
            grid-template-columns: 1fr 1fr;
            gap: 20px;
        }

        @media (max-width: 768px) {
            .container {
                padding: 16px 12px;
                max-width: 100%;
            }

            .header {
                flex-direction: column;
                align-items: flex-start;
                gap: 16px;
            }

            .header h1 {
                font-size: 20px;
            }

            .header p {
                font-size: 12px;
            }

            .form-card {
                padding: 20px;
            }

            .form-grid {
                grid-template-columns: 1fr;
                gap: 16px;
            }

            .field {
                width: 100%;
            }

            .file-upload-field {
                grid-column: span 1 !important;
            }

            .metrics-grid {
                grid-template-columns: repeat(2, 1fr);
                gap: 12px;
            }

            .metric-card {
                padding: 16px;
            }

            .metric-card .value {
                font-size: 20px;
            }

            .table-section h3 {
                font-size: 12px;
            }

            .table-wrapper {
                border-radius: 8px;
            }

            .table-scroll {
                overflow-x: auto;
                -webkit-overflow-scrolling: touch;
            }

            table {
                font-size: 11px;
                min-width: 800px;
            }

            th {
                padding: 10px 4px;
                font-size: 9px;
                line-height: 1.2;
            }

            td {
                padding: 10px 4px;
                font-size: 10px;
            }

            .status-badge {
                padding: 4px 8px;
                font-size: 10px;
                min-width: 50px;
            }

            .export-buttons {
                flex-direction: column;
            }

            .export-buttons button {
                padding: 8px 12px;
                font-size: 11px;
            }

            .table-header {
                flex-direction: column;
                align-items: flex-start;
                gap: 12px;
            }

            button[type="submit"] {
                width: 100%;
            }

            .chart-header {
                flex-direction: column;
                align-items: flex-start;
            }

            .chart-container {
                height: 300px;
            }

            .chart-buttons {
                flex-direction: column;
                width: 100%;
            }

            .chart-button {
                width: 100%;
                justify-content: center;
            }

            .comparison-table {
                font-size: 12px;
            }

            .comparison-table th,
            .comparison-table td {
                padding: 12px 8px;
            }

            .comparison-table th.metric-name-header {
                width: 35%;
            }

            .comparison-table th.before-header,
            .comparison-table th.after-header {
                width: 32.5%;
            }

            .comparison-table .metric-value {
                font-size: 16px;
            }

            .comparison-table .metric-label {
                font-size: 10px;
            }

            .charts-comparison-grid {
                grid-template-columns: 1fr;
            }
        }

        @media (max-width: 480px) {
            .container {
                padding: 12px 8px;
            }

            .header h1 {
                font-size: 18px;
            }

            .metrics-grid {
                grid-template-columns: 1fr;
            }

            .form-card {
                padding: 16px;
            }

            th, td {
                padding: 8px 3px;
            }

            th {
                font-size: 8px;
                line-height: 1.1;
            }

            td {
                font-size: 9px;
            }

            .status-badge {
                padding: 3px 6px;
                font-size: 9px;
                min-width: 45px;
            }

            .export-buttons button {
                padding: 6px 10px;
                font-size: 10px;
            }

            .export-buttons button svg {
                width: 12px;
                height: 12px;
            }
        }
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <div class="header-content">
                <h1>Line Balancing Optimizer</h1>
                <p>Upload operation data and configure parameters to optimize workstation balance</p>
            </div>
            <div class="header-actions">
                <nav class="nav-tabs">
                    <a href="/" class="nav-tab active">Line Balancing</a>
                    <a href="/takt-vs-pitch" class="nav-tab">Takt vs Pitch Comparison</a>
                </nav>
                <button class="theme-toggle" onclick="toggleTheme()">🌙 Dark</button>
            </div>
        </div>

        <form method="post" enctype="multipart/form-data" class="form-card">
            <h2>Configuration</h2>
            <div class="form-grid">
                <div class="field file-upload-field">
                    <label>Upload Excel/CSV file</label>
                    <input type="file" name="file" accept=".csv,.xlsx,.xls" required>
                </div>
                <div class="field">
                    <label>Time Method</label>
                    <select name="pitch_time_method" id="pitch_time_method" onchange="togglePitchTimeInput()">
                        <option value="target_direct">Takt time (By target direct)</option>
                        <option value="target">Takt time (By target)</option>
                        <option value="manual">Takt time (Manual)</option>
                        <option value="auto">Pitch time (Auto)</option>
                    </select>
                </div>
                <div class="field" id="pitch_time_field">
                    <label>Time Value</label>
                    <input type="number" name="pitch_time" id="pitch_time_input" value="" placeholder="Time in seconds" min="0" step="0.1">
                </div>
                <div class="field" id="tolerance_field">
                    <label>Tolerance %</label>
                    <input type="number" name="tolerance" id="tolerance_input" value="15" min="0" max="100" step="1">
                </div>
                <div class="field" id="production_target_field">
                    <label>Customer Demand</label>
                    <input type="number" name="production_target" id="production_target_input" value="" placeholder="Number of units" min="0" step="1">
                </div>
                <div class="field" id="shift_time_field">
                    <label>Available Time</label>
                    <input type="number" name="shift_time" id="shift_time_input" value="420" placeholder="Shift duration in minutes" min="0" step="1">
                </div>
                <div class="field" id="efficiency_field">
                    <label>Required Efficiency (Optional)</label>
                    <input type="number" name="efficiency" id="efficiency_input" value="" placeholder="Required efficiency (0-100)" min="0" max="100" step="0.1">
                </div>
                <div class="field" id="available_time_field">
                    <label>Available Time (Optional)</label>
                    <input type="number" name="available_time" id="available_time_input" value="420" placeholder="Available time in minutes" min="0" step="1">
                </div>
                <div class="field">
                    <label>&nbsp;</label>
                    <button type="submit">Run calculations</button>
                </div>
            </div>
        </form>

        {% if error %}
        <div class="error-box">
            <strong>Error:</strong> {{ error }}
        </div>
        {% endif %}

        {% if result %}
        <section class="results-section">
            {% if result.pitch_time_source == "By Target" and result.target_validation_message %}
            <div class="target-workflow-notice-container" id="targetWorkflowNoticeContainer">
                <!-- Step 2 Validation Banner -->
                <div class="target-validation-banner {% if result.demand_met %}success{% else %}warning{% endif %}" id="targetValidationBanner">
                    {% if result.demand_met %}
                    <svg xmlns="http://www.w3.org/2000/svg" style="width: 20px; height: 20px; flex-shrink: 0;" fill="none" viewBox="0 0 24 24" stroke="currentColor" stroke-width="2">
                        <path stroke-linecap="round" stroke-linejoin="round" d="M9 12l2 2 4-4m6 2a9 9 0 11-18 0 9 9 0 0118 0z" />
                    </svg>
                    {% else %}
                    <svg xmlns="http://www.w3.org/2000/svg" style="width: 20px; height: 20px; flex-shrink: 0;" fill="none" viewBox="0 0 24 24" stroke="currentColor" stroke-width="2">
                        <path stroke-linecap="round" stroke-linejoin="round" d="M12 9v2m0 4h.01m-6.938 4h13.856c1.54 0 2.502-1.667 1.732-3L13.732 4c-.77-1.333-2.694-1.333-3.464 0L3.34 16c-.77 1.333.192 3 1.732 3z" />
                    </svg>
                    {% endif %}
                    <div>
                        <strong>Customer Demand Validation:</strong> <span id="targetValidationText">{{ result.target_validation_message }}</span>
                    </div>
                </div>
            </div>
            {% endif %}

            <div style="display: flex; justify-content: space-between; align-items: center; margin-bottom: 16px;">
                <h3 style="font-size: 14px; font-weight: 600; color: var(--text-muted); text-transform: uppercase; letter-spacing: 0.05em; margin: 0;">Calculation Results</h3>
            </div>
            <!-- Constant Metrics (shown immediately) -->
            <div class="metrics-grid">
                {% if result.production_target %}
                <div class="metric-card">
                    <div class="label">Customer Demand<br></div>
                    <div class="value">{{ result.production_target }}<span style="font-size: 12px; color: var(--text-muted);"> units</span></div>
                </div>
                {% endif %}
                {% if result.shift_time_minutes %}
                <div class="metric-card">
                    <div class="label">Available Time<br><br></div>
                    <div class="value">{{ "%.1f"|format(result.shift_time_minutes) }}<span style="font-size: 12px; color: var(--text-muted);"> min</span></div>
                </div>
                {% endif %}
                {% if result.efficiency_percentage and result.available_time_minutes %}
                <div class="metric-card">
                    <div class="label">Required Efficiency<br><br></div>
                    <div class="value">{{ "%.1f"|format(result.efficiency_percentage) }}<span style="font-size: 12px; color: var(--text-muted);">%</span></div>
                </div>
                {% endif %}
                <div class="metric-card">
                    <div class="label">SAM<br><br></div>
                    <div class="value">{{ "%.1f"|format(result.total_basic_time) }}<span style="font-size: 12px; color: var(--text-muted);"> min</span></div>
                </div>
                {% if result.pitch_time_source == "manual" or result.pitch_time_source == "By Target" or result.pitch_time_source == "By Target Direct" %}
                <div class="metric-card">
                    <div class="label">Takt Time<br><br></div>
                    <div class="value">
                        {{ "%.1f"|format(result.pitch_time) }}<span style="font-size: 12px; color: var(--text-muted);">sec</span>
                        {% if result.pitch_time_source == "manual" %}
                        <span class="pitch-source-badge manual">Manual</span>
                        {% elif result.pitch_time_source == "By Target" %}
                        <span class="pitch-source-badge target">By Target</span>
                        {% elif result.pitch_time_source == "By Target Direct" %}
                        <span class="pitch-source-badge target">By Target Direct</span>
                        {% endif %}
                    </div>
                </div>
                {% else %}
                <div class="metric-card balancing-constant-metric" style="display: none;">
                    <div class="label">Pitch Time<br><br></div>
                    <div class="value">
                        {{ "%.1f"|format(result.pitch_time) }}<span style="font-size: 12px; color: var(--text-muted);">sec</span>
                        <span class="pitch-source-badge auto">Auto</span>
                    </div>
                </div>
                {% endif %}
            {% if result.pitch_time_source == "By Target" and result.auto_pitch_time is defined %}
            <div class="metric-card balancing-constant-metric" style="display: none;">
                <div class="label">Pitch Time<br><br></div>
                <div class="value">{{ "%.1f"|format(result.auto_pitch_time) }}<span style="font-size: 12px; color: var(--text-muted);">sec</span>
                <span class="pitch-source-badge auto">Auto</span>
                </div>
            </div>
            {% endif %}
            {% if result.pitch_time_source == "calculated" or result.pitch_time_source == "By Target" %}
            <div class="metric-card balancing-constant-metric" style="display: none;">
                <div class="label">Tolerance<br><br></div>
                <div class="value">
                    {{ "%.1f"|format(result.tolerance * 100) }}<span style="font-size: 12px; color: var(--text-muted);">%</span>
                    {% if result.tolerance * 100 != 15.0 %}
                    <span class="tolerance-badge manual">Manual</span>
                    {% endif %}
                </div>
            </div>
            <div class="metric-card balancing-constant-metric" style="display: none;">
                <div class="label">UCL<br><br></div>
                <div class="value">{{ "%.1f"|format(result.ucl) }}<span style="font-size: 12px; color: var(--text-muted);">sec</span></div>
            </div>
            <div class="metric-card balancing-constant-metric" style="display: none;">
                <div class="label">LCL<br><br></div>
                <div class="value">{{ "%.1f"|format(result.lcl) }}<span style="font-size: 12px; color: var(--text-muted);">sec</span></div>
            </div>
            {% endif %}
            </div>

            <!-- Show Balancing Results Button -->
            <div style="margin-top: 24px; margin-bottom: 24px;">
                <button type="button" id="showBalancingResultsBtn" onclick="showBalancingResults()" class="chart-button" style="width: auto; padding: 12px 24px;">
                    <svg xmlns="http://www.w3.org/2000/svg" fill="none" viewBox="0 0 24 24" stroke="currentColor" stroke-width="2">
                        <path stroke-linecap="round" stroke-linejoin="round" d="M9 5l7 7-7 7" />
                    </svg>
                    Balancing Results
                </button>
            </div>

            <!-- Balancing Results (hidden by default, shown on button click) -->
            <div id="balancingResults" style="display: none;">
                <!-- Side-by-side comparison metrics table -->
                <div class="comparison-section">

            {% if result.pitch_time_source == "By Target" and result.target_validation_message %}
            <div class="target-workflow-notice-container" id="targetWorkflowNoticeContainer">   
                 <!-- Step 4 Recheck Summary Banner -->
                {% if result.target_recheck_summary %}
                <div class="target-recheck-banner" id="targetRecheckBanner">
                    <div style="display: flex; align-items: center; gap: 8px;">
                        <svg xmlns="http://www.w3.org/2000/svg" style="width: 18px; height: 18px; color: var(--accent); flex-shrink: 0;" fill="none" viewBox="0 0 24 24" stroke="currentColor" stroke-width="2">
                            <path stroke-linecap="round" stroke-linejoin="round" d="M4 4v5h.582m15.356 2A8.001 8.001 0 004.582 9m0 0H9m11 11v-5h-.581m0 0a8.003 8.003 0 01-15.357-2m15.357 2H15" />
                        </svg>
                        <span><strong>Balancing Status:</strong> <span id="targetRecheckSummaryText">{{ result.target_recheck_summary }}</span></span>
                    </div>
                    {% if result.target_recheck_messages and result.target_recheck_messages|length > 1 %}
                    <details style="cursor: pointer; font-size: 12px; color: var(--text-muted);">
                        <summary>View {{ result.target_recheck_messages|length }} attempts</summary>
                        <ul style="margin-top: 8px; padding-left: 20px; color: var(--text-muted);">
                            {% for msg in result.target_recheck_messages %}
                            <li>{{ msg }}</li>
                            {% endfor %}
                        </ul>
                    </details>
                    {% endif %}
                </div>
                {% endif %}                 
            </div>
            {% endif %}
                        <div style="display: flex; justify-content: space-between; align-items: center; margin-bottom: 16px;">
                    <h3 style="font-size: 14px; font-weight: 600; color: var(--text-muted); text-transform: uppercase; letter-spacing: 0.05em; margin-bottom: 16px; margin-top: 32px;">Before vs After Comparison</h3>
                <div class="export-buttons">
                    <button class="export-button" onclick="window.location.href='/layout/{{ session_id }}'">
                        <svg xmlns="http://www.w3.org/2000/svg" fill="none" viewBox="0 0 24 24" stroke="currentColor" stroke-width="2">
                            <path stroke-linecap="round" stroke-linejoin="round" d="M4 6h16M4 12h16M4 18h16" />
                        </svg>
                        View Layout
                    </button>
                    <button class="export-button" onclick="exportFile('xlsx', '{{ session_id }}')">
                        <svg xmlns="http://www.w3.org/2000/svg" fill="none" viewBox="0 0 24 24" stroke="currentColor" stroke-width="2">
                            <path stroke-linecap="round" stroke-linejoin="round" d="M4 16v1a3 3 0 003 3h10a3 3 0 003-3v-1m-4-4l-4 4m0 0l-4-4m4 4V4" />
                        </svg>
                        Export Report
                    </button>
                </div>
                </div>

                    <div class="comparison-table-wrapper">
                        <table class="comparison-table">
                            <thead>
                                <tr>
                                    <th class="metric-name-header">Metric</th>
                                    <th class="before-header">Before Balancing</th>
                                    <th class="after-header">After Balancing</th>
                                </tr>
                            </thead>
                            <tbody>
                                <tr>
                                    <td class="metric-name">Total / Composite Operations</td>
                                    <td class="before-cell">
                                        <span class="metric-label">Total Operations</span>
                                        <span class="metric-value">{{ result.before_metrics.num_operations }}</span>
                                    </td>
                                    <td class="after-cell">
                                        <span class="metric-label">Composite Operations</span>
                                        {% set after_operations = result.workstations|length %}
                                        {% set arrow = '↑' if after_operations > result.before_metrics.num_operations else ('↓' if after_operations < result.before_metrics.num_operations else '') %}
                                        <span class="metric-value">{{ after_operations }}{% if arrow %} <span style="font-weight: 900; font-size: 1.2em;">{{ arrow }}</span>{% endif %}</span>
                                    </td>
                                </tr>
                                <tr>
                                    <td class="metric-name">Total Manpower</td>
                                    <td class="before-cell">
                                        <span class="metric-value">{{ result.before_metrics.total_manpower }}</span>
                                    </td>
                                    <td class="after-cell">
                                        {% set after_manpower = result.workstations|map(attribute='manpower')|sum %}
                                        {% set arrow = '↑' if after_manpower > result.before_metrics.total_manpower else ('↓' if after_manpower < result.before_metrics.total_manpower else '') %}
                                        <span class="metric-value">{{ after_manpower }}{% if arrow %} <span style="font-weight: 900; font-size: 1.2em;">{{ arrow }}</span>{% endif %}</span>
                                    </td>
                                </tr>
                                {% if result.target_before is not none and result.target_after is not none %}
                                <tr>
                                    <td class="metric-name">Target</td>
                                    <td class="before-cell">
                                        <span class="metric-value">{{ "%.2f"|format(result.target_before) }}<span class="unit"> units</span></span>
                                    </td>
                                    <td class="after-cell">
                                        {% set arrow = '↑' if (result.target_after|round(2)) > (result.target_before|round(2)) else ('↓' if (result.target_after|round(2)) < (result.target_before|round(2)) else '') %}
                                        <span class="metric-value">{{ "%.2f"|format(result.target_after) }}{% if arrow %} <span style="font-weight: 900; font-size: 1.2em;">{{ arrow }}</span>{% endif %}<span class="unit"> units</span></span>
                                    </td>
                                </tr>
                                {% endif %}
                                {% if result.labour_productivity_before is not none or result.labour_productivity_after is not none %}
                                <tr>
                                    <td class="metric-name">Labour Productivity</td>
                                    <td class="before-cell">
                                        <span class="metric-value">{% if result.labour_productivity_before is not none %}{{ "%.0f"|format(result.labour_productivity_before) }}{% else %}N/A{% endif %}<span class="unit"> units/person</span></span>
                                    </td>
                                    <td class="after-cell">
                                        {% if result.labour_productivity_before is not none and result.labour_productivity_after is not none %}
                                            {% set arrow = '↑' if (result.labour_productivity_after|round(0)) > (result.labour_productivity_before|round(0)) else ('↓' if (result.labour_productivity_after|round(0)) < (result.labour_productivity_before|round(0)) else '') %}
                                            <span class="metric-value">{{ "%.0f"|format(result.labour_productivity_after) }}<span class="unit"> units/person</span>{% if arrow %} <span style="font-weight: 900; font-size: 1.2em;">{{ arrow }}</span>{% endif %}</span>
                                        {% else %}
                                            <span class="metric-value">{% if result.labour_productivity_after is not none %}{{ "%.0f"|format(result.labour_productivity_after) }}{% else %}N/A{% endif %}<span class="unit"> units/person</span></span>
                                        {% endif %}
                                    </td>
                                </tr>
                                {% endif %}
                                {% if result.before_metrics.line_efficiency or result.line_efficiency %}
                                <tr>
                                    <td class="metric-name">Required Line Efficiency</td>
                                    <td class="before-cell">
                                        <span class="metric-value">{% if result.before_metrics.line_efficiency %}{{ "%.1f"|format(result.before_metrics.line_efficiency) }}{% else %}N/A{% endif %}<span class="unit">%</span></span>
                                    </td>
                                    <td class="after-cell">
                                        {% if result.before_metrics.line_efficiency and result.line_efficiency %}
                                            {% set arrow = '↑' if (result.line_efficiency|round(1)) > (result.before_metrics.line_efficiency|round(1)) else ('↓' if (result.line_efficiency|round(1)) < (result.before_metrics.line_efficiency|round(1)) else '') %}
                                            <span class="metric-value">{{ "%.1f"|format(result.line_efficiency) }}<span class="unit">%</span>{% if arrow %} <span style="font-weight: 900; font-size: 1.2em;">{{ arrow }}</span>{% endif %}</span>
                                        {% else %}
                                            <span class="metric-value">{% if result.line_efficiency %}{{ "%.1f"|format(result.line_efficiency) }}{% else %}N/A{% endif %}<span class="unit">%</span></span>
                                        {% endif %}
                                    </td>
                                </tr>
                                {% endif %}
                                {% if (result.pitch_time_source == "By Target" or result.pitch_time_source == "By Target Direct") and result.throughput_rate is not none and result.before_metrics.throughput_rate is not none %}
                                <tr>
                                    <td class="metric-name">Throughput Rate</td>
                                    <td class="before-cell">
                                        <span class="metric-value">{{ "%.1f"|format(result.before_metrics.throughput_rate) }}<span class="unit"> s</span></span>
                                    </td>
                                    <td class="after-cell">
                                        {% set arrow = '↑' if (result.throughput_rate|round(1)) > (result.before_metrics.throughput_rate|round(1)) else ('↓' if (result.throughput_rate|round(1)) < (result.before_metrics.throughput_rate|round(1)) else '') %}
                                        <span class="metric-value">{{ "%.1f"|format(result.throughput_rate) }}<span class="unit"> s</span>{% if arrow %} <span style="font-weight: 900; font-size: 1.2em;">{{ arrow }}</span>{% endif %}</span>
                                    </td>
                                </tr>
                                {% endif %}
                                {% if (result.pitch_time_source == "By Target" or result.pitch_time_source == "By Target Direct") and result.required_minutes is not none and result.before_metrics.required_minutes is not none %}
                                <tr>
                                    <td class="metric-name">Required Minutes</td>
                                    <td class="before-cell">
                                        <span class="metric-value">{{ "%.1f"|format(result.before_metrics.required_minutes) }}<span class="unit"> min</span></span>
                                    </td>
                                    <td class="after-cell">
                                        {% set arrow = '↑' if (result.required_minutes|round(1)) > (result.before_metrics.required_minutes|round(1)) else ('↓' if (result.required_minutes|round(1)) < (result.before_metrics.required_minutes|round(1)) else '') %}
                                        <span class="metric-value">{{ "%.1f"|format(result.required_minutes) }}<span class="unit"> min</span>{% if arrow %} <span style="font-weight: 900; font-size: 1.2em;">{{ arrow }}</span>{% endif %}</span>
                                    </td>
                                </tr>
                                {% endif %}
                                <tr>
                                    <td class="metric-name">Balancing Rate</td>
                                    <td class="before-cell">
                                        <span class="metric-value">{{ "%.1f"|format(result.before_metrics.balancing_rate) }}<span class="unit">%</span></span>
                                    </td>
                                    <td class="after-cell">
                                        {% set arrow = '↑' if (result.line_balancing_rate|round(1)) > (result.before_metrics.balancing_rate|round(1)) else ('↓' if (result.line_balancing_rate|round(1)) < (result.before_metrics.balancing_rate|round(1)) else '') %}
                                        <span class="metric-value">{{ "%.1f"|format(result.line_balancing_rate) }}<span class="unit">%</span>{% if arrow %} <span style="font-weight: 900; font-size: 1.2em;">{{ arrow }}</span>{% endif %}</span>
                                    </td>
                                </tr>
                                <tr>
                                    <td class="metric-name">Balance Delay</td>
                                    <td class="before-cell">
                                        <span class="metric-value">{{ "%.1f"|format(result.before_metrics.balance_delay) }}<span class="unit">%</span></span>
                                    </td>
                                    <td class="after-cell">
                                        {% set arrow = '↑' if (result.balance_delay|round(1)) > (result.before_metrics.balance_delay|round(1)) else ('↓' if (result.balance_delay|round(1)) < (result.before_metrics.balance_delay|round(1)) else '') %}
                                        <span class="metric-value">{{ "%.1f"|format(result.balance_delay) }}<span class="unit">%</span>{% if arrow %} <span style="font-weight: 900; font-size: 1.2em;">{{ arrow }}</span>{% endif %}</span>
                                    </td>
                                </tr>
                                <tr>
                                    <td class="metric-name">Smoothing Index</td>
                                    <td class="before-cell">
                                        <span class="metric-value">{{ "%.2f"|format(result.before_metrics.smoothing_index) }}<span class="unit"> min</span></span>
                                    </td>
                                    <td class="after-cell">
                                        {% set arrow = '↑' if (result.smoothing_index|round(2)) > (result.before_metrics.smoothing_index|round(2)) else ('↓' if (result.smoothing_index|round(2)) < (result.before_metrics.smoothing_index|round(2)) else '') %}
                                        <span class="metric-value">{{ "%.2f"|format(result.smoothing_index) }}<span class="unit"> min</span>{% if arrow %} <span style="font-weight: 900; font-size: 1.2em;">{{ arrow }}</span>{% endif %}</span>
                                    </td>
                                </tr>
                            </tbody>
                        </table>
                    </div>
                </div>

                <!-- Side-by-side charts -->
                <div class="charts-comparison-section">
                    <h3 style="font-size: 14px; font-weight: 600; color: var(--text-muted); text-transform: uppercase; letter-spacing: 0.05em; margin-bottom: 16px; margin-top: 32px;">Before vs After Charts</h3>
                    
                    <div class="charts-comparison-grid">
                        <div class="chart-section">
                            <div class="chart-header">
                                <div class="chart-title">
                                    <svg xmlns="http://www.w3.org/2000/svg" fill="none" viewBox="0 0 24 24" stroke="currentColor">
                                        <path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M9 19v-6a2 2 0 00-2-2H5a2 2 0 00-2 2v6a2 2 0 002 2h2a2 2 0 002-2zm0 0V9a2 2 0 012-2h2a2 2 0 012 2v10m-6 0a2 2 0 002 2h2a2 2 0 002-2m0 0V5a2 2 0 012-2h2a2 2 0 012 2v14a2 2 0 01-2 2h-2a2 2 0 01-2-2z" />
                                    </svg>
                                    Before Balancing
                                </div>
                            </div>
                            <div class="chart-container">
                                <canvas id="beforeBalanceChart"></canvas>
                            </div>
                        </div>

                        <div class="chart-section">
                            <div class="chart-header">
                                <div class="chart-title">
                                    <svg xmlns="http://www.w3.org/2000/svg" fill="none" viewBox="0 0 24 24" stroke="currentColor">
                                        <path stroke-linecap="round" stroke-linejoin="round" stroke-width="2" d="M9 19v-6a2 2 0 00-2-2H5a2 2 0 00-2 2v6a2 2 0 002 2h2a2 2 0 002-2zm0 0V9a2 2 0 012-2h2a2 2 0 012 2v10m-6 0a2 2 0 002 2h2a2 2 0 002-2m0 0V5a2 2 0 012-2h2a2 2 0 012 2v14a2 2 0 01-2 2h-2a2 2 0 01-2-2z" />
                                    </svg>
                                    After Balancing
                                </div>
                            </div>
                            <div class="chart-container">
                                <canvas id="balanceChart"></canvas>
                            </div>
                        </div>
                    </div>
                </div>
            </div>
        </section>
        {% endif %}
    </div>

    <script>
        // Theme Toggle
        function toggleTheme() {
            const html = document.documentElement;
            const currentTheme = html.getAttribute('data-theme') || 'dark';
            const newTheme = currentTheme === 'dark' ? 'light' : 'dark';
            html.setAttribute('data-theme', newTheme);
            localStorage.setItem('theme', newTheme);
            // Show what you're switching TO (opposite of current/new theme)
            // Update all theme toggle buttons on the page
            document.querySelectorAll('.theme-toggle').forEach(button => {
                button.textContent = newTheme === 'dark' ? '☀️ Light' : '🌙 Dark';
            });
        }

        // Toggle pitch time input field based on method selection
        function togglePitchTimeInput() {
            const method = document.getElementById('pitch_time_method').value;
            const pitchTimeField = document.getElementById('pitch_time_field');
            const pitchTimeInput = document.getElementById('pitch_time_input');
            const productionTargetField = document.getElementById('production_target_field');
            const shiftTimeField = document.getElementById('shift_time_field');
            const shiftTimeInput = document.getElementById('shift_time_input');
            const efficiencyField = document.getElementById('efficiency_field');
            const efficiencyInput = document.getElementById('efficiency_input');
            const availableTimeField = document.getElementById('available_time_field');
            const availableTimeInput = document.getElementById('available_time_input');
            const toleranceField = document.getElementById('tolerance_field');
            const toleranceInput = document.getElementById('tolerance_input');
            
            if (method === 'manual') {
                pitchTimeField.style.display = 'flex';
                pitchTimeInput.required = true;
                productionTargetField.style.display = 'none';
                shiftTimeField.style.display = 'none';
                shiftTimeInput.value = '';
                // Hide tolerance field for manual method
                toleranceField.style.display = 'none';
                toleranceInput.required = false;
                toleranceInput.value = '';
                // Show efficiency and available time fields for manual method
                efficiencyField.style.display = 'flex';
                efficiencyInput.required = false;
                availableTimeField.style.display = 'flex';
                availableTimeInput.required = false;
                // Set default available time only if field is empty
                if (!availableTimeInput.value) {
                    availableTimeInput.value = '420';
                }
            } else if (method === 'target' || method === 'target_direct') {
                pitchTimeField.style.display = 'none';
                pitchTimeInput.required = false;
                pitchTimeInput.value = '';
                productionTargetField.style.display = 'flex';
                shiftTimeField.style.display = 'flex';
                // Set default shift time only if field is empty
                if (!shiftTimeInput.value) {
                    shiftTimeInput.value = '420';
                }
                // Hide tolerance field for target/target_direct methods
                toleranceField.style.display = 'none';
                toleranceInput.required = false;
                toleranceInput.value = '';
                // Hide efficiency and available time fields for target/target_direct methods
                efficiencyField.style.display = 'none';
                efficiencyInput.required = false;
                efficiencyInput.value = '';
                availableTimeField.style.display = 'none';
                availableTimeInput.required = false;
                availableTimeInput.value = '';
            } else { // auto
                pitchTimeField.style.display = 'none';
                pitchTimeInput.required = false;
                pitchTimeInput.value = '';
                productionTargetField.style.display = 'none';
                shiftTimeField.style.display = 'none';
                shiftTimeInput.value = '';
                // Show tolerance field for auto method
                toleranceField.style.display = 'flex';
                toleranceInput.required = false;
                // Set default tolerance only if field is empty
                if (!toleranceInput.value) {
                    toleranceInput.value = '15';
                }
                // Show efficiency and available time fields for auto method
                efficiencyField.style.display = 'flex';
                efficiencyInput.required = false;
                availableTimeField.style.display = 'flex';
                availableTimeInput.required = false;
                // Set default available time only if field is empty
                if (!availableTimeInput.value) {
                    availableTimeInput.value = '420';
                }
            }
        }

        // Function to reset chart state when new calculations are run
        function resetChartState() {
            // Destroy existing chart instances
            if (window.beforeBalanceChartInstance) {
                try {
                    window.beforeBalanceChartInstance.destroy();
                } catch (e) {
                    console.log('Error destroying before chart:', e);
                }
                window.beforeBalanceChartInstance = null;
            }
            if (window.balanceChartInstance) {
                try {
                    window.balanceChartInstance.destroy();
                } catch (e) {
                    console.log('Error destroying after chart:', e);
                }
                window.balanceChartInstance = null;
            }
            // Reset chart data and loaded state
            window.beforeChartData = null;
            window.chartsLoaded = false;
            
            // Hide balancing results section if it's visible
            const resultsSection = document.getElementById('balancingResults');
            if (resultsSection) {
                resultsSection.style.display = 'none';
            }
            
            // Hide constant metric cards
            document.querySelectorAll('.balancing-constant-metric').forEach(el => {
                el.style.display = 'none';
            });
            
            // Reset button to show state
            const button = document.getElementById('showBalancingResultsBtn');
            if (button) {
                button.innerHTML = `
                    <svg xmlns="http://www.w3.org/2000/svg" fill="none" viewBox="0 0 24 24" stroke="currentColor" stroke-width="2">
                        <path stroke-linecap="round" stroke-linejoin="round" d="M9 5l7 7-7 7" />
                    </svg>
                    Balancing Results
                `;
                button.onclick = showBalancingResults;
            }
        }

        // Restore theme on load
        window.addEventListener('DOMContentLoaded', function() {
            const savedTheme = localStorage.getItem('theme') || 'dark';
            document.documentElement.setAttribute('data-theme', savedTheme);
            // Show what you're switching TO (opposite of current theme)
            // Update all theme toggle buttons on the page
            document.querySelectorAll('.theme-toggle').forEach(button => {
                button.textContent = savedTheme === 'dark' ? '☀️ Light' : '🌙 Dark';
            });

            // Initialize pitch time field visibility
            togglePitchTimeInput();

            // Initialize balancing results section state
            // Only load charts if session_id is available and results are already shown
            {% if session_id %}
            // Store session ID for later use
            const newSessionId = '{{ session_id }}';
            if (window.currentSessionId && window.currentSessionId !== newSessionId) {
                // New session detected, reset chart state
                resetChartState();
            }
            window.currentSessionId = newSessionId;
            window.chartsLoaded = false; // Track if charts have been loaded
            {% else %}
            // Reset charts loaded state when no session
            window.chartsLoaded = false;
            window.currentSessionId = null;
            // Reset chart state when page loads without results
            resetChartState();
            {% endif %}
        });

        // Function to show balancing results
        function showBalancingResults() {
            const resultsSection = document.getElementById('balancingResults');
            const button = document.getElementById('showBalancingResultsBtn');
            
            if (resultsSection) {
                resultsSection.style.display = 'block';
                
                // Show constant metric cards
                document.querySelectorAll('.balancing-constant-metric').forEach(el => {
                    el.style.display = '';
                });

                // Update button to hide results
                button.innerHTML = `
                    <svg xmlns="http://www.w3.org/2000/svg" fill="none" viewBox="0 0 24 24" stroke="currentColor" stroke-width="2">
                        <path stroke-linecap="round" stroke-linejoin="round" d="M19 9l-7 7-7-7" />
                    </svg>
                    Hide Balancing Results
                `;
                button.onclick = hideBalancingResults;
                
                // Load charts only when section is revealed and if not already loaded
                if (window.currentSessionId && !window.chartsLoaded) {
                    loadBeforeChart(window.currentSessionId);
                }
            }
        }

        // Function to hide balancing results
        function hideBalancingResults() {
            const resultsSection = document.getElementById('balancingResults');
            const button = document.getElementById('showBalancingResultsBtn');
            
            if (resultsSection) {
                resultsSection.style.display = 'none';
                
                // Hide constant metric cards
                document.querySelectorAll('.balancing-constant-metric').forEach(el => {
                    el.style.display = 'none';
                });

                // Update button to show results
                button.innerHTML = `
                    <svg xmlns="http://www.w3.org/2000/svg" fill="none" viewBox="0 0 24 24" stroke="currentColor" stroke-width="2">
                        <path stroke-linecap="round" stroke-linejoin="round" d="M9 5l7 7-7 7" />
                    </svg>
                    Balancing Results
                `;
                button.onclick = showBalancingResults;
                
                // Don't destroy chart instances - they will be reused when section is shown again
                // This prevents the "Canvas is already in use" error
            }
        }

        // Export function
        function exportFile(format, sessionId) {
            window.location.href = `/api/export/${format}/${sessionId}`;
        }

        async function loadBeforeChart(sessionId) {
            try {
                const response = await fetch(`/api/before-chart-data/${sessionId}`);
                if (!response.ok) {
                    const errorData = await response.json();
                    if (response.status === 404) {
                        // Session expired, redirect to home with message
                        alert(errorData.message || 'Session expired. Please reload the data.');
                        window.location.href = '/';
                        return;
                    }
                    throw new Error(errorData.error || 'Failed to load before chart data');
                }
                
                const data = await response.json();
                
                // Store before chart data for shared Y-axis calculation
                window.beforeChartData = data;
                
                // Load after chart data to calculate shared Y-axis
                await loadChart(sessionId);
            } catch (error) {
                console.error('Error loading before chart:', error);
                alert('Error loading before chart data: ' + error.message);
                window.location.href = '/';
            }
        }

        async function loadChart(sessionId) {
            try {
                const response = await fetch(`/api/chart-data/${sessionId}`);
                if (!response.ok) {
                    const errorData = await response.json();
                    if (response.status === 404) {
                        // Session expired, redirect to home with message
                        alert(errorData.message || 'Session expired. Please reload the data.');
                        window.location.href = '/';
                        return;
                    }
                    throw new Error(errorData.error || 'Failed to load chart data');
                }
                
                const data = await response.json();
                
                // Calculate shared Y-axis maximum
                let sharedYMax = 0;
                
                // Find max from before chart data
                if (window.beforeChartData && window.beforeChartData.basic_times) {
                    const beforeMax = Math.max(...window.beforeChartData.basic_times);
                    sharedYMax = Math.max(sharedYMax, beforeMax);
                }
                
                // Find max from after chart data
                if (data.balancing_sam) {
                    const afterMax = Math.max(...data.balancing_sam);
                    sharedYMax = Math.max(sharedYMax, afterMax);
                }
                
                // Add some padding (10%) to the max value
                sharedYMax = sharedYMax * 1.1;
                
                // Create both charts with shared Y-axis
                if (window.beforeChartData) {
                    createBeforeChart(window.beforeChartData, sharedYMax);
                }
                createChart(data, sharedYMax);
                
                // Mark charts as loaded
                window.chartsLoaded = true;
            } catch (error) {
                console.error('Error loading chart:', error);
                alert('Error loading chart data: ' + error.message);
                window.location.href = '/';
            }
        }

        function createBeforeChart(chartData, sharedYMax) {
            const ctx = document.getElementById('beforeBalanceChart').getContext('2d');
            
            // Destroy existing chart instance if it exists
            if (window.beforeBalanceChartInstance) {
                window.beforeBalanceChartInstance.destroy();
                window.beforeBalanceChartInstance = null;
            }
            
            // Get theme colors - grey text for light mode, white for dark mode
            const isDark = document.documentElement.getAttribute('data-theme') === 'dark';
            const textColor = isDark ? '#ffffff' : '#64748b'; // White in dark, grey in light
            const gridColor = isDark ? 'rgba(255, 255, 255, 0.1)' : 'rgba(0, 0, 0, 0.1)';
            
            // Determine pitch time label based on source
            let pitchTimeLabel, pitchTimeValue;
            if (chartData.pitch_time_source === "manual" || chartData.pitch_time_source === "By Target Direct") {
                pitchTimeLabel = 'Takt Time';
                pitchTimeValue = chartData.pitch_time;
            } else if (chartData.pitch_time_source === "By Target") {
                pitchTimeLabel = 'Pitch Time (Auto)';
                pitchTimeValue = chartData.auto_pitch_time || chartData.pitch_time;
            } else {
                pitchTimeLabel = 'Pitch Time';
                pitchTimeValue = chartData.pitch_time;
            }
            
            // Build datasets array - always include basic time
            const datasets = [
                {
                    label: 'Basic Time (SAM)',
                    data: chartData.basic_times,
                    backgroundColor: 'rgba(59, 130, 246, 0.8)',
                    borderColor: 'rgba(59, 130, 246, 1)',
                    borderWidth: 1,
                    borderRadius: 4
                }
            ];
            
            // Add pitch time line for manual/By Target Direct method, or auto pitch time for auto/By Target methods
            if (chartData.pitch_time_source === "manual" || chartData.pitch_time_source === "By Target Direct") {
                datasets.push({
                    label: `${pitchTimeLabel} ${pitchTimeValue.toFixed(1)}s`,
                    data: Array(chartData.operations.length).fill(pitchTimeValue),
                    borderColor: 'rgb(34, 197, 94)',
                    backgroundColor: 'rgb(34, 197, 94)',
                    borderWidth: 2,
                    borderDash: [5, 5],
                    pointRadius: 0,
                    type: 'line',
                    fill: false,
                    hidden: false
                });
            } else if (chartData.pitch_time_source === "calculated" || chartData.pitch_time_source === "By Target") {
                // For auto and By Target methods, add pitch time, UCL, and LCL lines
                datasets.push(
                    {
                        label: `${pitchTimeLabel} ${pitchTimeValue.toFixed(1)}s`,
                        data: Array(chartData.operations.length).fill(pitchTimeValue),
                        borderColor: 'rgb(34, 197, 94)',
                        backgroundColor: 'rgb(34, 197, 94)',
                        borderWidth: 2,
                        borderDash: [5, 5],
                        pointRadius: 0,
                        type: 'line',
                        fill: false,
                        hidden: false
                    },
                    {
                        label: `UCL ${chartData.ucl.toFixed(1)}s`,
                        data: Array(chartData.operations.length).fill(chartData.ucl),
                        borderColor: 'rgb(239, 68, 68)',
                        backgroundColor: 'rgb(239, 68, 68)',
                        borderWidth: 2,
                        borderDash: [5, 5],
                        pointRadius: 0,
                        type: 'line',
                        fill: false,
                        hidden: false
                    },
                    {
                        label: `LCL ${chartData.lcl.toFixed(1)}s`,
                        data: Array(chartData.operations.length).fill(chartData.lcl),
                        borderColor: 'rgb(249, 115, 22)',
                        backgroundColor: 'rgb(249, 115, 22)',
                        borderWidth: 2,
                        borderDash: [5, 5],
                        pointRadius: 0,
                        type: 'line',
                        fill: false,
                        hidden: false
                    }
                );
            }
            
            window.beforeBalanceChartInstance = new Chart(ctx, {
                type: 'bar',
                data: {
                    labels: chartData.operations,
                    datasets: datasets
                },
                options: {
                    responsive: true,
                    maintainAspectRatio: false,
                    plugins: {
                        legend: {
                            display: true,
                            position: 'top',
                            labels: {
                                color: textColor,
                                font: {
                                    size: 13,
                                    weight: 500
                                }
                            }
                        },
                        tooltip: {
                            backgroundColor: '#1a2332', // Always dark background for tooltip
                            titleColor: '#ffffff', // Always white in tooltip
                            bodyColor: '#ffffff', // Always white in tooltip
                            borderColor: gridColor,
                            borderWidth: 1,
                            padding: 12,
                            displayColors: true,
                            titleFont: {
                                size: 11,
                                weight: 'bold'
                            },
                            bodyFont: {
                                size: 12
                            },
                            callbacks: {
                                title: function(context) {
                                    // Show full operation name in tooltip
                                    return context[0].label;
                                },
                                label: function(context) {
                                    if (context.datasetIndex === 0) {
                                        return `Basic Time (SAM): ${context.raw.toFixed(1)}s`;
                                    }
                                    return null; // Don't show tooltips for reference lines
                                }
                            }
                        }
                    },
                    scales: {
                        x: {
                            grid: {
                                color: gridColor
                            },
                            ticks: {
                                color: textColor,
                                font: {
                                    size: 9
                                },
                                // Show all labels without skipping
                                autoSkip: false,
                                maxRotation: 45,
                                minRotation: 45,
                                callback: function(value, index, values) {
                                    const label = this.getLabelForValue(value);
                                    // Truncate very long labels for display
                                    if (label.length > 25) {
                                        return label.substring(0, 25) + '...';
                                    }
                                    return label;
                                }
                            }
                        },
                        y: {
                            beginAtZero: true,
                            max: sharedYMax, // Use shared Y-axis maximum
                            grid: {
                                color: gridColor
                            },
                            ticks: {
                                color: textColor,
                                font: {
                                    size: 12
                                },
                                callback: function(value) {
                                    return value.toFixed(1) + 's';
                                }
                            },
                            title: {
                                display: true,
                                text: 'Time (seconds)',
                                color: textColor,
                                font: {
                                    size: 13,
                                    weight: 500
                                }
                            }
                        }
                    }
                }
            });
        }

        function createChart(chartData, sharedYMax) {
            const ctx = document.getElementById('balanceChart').getContext('2d');
            
            // Destroy existing chart instance if it exists
            if (window.balanceChartInstance) {
                window.balanceChartInstance.destroy();
                window.balanceChartInstance = null;
            }
            
            // Get theme colors - grey text for light mode, white for dark mode
            const isDark = document.documentElement.getAttribute('data-theme') === 'dark';
            const textColor = isDark ? '#ffffff' : '#333333'; // White in dark, dark grey in light
            const gridColor = isDark ? 'rgba(255, 255, 255, 0.1)' : 'rgba(0, 0, 0, 0.1)';
            
            // Determine pitch time label based on source
            let pitchTimeLabel, pitchTimeValue;
            if (chartData.pitch_time_source === "manual" || chartData.pitch_time_source === "By Target Direct") {
                pitchTimeLabel = 'Takt Time';
                pitchTimeValue = chartData.pitch_time;
            } else if (chartData.pitch_time_source === "By Target") {
                pitchTimeLabel = 'Pitch Time (Auto)';
                pitchTimeValue = chartData.auto_pitch_time || chartData.pitch_time;
            } else {
                pitchTimeLabel = 'Pitch Time';
                pitchTimeValue = chartData.pitch_time;
            }
            
            // Build datasets array - always include balancing SAM
            const datasets = [
                {
                    label: 'Balancing SAM',
                    data: chartData.balancing_sam,
                    backgroundColor: 'rgba(59, 130, 246, 0.8)',
                    borderColor: 'rgba(59, 130, 246, 1)',
                    borderWidth: 1,
                    borderRadius: 4
                }
            ];
            
            // Add pitch time line for manual/By Target Direct method, or auto pitch time for auto/By Target methods
            if (chartData.pitch_time_source === "manual" || chartData.pitch_time_source === "By Target Direct") {
                datasets.push({
                    label: `${pitchTimeLabel} ${pitchTimeValue.toFixed(1)}s`,
                    data: Array(chartData.workstations.length).fill(pitchTimeValue),
                    borderColor: 'rgb(34, 197, 94)',
                    backgroundColor: 'rgb(34, 197, 94)',
                    borderWidth: 2,
                    borderDash: [5, 5],
                    pointRadius: 0,
                    type: 'line',
                    fill: false,
                    hidden: false
                });
            } else if (chartData.pitch_time_source === "calculated" || chartData.pitch_time_source === "By Target") {
                // For auto and By Target methods, add pitch time, UCL, and LCL lines
                datasets.push(
                    {
                        label: `${pitchTimeLabel} ${pitchTimeValue.toFixed(1)}s`,
                        data: Array(chartData.workstations.length).fill(pitchTimeValue),
                        borderColor: 'rgb(34, 197, 94)',
                        backgroundColor: 'rgb(34, 197, 94)',
                        borderWidth: 2,
                        borderDash: [5, 5],
                        pointRadius: 0,
                        type: 'line',
                        fill: false,
                        hidden: false
                    },
                    {
                        label: `UCL ${chartData.ucl.toFixed(1)}s`,
                        data: Array(chartData.workstations.length).fill(chartData.ucl),
                        borderColor: 'rgb(239, 68, 68)',
                        backgroundColor: 'rgb(239, 68, 68)',
                        borderWidth: 2,
                        borderDash: [5, 5],
                        pointRadius: 0,
                        type: 'line',
                        fill: false,
                        hidden: false
                    },
                    {
                        label: `LCL ${chartData.lcl.toFixed(1)}s`,
                        data: Array(chartData.workstations.length).fill(chartData.lcl),
                        borderColor: 'rgb(249, 115, 22)',
                        backgroundColor: 'rgb(249, 115, 22)',
                        borderWidth: 2,
                        borderDash: [5, 5],
                        pointRadius: 0,
                        type: 'line',
                        fill: false,
                        hidden: false
                    }
                );
            }
            
            window.balanceChartInstance = new Chart(ctx, {
                type: 'bar',
                data: {
                    labels: chartData.workstations,
                    datasets: datasets
                },
                options: {
                    responsive: true,
                    maintainAspectRatio: false,
                    plugins: {
                        legend: {
                            display: true,
                            position: 'top',
                            labels: {
                                color: textColor,
                                font: {
                                    size: 13,
                                    weight: 500
                                }
                            }
                        },
                        tooltip: {
                            backgroundColor: '#1a2332', // Always dark background for tooltip
                            titleColor: '#ffffff', // Always white in tooltip
                            bodyColor: '#ffffff', // Always white in tooltip
                            borderColor: gridColor,
                            borderWidth: 1,
                            padding: 12,
                            displayColors: true,
                            titleFont: {
                                size: 11,
                                weight: 'bold'
                            },
                            bodyFont: {
                                size: 12
                            },
                            callbacks: {
                                title: function(context) {
                                    // Show full operation name in tooltip
                                    return context[0].label;
                                },
                                label: function(context) {
                                    if (context.datasetIndex === 0) {
                                        return `Balancing SAM: ${context.raw.toFixed(1)}s`;
                                    }
                                    return null; // Don't show tooltips for reference lines
                                }
                            }
                        }
                    },
                    scales: {
                        x: {
                            grid: {
                                color: gridColor
                            },
                            ticks: {
                                color: textColor,
                                font: {
                                    size: 9
                                },
                                // Show all labels without skipping
                                autoSkip: false,
                                maxRotation: 45,
                                minRotation: 45,
                                callback: function(value, index, values) {
                                    const label = this.getLabelForValue(value);
                                    // Truncate very long labels for display
                                    if (label.length > 25) {
                                        return label.substring(0, 25) + '...';
                                    }
                                    return label;
                                }
                            }
                        },
                        y: {
                            beginAtZero: true,
                            max: sharedYMax, // Use shared Y-axis maximum
                            grid: {
                                color: gridColor
                            },
                            ticks: {
                                color: textColor,
                                font: {
                                    size: 12
                                },
                                callback: function(value) {
                                    return value.toFixed(1) + 's';
                                }
                            },
                            title: {
                                display: true,
                                text: 'Time (seconds)',
                                color: textColor,
                                font: {
                                    size: 13,
                                    weight: 500
                                }
                            }
                        }
                    }
                }
            });
        }
    </script>
</body>
</html>
"""

LAYOUT_TEMPLATE = """
<!DOCTYPE html>
<html lang="en" data-theme="dark">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Layout View</title>
    <style>
        :root {
            --bg: #0f1419;
            --surface: #1a2332;
            --surface-2: #243044;
            --border: rgba(255, 255, 255, 0.08);
            --text: #e8edf4;
            --text-muted: #8b9cb3;
            --accent: #3b82f6;
            --accent-hover: #2563eb;
            --success: #22c55e;
            --warning: #f59e0b;
            --danger: #ef4444;
            --shadow: 0 4px 24px rgba(0, 0, 0, 0.35);
            --radius: 12px;
            --radius-sm: 8px;
            --transition: 0.2s ease;
        }

        [data-theme="light"] {
            --bg: #f1f5f9;
            --surface: #ffffff;
            --surface-2: #f8fafc;
            --border: rgba(15, 23, 42, 0.1);
            --text: #0f172a;
            --text-muted: #64748b;
            --accent: #2563eb;
            --accent-hover: #1d4ed8;
            --success: #16a34a;
            --warning: #d97706;
            --danger: #dc2626;
            --shadow: 0 4px 24px rgba(15, 23, 42, 0.08);
        }

        * { 
            margin: 0; 
            padding: 0; 
            box-sizing: border-box; 
        }

        body {
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
            background: var(--bg);
            color: var(--text);
            line-height: 1.6;
            transition: background var(--transition), color var(--transition);
        }

        .container {
            max-width: 1400px;
            margin: 0 auto;
            padding: 32px 24px;
        }

        /* Header */
        .header {
            display: flex;
            justify-content: space-between;
            align-items: center;
            margin-bottom: 40px;
            flex-wrap: wrap;
            gap: 20px;
        }

        .header-content {
            flex: 1;
        }

        .header h1 {
            font-size: 28px;
            font-weight: 700;
            background: linear-gradient(135deg, #fff 0%, #94a3b8 100%);
            -webkit-background-clip: text;
            -webkit-text-fill-color: transparent;
            background-clip: text;
        }

        [data-theme="light"] .header h1 {
            background: linear-gradient(135deg, #1e40af 0%, #3b82f6 100%);
            -webkit-background-clip: text;
            -webkit-text-fill-color: transparent;
            background-clip: text;
        }

        .header p {
            color: var(--text-muted);
            font-size: 14px;
        }

        .theme-toggle {
            background: var(--surface-2);
            border: 1px solid var(--border);
            border-radius: 999px;
            padding: 8px 16px;
            cursor: pointer;
            font-size: 14px;
            font-weight: 500;
            color: var(--text-muted);
            transition: all var(--transition);
        }

        .theme-toggle:hover {
            border-color: var(--accent);
            box-shadow: 0 0 0 3px rgba(59, 130, 246, 0.25);
        }

        /* Target Workflow Notifications */
        .target-workflow-notice-container {
            margin-bottom: 24px;
            display: flex;
            flex-direction: column;
            gap: 12px;
        }

        .target-validation-banner {
            padding: 14px 18px;
            border-radius: var(--radius-sm);
            font-size: 14px;
            font-weight: 500;
            display: flex;
            align-items: center;
            gap: 12px;
            box-shadow: var(--shadow);
            transition: all var(--transition);
        }

        .target-validation-banner.success {
            background: rgba(34, 197, 94, 0.12);
            border: 1px solid var(--success);
            color: var(--success);
        }

        .target-validation-banner.warning {
            background: rgba(245, 158, 11, 0.12);
            border: 1px solid var(--warning);
            color: var(--warning);
        }

        .target-recheck-banner {
            padding: 12px 18px;
            border-radius: var(--radius-sm);
            font-size: 13px;
            display: flex;
            align-items: center;
            justify-content: space-between;
            gap: 12px;
            background: var(--surface);
            border: 1px solid var(--border);
            color: var(--text);
            box-shadow: var(--shadow);
            flex-wrap: wrap;
        }

        /* Metrics Grid */
        .metrics-grid {
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(140px, 1fr));
            gap: 16px;
            margin-bottom: 32px;
        }

        .metric-card {
            background: var(--surface);
            border: 1px solid var(--border);
            border-radius: var(--radius);
            padding: 20px;
            transition: all var(--transition);
        }

        .metric-card:hover {
            border-color: rgba(59, 130, 246, 0.3);
            transform: translateY(-2px);
        }

        .metric-card .label {
            font-size: 12px;
            font-weight: 500;
            text-transform: uppercase;
            letter-spacing: 0.05em;
            color: var(--text-muted);
            margin-bottom: 8px;
        }

        .metric-card .value {
            font-size: 24px;
            font-weight: 600;
            display: flex;
            align-items: center;
            gap: 8px;
            flex-wrap: wrap;
        }

        .pitch-source-badge {
            font-size: 10px;
            font-weight: 600;
            padding: 2px 8px;
            border-radius: 4px;
            text-transform: uppercase;
            letter-spacing: 0.05em;
        }

        .pitch-source-badge.manual {
            background: rgba(59, 130, 246, 0.2);
            color: #3b82f6;
            border: 1px solid rgba(59, 130, 246, 0.3);
        }

        .pitch-source-badge.auto {
            background: rgba(34, 197, 94, 0.2);
            color: #22c55e;
            border: 1px solid rgba(34, 197, 94, 0.3);
        }

        .pitch-source-badge.target {
            background: rgba(245, 158, 11, 0.2);
            color: #f59e0b;
            border: 1px solid rgba(245, 158, 11, 0.3);
        }

        .tolerance-badge {
            font-size: 10px;
            font-weight: 600;
            padding: 2px 8px;
            border-radius: 4px;
            text-transform: uppercase;
            letter-spacing: 0.05em;
        }

        .tolerance-badge.manual {
            background: rgba(59, 130, 246, 0.2);
            color: #3b82f6;
            border: 1px solid rgba(59, 130, 246, 0.3);
        }

        .metric-card .value {
            font-weight: 700;
            font-variant-numeric: tabular-nums;
            color: var(--text);
        }

        .metric-card.highlight .value {
            color: var(--accent);
        }

        /* Table Section */
        .table-section {
            position: relative;
        }

        .table-header {
            display: flex;
            justify-content: space-between;
            align-items: center;
            margin-bottom: 16px;
        }

        .table-section h3 {
            font-size: 14px;
            font-weight: 600;
            color: var(--text-muted);
            text-transform: uppercase;
            letter-spacing: 0.05em;
            margin: 0;
        }

        .table-wrapper {
            background: var(--surface);
            border: 1px solid var(--border);
            border-radius: var(--radius);
            overflow: hidden;
            box-shadow: var(--shadow);
        }

        .table-scroll {
            overflow-x: auto;
        }

        table {
            width: 100%;
            border-collapse: collapse;
            font-size: 13px;
            table-layout: fixed;
        }

        th {
            background: var(--surface-2);
            padding: 12px 6px;
            font-size: 10px;
            font-weight: 600;
            text-transform: uppercase;
            letter-spacing: 0.03em;
            color: var(--text-muted);
            text-align: center;
            border-bottom: 1px solid var(--border);
            white-space: normal;
            line-height: 1.3;
        }

        th:nth-child(1) { width: 7%; }  /* Workstation */
        th:nth-child(2) { width: 7%; }  /* Serial/Id */
        th:nth-child(3) { width: 16%; } /* Operations */
        th:nth-child(4) { width: 9%; }  /* Machine */
        th:nth-child(5) { width: 9%; }  /* Predecessor */
        th:nth-child(6) { width: 9%; }  /* Basic Time */
        th:nth-child(7) { width: 11%; } /* Combined Basic Time */
        th:nth-child(8) { width: 9%; }  /* Balancing SAM */
        th:nth-child(9) { width: 6%; }  /* M/P */
        th:nth-child(10) { width: 10%; } /* Pitch Time / Pitch Time (Auto) */
        th:nth-child(11) { width: 8%; } /* UCL */
        th:nth-child(12) { width: 8%; } /* LCL */
        th:nth-child(13) { width: 10%; } /* Status */

        td {
            padding: 12px 8px;
            border-bottom: 1px solid var(--border);
            color: var(--text);
            font-variant-numeric: tabular-nums;
            text-align: center;
            vertical-align: middle;
            word-wrap: break-word;
            overflow-wrap: break-word;
            white-space: normal;
        }

        .smart-break-cell {
            white-space: pre-wrap;
            line-height: 1.4;
            word-break: break-word;
        }

        tbody tr:hover {
            background: rgba(59, 130, 246, 0.06);
        }

        .status-badge {
            display: inline-block;
            padding: 6px 12px;
            border-radius: 20px;
            font-size: 11px;
            font-weight: 600;
            text-transform: uppercase;
            letter-spacing: 0.03em;
            min-width: 60px;
            white-space: nowrap;
        }

        .status-ok {
            background: rgba(34, 197, 94, 0.15);
            color: var(--success);
        }

        .status-ucl {
            background: rgba(239, 68, 68, 0.15);
            color: var(--danger);
        }

        .status-lcl {
            background: rgba(245, 158, 11, 0.15);
            color: var(--warning);
        }

        /* Export Buttons */
        .export-buttons {
            display: flex;
            gap: 8px;
        }

        .export-button, .export-buttons button, .export-buttons a {
            background: var(--surface-2);
            color: var(--text);
            border: 1px solid var(--border);
            border-radius: var(--radius-sm);
            padding: 8px 16px;
            font-size: 12px;
            font-weight: 500;
            cursor: pointer;
            transition: all var(--transition);
            display: flex;
            align-items: center;
            gap: 6px;
            text-decoration: none;
        }

        .export-button:hover, .export-buttons button:hover, .export-buttons a:hover {
            background: var(--accent);
            color: white;
            border-color: var(--accent);
        }

        .export-button svg, .export-buttons button svg, .export-buttons a svg {
            width: 14px;
            height: 14px;
        }

        /* Status Section */
        .status {
            background: var(--surface);
            border: 1px solid var(--border);
            border-radius: var(--radius);
            padding: 20px;
            text-align: center;
            margin-top: 40px;
        }

        .status p {
            color: var(--text-muted);
            font-size: 16px;
        }

        .status-link {
            color: var(--accent);
            text-decoration: none;
            font-weight: 500;
        }

        .status-link:hover {
            text-decoration: underline;
        }

        @media (max-width: 768px) {
            .container {
                padding: 16px 12px;
                max-width: 100%;
            }

            .header {
                flex-direction: column;
                align-items: flex-start;
                gap: 16px;
            }

            .header h1 {
                font-size: 20px;
            }

            .header p {
                font-size: 12px;
            }

            .metrics-grid {
                grid-template-columns: repeat(2, 1fr);
                gap: 12px;
            }

            .metric-card {
                padding: 16px;
            }

            .metric-card .value {
                font-size: 20px;
            }

            .table-section h3 {
                font-size: 12px;
            }

            .table-wrapper {
                border-radius: 8px;
            }

            .table-scroll {
                overflow-x: auto;
                -webkit-overflow-scrolling: touch;
            }

            table {
                font-size: 11px;
                min-width: 800px;
            }

            th {
                padding: 10px 4px;
                font-size: 9px;
                line-height: 1.2;
            }

            td {
                padding: 10px 4px;
                font-size: 10px;
            }

            .status-badge {
                padding: 4px 8px;
                font-size: 10px;
                min-width: 50px;
            }

            .export-buttons {
                flex-direction: column;
            }

            .export-buttons button, .export-buttons a {
                padding: 8px 12px;
                font-size: 11px;
            }

            .table-header {
                flex-direction: column;
                align-items: flex-start;
                gap: 12px;
            }
        }
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <div class="header-content">
                <h1>Layout View</h1>
                <p>Detailed line balancing report table</p>
            </div>
            <div style="display: flex; align-items: center; gap: 12px;">
                <a href="/" class="export-button" style="text-decoration: none; display: flex; align-items: center; gap: 6px;">
                    <svg xmlns="http://www.w3.org/2000/svg" fill="none" viewBox="0 0 24 24" stroke="currentColor" stroke-width="2" style="width: 14px; height: 14px;">
                        <path stroke-linecap="round" stroke-linejoin="round" d="M3 12l2-2m0 0l7-7 7 7M5 10v10a1 1 0 001 1h3m10-11l2 2m-2-2v10a1 1 0 01-1 1h-3m-6 0a1 1 0 001-1v-4a1 1 0 011-1h2a1 1 0 011 1v4a1 1 0 001 1m-6 0h6" />
                    </svg>
                    Home
                </a>
                <button class="theme-toggle" onclick="toggleTheme()">🌙 Dark</button>
            </div>
        </div>

        {% if has_data %}
        <div style="display: flex; justify-content: space-between; align-items: center; margin-bottom: 16px;">
            <h3 style="font-size: 14px; font-weight: 600; color: var(--text-muted); text-transform: uppercase; letter-spacing: 0.05em; margin: 0;">After Balancing</h3>
            <div class="export-buttons">
                <button onclick="exportFile('xlsx', '{{ session_id }}')">
                    <svg xmlns="http://www.w3.org/2000/svg" fill="none" viewBox="0 0 24 24" stroke="currentColor" stroke-width="2">
                        <path stroke-linecap="round" stroke-linejoin="round" d="M4 16v1a3 3 0 003 3h10a3 3 0 003-3v-1m-4-4l-4 4m0 0l-4-4m4 4V4" />
                    </svg>
                    Excel
                </button>
            </div>
        </div>

        <div class="metrics-grid">
            {% if result.production_target %}
            <div class="metric-card">
                <div class="label">Customer Demand</div>
                <div class="value">{{ result.production_target }}<span style="font-size: 12px; color: var(--text-muted);"> units</span></div>
            </div>
            {% endif %}
            {% if result.shift_time_minutes %}
            <div class="metric-card">
                <div class="label">Available Time</div>
                <div class="value">{{ "%.1f"|format(result.shift_time_minutes) }}<span style="font-size: 12px; color: var(--text-muted);"> min</span></div>
            </div>
            {% endif %}
            {% if result.efficiency_percentage and result.available_time_minutes %}
            <div class="metric-card">
                <div class="label">Required Efficiency</div>
                <div class="value">{{ "%.1f"|format(result.efficiency_percentage) }}<span style="font-size: 12px; color: var(--text-muted);">%</span></div>
            </div>
            {% endif %}
            <div class="metric-card">
                <div class="label">Composite Operations</div>
                <div class="value">{{ result.workstations|length }}</div>
            </div>
            <div class="metric-card">
                <div class="label">SAM<br></div>
                <div class="value">{{ "%.1f"|format(result.total_basic_time) }}<span style="font-size: 12px; color: var(--text-muted);"> min</span></div>
            </div>
            <div class="metric-card">
                <div class="label">Manpower<br></div>
                <div class="value">{{ result.workstations|map(attribute='manpower')|sum }}</div>
            </div>
        </div>

        <div class="metrics-grid">
            <div class="metric-card">
                <div class="label">{% if result.pitch_time_source == "manual" or result.pitch_time_source == "By Target Direct" or result.pitch_time_source == "By Target" %}Takt Time{% else %}Pitch Time{% endif %}</div>
                <div class="value">
                    {{ "%.1f"|format(result.pitch_time) }}<span style="font-size: 12px; color: var(--text-muted);">sec</span>
                    {% if result.pitch_time_source == "manual" %}
                    <span class="pitch-source-badge manual">Manual</span>
                    {% elif result.pitch_time_source == "By Target" %}
                    <span class="pitch-source-badge target">By Target</span>
                    {% elif result.pitch_time_source == "By Target Direct" %}
                    <span class="pitch-source-badge target">By Target Direct</span>
                    {% else %}
                    <span class="pitch-source-badge auto">Auto</span>
                    {% endif %}
                </div>
            </div>
            {% if result.pitch_time_source == "By Target" and result.auto_pitch_time is defined %}
            <div class="metric-card">
                <div class="label">Pitch Time</div>
                <div class="value">{{ "%.1f"|format(result.auto_pitch_time) }}<span style="font-size: 12px; color: var(--text-muted);">sec</span>
                    <span class="pitch-source-badge auto">Auto</span>
                </div>
            </div>
            {% endif %}
            {% if result.pitch_time_source == "calculated" or result.pitch_time_source == "By Target" %}
            <div class="metric-card">
                <div class="label">Tolerance</div>
                <div class="value">
                    {{ "%.1f"|format(result.tolerance * 100) }}<span style="font-size: 12px; color: var(--text-muted);">%</span>
                    {% if result.tolerance * 100 != 15.0 %}
                    <span class="tolerance-badge manual">Manual</span>
                    {% endif %}
                </div>
            </div>
            <div class="metric-card">
                <div class="label">UCL</div>
                <div class="value">{{ "%.1f"|format(result.ucl) }}<span style="font-size: 12px; color: var(--text-muted);">sec</span></div>
            </div>
            <div class="metric-card">
                <div class="label">LCL</div>
                <div class="value">{{ "%.1f"|format(result.lcl) }}<span style="font-size: 12px; color: var(--text-muted);">sec</span></div>
            </div>
            {% endif %}
        </div>


        <div class="table-section">
            <h3 style="font-size: 14px; font-weight: 600; color: var(--text-muted); text-transform: uppercase; letter-spacing: 0.05em; margin-bottom: 16px;">Line Balancing Report</h3>
            <div class="table-wrapper">
                <div class="table-scroll">
                    <table>
                        <thead>
                            <tr>
                                <th>Composite Operations</th>
                                <th>Serial/Id</th>
                                <th>Operations</th>
                                <th>Machine</th>
                                <th>Predecessor</th>
                                <th>Basic<br>SAM</th>
                                <th>Combined Basic<br>SAM</th>
                                <th>Balanced<br>SAM</th>
                                <th>M/P</th>
                                <th>{% if result.pitch_time_source == "manual" or result.pitch_time_source == "By Target Direct" %}Takt Time{% elif result.pitch_time_source == "By Target" %}Pitch Time (Auto){% else %}Pitch Time{% endif %}</th>
                                {% if result.pitch_time_source == "calculated" or result.pitch_time_source == "By Target" %}
                                <th>UCL</th>
                                <th>LCL</th>
                                {% endif %}
                                <th>Status</th>
                            </tr>
                        </thead>
                        <tbody>
                            {% for row in rows %}
                            <tr>
                                <td>{{ row['Composite Operations'] }}</td>
                                <td class="smart-break-cell">{{ row['Serial/Id'] }}</td>
                                <td class="smart-break-cell">{{ row['Operations'] }}</td>
                                <td class="smart-break-cell">{{ row['Machine'] }}</td>
                                <td class="smart-break-cell">{{ row['Predecessor'] }}</td>
                                <td class="smart-break-cell">{{ row['Basic Time'] }}</td>
                                <td>{{ row['Combined Basic Time'] }}</td>
                                <td>{{ row['Balancing SAM'] }}</td>
                                <td>{{ row['M/P'] }}</td>
                                <td>{% if result.pitch_time_source == "manual" or result.pitch_time_source == "By Target Direct" %}{{ row['Takt Time'] }}{% elif result.pitch_time_source == "By Target" %}{{ row['Pitch Time (Auto)'] }}{% else %}{{ row['Pitch Time'] }}{% endif %}</td>
                                {% if result.pitch_time_source == "calculated" or result.pitch_time_source == "By Target" %}
                                <td>{{ row['UCL'] }}</td>
                                <td>{{ row['LCL'] }}</td>
                                {% endif %}
                                <td>
                                    {% if '< Takt Time' in row['Status'] or row['Status'] == 'OK' %}
                                        <span class="status-badge status-ok">{{ row['Status'] }}</span>
                                    {% elif '> Takt Time' in row['Status'] %}
                                        <span class="status-badge status-ucl">> Takt Time</span>
                                    {% elif 'UCL' in row['Status'] or 'Target' in row['Status'] %}
                                        <span class="status-badge status-ucl">{% if 'Target' in row['Status'] %}> Target{% else %}> UCL{% endif %}</span>
                                    {% else %}
                                        <span class="status-badge status-lcl">< LCL</span>
                                    {% endif %}
                                </td>
                            </tr>
                            {% endfor %}
                        </tbody>
                    </table>
                </div>
            </div>
        </div>
        {% else %}
        <div class="status">
            <p>Layout view - detailed line balancing report table</p>
            <p style="margin-top: 10px; font-size: 14px;">
                <a href="/" class="status-link">Load a calculation from the main view</a> to display the layout
            </p>
        </div>
        {% endif %}
    </div>

    <script>
        // Theme Toggle
        function toggleTheme() {
            const html = document.documentElement;
            const currentTheme = html.getAttribute('data-theme') || 'dark';
            const newTheme = currentTheme === 'dark' ? 'light' : 'dark';
            html.setAttribute('data-theme', newTheme);
            localStorage.setItem('theme', newTheme);
            // Show what you're switching TO (opposite of current/new theme)
            // Update all theme toggle buttons on the page
            document.querySelectorAll('.theme-toggle').forEach(button => {
                button.textContent = newTheme === 'dark' ? '☀️ Light' : '🌙 Dark';
            });
        }

        // Restore theme on load
        window.addEventListener('DOMContentLoaded', function() {
            const savedTheme = localStorage.getItem('theme') || 'dark';
            document.documentElement.setAttribute('data-theme', savedTheme);
            // Show what you're switching TO (opposite of current theme)
            // Update all theme toggle buttons on the page
            document.querySelectorAll('.theme-toggle').forEach(button => {
                button.textContent = savedTheme === 'dark' ? '☀️ Light' : '🌙 Dark';
        // Export function
        function exportFile(format, sessionId) {
            window.location.href = `/api/export/${format}/${sessionId}`;
        }
    </script>
</body>
</html>
"""

COMPARISON_TEMPLATE = """
<!DOCTYPE html>
<html lang="en" data-theme="dark">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Takt vs Pitch Comparison — Line Balancing</title>
    <script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
    <style>
        :root {
            --bg: #0f1419;
            --surface: #1a2332;
            --surface-2: #243044;
            --border: rgba(255, 255, 255, 0.08);
            --text: #e8edf4;
            --text-muted: #8b9cb3;
            --accent: #3b82f6;
            --accent-hover: #2563eb;
            --warning: #f59e0b;
            --shadow: 0 4px 24px rgba(0, 0, 0, 0.35);
            --radius: 12px;
            --radius-sm: 8px;
            --transition: 0.2s ease;
        }

        [data-theme="light"] {
            --bg: #f1f5f9;
            --surface: #ffffff;
            --surface-2: #f8fafc;
            --border: rgba(15, 23, 42, 0.1);
            --text: #0f172a;
            --text-muted: #64748b;
            --accent: #2563eb;
            --accent-hover: #1d4ed8;
            --shadow: 0 4px 24px rgba(15, 23, 42, 0.08);
        }

        * { margin: 0; padding: 0; box-sizing: border-box; }

        body {
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
            background: var(--bg);
            color: var(--text);
            line-height: 1.6;
            transition: background var(--transition), color var(--transition);
        }

        .container {
            max-width: 1480px;
            margin: 0 auto;
            padding: 32px 24px 64px;
        }

        /* Header */
        .header {
            display: flex;
            justify-content: space-between;
            align-items: center;
            margin-bottom: 28px;
            flex-wrap: wrap;
            gap: 20px;
        }

        .header-content h1 {
            font-size: 26px;
            font-weight: 700;
            background: linear-gradient(135deg, #fff 0%, #94a3b8 100%);
            -webkit-background-clip: text;
            -webkit-text-fill-color: transparent;
            background-clip: text;
        }

        .header-content p {
            color: var(--text-muted);
            font-size: 14px;
            margin-top: 4px;
        }

        .header-actions {
            display: flex;
            align-items: center;
            gap: 12px;
            flex-wrap: wrap;
        }

        .nav-tabs {
            display: flex;
            gap: 6px;
            background: var(--surface-2);
            padding: 4px;
            border-radius: var(--radius-sm);
            border: 1px solid var(--border);
        }

        .nav-tab {
            padding: 6px 14px;
            font-size: 13px;
            font-weight: 600;
            color: var(--text-muted);
            text-decoration: none;
            border-radius: 6px;
            transition: var(--transition);
        }

        .nav-tab:hover {
            color: var(--text);
            background: rgba(255, 255, 255, 0.05);
        }

        .nav-tab.active {
            color: #fff;
            background: var(--accent);
        }

        .btn-export {
            display: inline-flex;
            align-items: center;
            gap: 6px;
            padding: 8px 16px;
            font-size: 13px;
            font-weight: 600;
            background: #10b981;
            color: #fff;
            border: none;
            border-radius: var(--radius-sm);
            cursor: pointer;
            text-decoration: none;
            transition: var(--transition);
        }

        .btn-export:hover {
            background: #059669;
        }

        .theme-toggle {
            padding: 8px 14px;
            background: var(--surface-2);
            border: 1px solid var(--border);
            border-radius: var(--radius-sm);
            color: var(--text);
            cursor: pointer;
            font-size: 13px;
            font-weight: 500;
            transition: var(--transition);
        }

        .theme-toggle:hover {
            border-color: var(--accent);
        }

        /* Form Card */
        .form-card {
            background: var(--surface);
            border: 1px solid var(--border);
            border-radius: var(--radius);
            padding: 24px;
            margin-bottom: 30px;
            box-shadow: var(--shadow);
        }

        .form-card h2 {
            font-size: 17px;
            font-weight: 600;
            margin-bottom: 18px;
            color: var(--text);
            display: flex;
            align-items: center;
            gap: 8px;
        }

        .form-grid {
            display: grid;
            grid-template-columns: 2fr 1fr 1fr auto;
            gap: 16px;
            align-items: end;
        }

        @media (max-width: 900px) {
            .form-grid {
                grid-template-columns: 1fr;
            }
        }

        .field label {
            display: block;
            font-size: 12px;
            font-weight: 600;
            color: var(--text-muted);
            margin-bottom: 6px;
            text-transform: uppercase;
            letter-spacing: 0.5px;
        }

        .field input[type="file"],
        .field input[type="number"] {
            width: 100%;
            padding: 10px 14px;
            font-size: 14px;
            border-radius: var(--radius-sm);
            border: 1px solid var(--border);
            background: var(--surface-2);
            color: var(--text);
            outline: none;
            transition: var(--transition);
        }

        .field input[type="file"] {
            padding: 7px 10px;
        }

        .field input:focus {
            border-color: var(--accent);
            box-shadow: 0 0 0 3px rgba(59, 130, 246, 0.2);
        }

        .form-card button[type="submit"] {
            padding: 11px 24px;
            font-size: 14px;
            font-weight: 600;
            background: var(--accent);
            color: #fff;
            border: none;
            border-radius: var(--radius-sm);
            cursor: pointer;
            transition: var(--transition);
            white-space: nowrap;
        }

        .form-card button[type="submit"]:hover {
            background: var(--accent-hover);
        }

        /* Error Box */
        .error-box {
            background: rgba(239, 68, 68, 0.1);
            border: 1px solid rgba(239, 68, 68, 0.3);
            color: #f87171;
            padding: 16px 20px;
            border-radius: var(--radius-sm);
            margin-bottom: 24px;
            font-size: 14px;
        }

        /* Metrics Card Rows */
        .section-header {
            display: flex;
            justify-content: space-between;
            align-items: center;
            margin-bottom: 14px;
            flex-wrap: wrap;
            gap: 10px;
        }

        .section-title {
            font-size: 17px;
            font-weight: 700;
            color: var(--text);
            display: flex;
            align-items: center;
            gap: 8px;
        }

        .metrics-grid {
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(150px, 1fr));
            gap: 12px;
            margin-bottom: 28px;
        }

        .metric-card {
            background: var(--surface);
            border: 1px solid var(--border);
            border-radius: var(--radius-sm);
            padding: 14px;
            transition: var(--transition);
        }

        .metric-card:hover {
            border-color: rgba(59, 130, 246, 0.4);
            transform: translateY(-2px);
        }

        .metric-card .label {
            font-size: 11px;
            font-weight: 600;
            color: var(--text-muted);
            text-transform: uppercase;
            letter-spacing: 0.5px;
            margin-bottom: 4px;
        }

        .metric-card .value {
            font-size: 20px;
            font-weight: 700;
            color: var(--text);
        }

        .metric-card .unit {
            font-size: 11px;
            font-weight: 500;
            color: var(--text-muted);
            margin-left: 2px;
        }

        /* Workstation Layout Single View & Switcher */
        .layout-control-card {
            background: var(--surface);
            border: 1px solid var(--border);
            border-radius: var(--radius);
            padding: 24px;
            margin-bottom: 32px;
            box-shadow: var(--shadow);
        }

        .layout-control-header {
            display: flex;
            justify-content: space-between;
            align-items: center;
            flex-wrap: wrap;
            gap: 16px;
            margin-bottom: 20px;
            padding-bottom: 16px;
            border-bottom: 1px solid var(--border);
        }

        .layout-tabs {
            display: flex;
            background: var(--surface-2);
            padding: 4px;
            border-radius: 8px;
            border: 1px solid var(--border);
            gap: 4px;
        }

        .layout-tab-btn {
            padding: 7px 16px;
            font-size: 13px;
            font-weight: 600;
            border: none;
            border-radius: 6px;
            background: transparent;
            color: var(--text-muted);
            cursor: pointer;
            transition: var(--transition);
        }

        .layout-tab-btn:hover {
            color: var(--text);
        }

        .layout-tab-btn.active {
            background: var(--accent);
            color: #ffffff;
            box-shadow: 0 2px 8px rgba(37, 99, 235, 0.3);
        }

        .method-layout-container {
            animation: fadeIn 0.2s ease-in-out;
        }

        @keyframes fadeIn {
            from { opacity: 0; transform: translateY(4px); }
            to { opacity: 1; transform: translateY(0); }
        }

        .method-col-header {
            display: flex;
            justify-content: space-between;
            align-items: center;
            margin-bottom: 16px;
            padding-bottom: 14px;
            border-bottom: 1px solid var(--border);
            flex-wrap: wrap;
            gap: 12px;
        }

        .method-header-left {
            display: flex;
            align-items: center;
            gap: 14px;
            flex-wrap: wrap;
        }

        .btn-method-switch {
            display: inline-flex;
            align-items: center;
            gap: 8px;
            padding: 8px 18px;
            font-size: 13px;
            font-weight: 700;
            font-family: inherit;
            background: linear-gradient(135deg, #2563eb 0%, #059669 100%);
            color: #ffffff !important;
            border: 1px solid rgba(255, 255, 255, 0.2);
            border-radius: 8px;
            cursor: pointer;
            text-decoration: none;
            transition: all 0.2s cubic-bezier(0.4, 0, 0.2, 1);
            box-shadow: 0 4px 14px rgba(37, 99, 235, 0.3);
        }

        .btn-method-switch:hover {
            background: linear-gradient(135deg, #1d4ed8 0%, #047857 100%);
            transform: translateY(-2px);
            box-shadow: 0 6px 18px rgba(5, 150, 105, 0.4);
        }

        .btn-method-switch:active {
            transform: translateY(0);
        }

        .switch-icon {
            font-size: 14px;
            display: inline-block;
            transition: transform 0.3s ease;
        }

        .btn-method-switch:hover .switch-icon {
            transform: rotate(180deg);
        }

        .method-badge {
            display: inline-flex;
            align-items: center;
            gap: 6px;
            padding: 6px 14px;
            border-radius: 9999px;
            font-size: 13px;
            font-weight: 700;
        }

        .badge-takt {
            background: rgba(34, 197, 94, 0.15);
            color: #22c55e;
            border: 1px solid rgba(34, 197, 94, 0.3);
        }

        .badge-pitch {
            background: rgba(251, 146, 60, 0.15);
            color: #fb923c;
            border: 1px solid rgba(251, 146, 60, 0.3);
        }

        .method-desc {
            font-size: 12px;
            color: var(--text-muted);
        }

        .mini-kpi-grid {
            display: grid;
            grid-template-columns: repeat(4, 1fr);
            gap: 8px;
            margin-bottom: 16px;
        }

        @media (max-width: 700px) {
            .mini-kpi-grid {
                grid-template-columns: repeat(2, 1fr);
            }
        }

        .mini-kpi {
            background: var(--surface-2);
            padding: 8px 10px;
            border-radius: 6px;
            border: 1px solid var(--border);
        }

        .mini-kpi .k-lbl {
            font-size: 10px;
            font-weight: 600;
            color: var(--text-muted);
            text-transform: uppercase;
        }

        .mini-kpi .k-val {
            font-size: 14px;
            font-weight: 700;
            color: var(--text);
        }

        /* Tables */
        .table-scroll {
            overflow-x: auto;
            max-height: 480px;
            overflow-y: auto;
            border-radius: var(--radius-sm);
            border: 1px solid var(--border);
        }

        .data-table {
            width: 100%;
            border-collapse: collapse;
            font-size: 12px;
            text-align: left;
        }

        .data-table th {
            background: var(--surface-2);
            color: var(--text-muted);
            font-weight: 600;
            padding: 10px 12px;
            text-transform: uppercase;
            font-size: 10px;
            letter-spacing: 0.5px;
            position: sticky;
            top: 0;
            z-index: 2;
            border-bottom: 1px solid var(--border);
            white-space: nowrap;
        }

        .data-table td {
            padding: 9px 12px;
            border-bottom: 1px solid var(--border);
            color: var(--text);
            white-space: nowrap;
        }

        .data-table tbody tr:hover {
            background: rgba(255, 255, 255, 0.02);
        }

        .row-flagged {
            background: rgba(245, 158, 11, 0.08) !important;
        }

        .row-flagged:hover {
            background: rgba(245, 158, 11, 0.12) !important;
        }

        /* Status Badges */
        .status-badge {
            display: inline-block;
            padding: 2px 8px;
            border-radius: 9999px;
            font-size: 11px;
            font-weight: 600;
            text-align: center;
        }

        .status-ok {
            background: rgba(34, 197, 94, 0.15);
            color: #4ade80;
            border: 1px solid rgba(34, 197, 94, 0.3);
        }

        .status-warning {
            background: rgba(245, 158, 11, 0.2);
            color: #fbbf24;
            border: 1px solid rgba(245, 158, 11, 0.4);
        }

        .status-danger {
            background: rgba(239, 68, 68, 0.15);
            color: #f87171;
            border: 1px solid rgba(239, 68, 68, 0.3);
        }

        /* Master Comparison Section */
        .comparison-section {
            background: var(--surface);
            border: 1px solid var(--border);
            border-radius: var(--radius);
            padding: 24px;
            margin-bottom: 32px;
            box-shadow: var(--shadow);
        }

        .comparison-table-wrapper {
            overflow-x: auto;
            border-radius: var(--radius-sm);
            border: 1px solid var(--border);
        }

        .comparison-table {
            width: 100%;
            border-collapse: collapse;
            font-size: 13px;
        }

        .comparison-table th {
            background: var(--surface-2);
            color: var(--text-muted);
            font-weight: 700;
            padding: 12px 16px;
            text-transform: uppercase;
            font-size: 11px;
            letter-spacing: 0.5px;
            border-bottom: 1px solid var(--border);
            text-align: center;
        }

        .comparison-table th:first-child {
            text-align: left;
        }

        .comparison-table td {
            padding: 12px 16px;
            border-bottom: 1px solid var(--border);
            text-align: center;
            font-size: 14px;
        }

        .comparison-table td:first-child {
            text-align: left;
        }

        .comparison-table tbody tr:nth-child(even) {
            background: rgba(255, 255, 255, 0.015);
        }

        .metric-col-title {
            font-weight: 600;
            color: var(--text);
        }

        .metric-unit {
            font-size: 11px;
            color: var(--text-muted);
            margin-left: 4px;
        }

        .col-before {
            color: var(--text-muted);
            font-weight: 500;
        }

        .col-method-a {
            color: #22c55e;
            font-weight: 600;
        }

        [data-theme="dark"] .col-method-a {
            color: #4ade80;
        }

        .col-method-b {
            color: #ea580c;
            font-weight: 600;
        }

        [data-theme="dark"] .col-method-b {
            color: #fb923c;
        }

        /* Winner Highlighting */
        .cell-winner {
            background: rgba(34, 197, 94, 0.15) !important;
            color: #22c55e !important;
            font-weight: 700 !important;
        }

        [data-theme="light"] .cell-winner {
            background: #dcfce7 !important;
            color: #15803d !important;
        }

        .winner-pill {
            display: inline-block;
            background: #22c55e;
            color: #ffffff;
            font-size: 9px;
            font-weight: 700;
            padding: 1px 5px;
            border-radius: 4px;
            margin-left: 6px;
            vertical-align: middle;
            text-transform: uppercase;
        }

        [data-theme="light"] .winner-pill {
            background: #16a34a;
        }

        /* Chart Card */
        .chart-card {
            background: var(--surface);
            border: 1px solid var(--border);
            border-radius: var(--radius);
            padding: 24px;
            margin-bottom: 32px;
            box-shadow: var(--shadow);
        }

        .chart-tabs {
            display: flex;
            gap: 8px;
        }

        .chart-tab-btn {
            padding: 6px 14px;
            font-size: 12px;
            font-weight: 600;
            border: 1px solid var(--border);
            border-radius: 6px;
            background: var(--surface-2);
            color: var(--text-muted);
            cursor: pointer;
            transition: var(--transition);
        }

        .chart-tab-btn:hover {
            color: var(--text);
            border-color: var(--accent);
        }

        .chart-tab-btn.active {
            background: var(--accent);
            color: #fff;
            border-color: var(--accent);
        }

        .chart-container {
            position: relative;
            height: 420px;
            width: 100%;
            margin-top: 14px;
        }

        /* ──── Understanding Production Optimization Charts Section ──── */
        .opt-overview {
            background: var(--surface);
            border: 1px solid var(--border);
            border-radius: var(--radius);
            padding: 28px 24px 20px 24px;
            margin-bottom: 32px;
            box-shadow: var(--shadow);
            min-height: calc(100vh - 180px);
            display: flex;
            flex-direction: column;
        }

        .opt-overview__title {
            font-size: 22px;
            font-weight: 800;
            color: var(--text);
            text-align: center;
            margin: 0 0 4px 0;
            letter-spacing: -0.4px;
        }

        [data-theme="light"] .opt-overview__title {
            color: #1a2b49;
        }

        .opt-overview__subtitle {
            font-size: 13px;
            color: var(--text-muted);
            text-align: center;
            margin: 0 0 20px 0;
        }

        .opt-overview__body {
            display: grid;
            grid-template-columns: 280px 1fr;
            gap: 20px;
            flex: 1;
            min-height: 0;
        }

        @media (max-width: 1100px) {
            .opt-overview__body {
                grid-template-columns: 1fr;
            }
        }

        /* Key / Legend Panel */
        .opt-key-panel {
            background: var(--surface-2);
            border: 1px solid var(--border);
            border-radius: var(--radius-sm);
            padding: 20px 18px;
            display: flex;
            flex-direction: column;
            gap: 16px;
        }

        [data-theme="light"] .opt-key-panel {
            background: #f8fafc;
            border-color: #cbd5e1;
        }

        .opt-key-panel__heading {
            font-size: 14px;
            font-weight: 700;
            color: var(--text);
            margin: 0;
            padding-bottom: 10px;
            border-bottom: 1px solid var(--border);
        }

        .opt-key-item {
            display: flex;
            align-items: flex-start;
            gap: 10px;
            font-size: 13px;
            line-height: 1.5;
            color: var(--text);
        }

        .opt-key-item__icon {
            flex-shrink: 0;
            width: 22px;
            height: 22px;
            display: flex;
            align-items: center;
            justify-content: center;
            font-size: 15px;
            margin-top: 1px;
        }

        .opt-key-item__text strong {
            color: var(--text);
        }

        .opt-key-item__text {
            color: var(--text-muted);
        }

        /* Charts Grid (3 charts) */
        .opt-charts-grid {
            display: grid;
            grid-template-columns: 1fr;
            grid-template-rows: 1fr 1fr;
            gap: 16px;
            min-height: 0;
        }

        .opt-charts-grid__top-row {
            display: grid;
            grid-template-columns: 1fr;
            gap: 0;
            min-height: 0;
        }

        .opt-charts-grid__bottom-row {
            display: grid;
            grid-template-columns: 1fr 1fr;
            gap: 16px;
            min-height: 0;
        }

        @media (max-width: 800px) {
            .opt-charts-grid__bottom-row {
                grid-template-columns: 1fr;
            }
        }

        .opt-chart-box {
            background: var(--surface-2);
            border: 1px solid var(--border);
            border-radius: var(--radius-sm);
            padding: 14px 14px 10px 14px;
            display: flex;
            flex-direction: column;
            min-height: 0;
        }

        [data-theme="light"] .opt-chart-box {
            background: #ffffff;
            border-color: #cbd5e1;
        }

        .opt-chart-box__label {
            font-size: 13px;
            font-weight: 700;
            color: var(--text);
            margin: 0 0 8px 0;
            letter-spacing: 0.3px;
            text-transform: uppercase;
        }

        .opt-chart-box__canvas-wrap {
            position: relative;
            flex: 1;
            min-height: 140px;
        }

        /* Visual Analysis & Grouped KPI 8-Card Section */
        .kpi-visual-container {
            padding: 12px 4px 8px 4px;
        }

        .kpi-visual-header {
            display: flex;
            justify-content: space-between;
            align-items: flex-end;
            flex-wrap: wrap;
            gap: 16px;
            margin-bottom: 24px;
            padding-bottom: 12px;
            border-bottom: 1px solid var(--border);
        }

        .kpi-header-left {
            max-width: 650px;
        }

        .kpi-eyebrow {
            font-size: 11px;
            font-weight: 800;
            letter-spacing: 1.5px;
            color: #d9532f;
            text-transform: uppercase;
            margin-bottom: 6px;
        }

        .kpi-main-title {
            font-size: 26px;
            font-weight: 800;
            color: var(--text);
            margin: 0 0 6px 0;
            line-height: 1.25;
            letter-spacing: -0.5px;
            font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, "Helvetica Neue", Arial, sans-serif;
        }

        [data-theme="light"] .kpi-main-title {
            color: #1a2b49;
        }

        .kpi-accent-glance {
            border-bottom: 2px dotted #d9532f;
            padding-bottom: 2px;
        }

        .kpi-subtitle {
            font-size: 13px;
            color: var(--text-muted);
            margin: 0;
            line-height: 1.4;
        }

        .kpi-legend {
            display: flex;
            align-items: center;
            gap: 20px;
            flex-wrap: wrap;
            padding-bottom: 4px;
        }

        .kpi-legend-item {
            display: flex;
            align-items: center;
            gap: 8px;
            font-size: 12px;
            font-weight: 600;
            color: var(--text);
        }

        .legend-box {
            width: 13px;
            height: 13px;
            border-radius: 2px;
            display: inline-block;
        }

        .legend-before {
            background-color: #ef4444;
        }

        .legend-takt {
            background-color: #22c55e;
        }

        .legend-pitch {
            background-color: #fb923c;
        }

        /* 8-Card Grid */
        .kpi-cards-grid {
            display: grid;
            grid-template-columns: repeat(4, 1fr);
            gap: 16px;
        }

        @media (max-width: 1200px) {
            .kpi-cards-grid {
                grid-template-columns: repeat(2, 1fr);
            }
        }

        @media (max-width: 640px) {
            .kpi-cards-grid {
                grid-template-columns: 1fr;
            }
        }

        .kpi-mini-card {
            background: var(--surface-2);
            border: 1px solid var(--border);
            border-radius: 4px;
            padding: 16px 14px 12px 14px;
            display: flex;
            flex-direction: column;
            justify-content: space-between;
            min-height: 220px;
            box-shadow: 0 1px 3px rgba(0, 0, 0, 0.04);
            transition: transform 0.2s ease, box-shadow 0.2s ease;
        }

        [data-theme="light"] .kpi-mini-card {
            background: #ffffff;
            border-color: #cbd5e1;
        }

        .kpi-mini-card:hover {
            transform: translateY(-2px);
            box-shadow: 0 4px 12px rgba(0, 0, 0, 0.08);
        }

        .kpi-mini-title {
            font-size: 13px;
            font-weight: 700;
            color: #1e3a8a;
            margin-bottom: 12px;
            line-height: 1.3;
            min-height: 32px;
        }

        [data-theme="dark"] .kpi-mini-title {
            color: #93c5fd;
        }

        .kpi-mini-chart {
            display: flex;
            flex-direction: column;
            width: 100%;
            height: 150px;
            justify-content: flex-end;
        }

        .kpi-bars-area {
            display: flex;
            justify-content: space-around;
            align-items: flex-end;
            height: 110px;
            width: 100%;
            padding: 0 8px;
        }

        .kpi-bar-col {
            display: flex;
            flex-direction: column;
            align-items: center;
            justify-content: flex-end;
            width: 28%;
            height: 100%;
        }

        .kpi-bar-val {
            font-size: 11.5px;
            font-weight: 700;
            color: var(--text);
            margin-bottom: 4px;
            text-align: center;
            white-space: nowrap;
        }

        [data-theme="light"] .kpi-bar-val {
            color: #1e293b;
        }

        .kpi-bar {
            width: 24px;
            border-radius: 2px 2px 0 0;
            transition: height 0.4s ease;
            min-height: 4px;
        }

        .bar-before {
            background-color: #ef4444;
        }

        .bar-takt {
            background-color: #22c55e;
        }

        .bar-pitch {
            background-color: #fb923c;
        }

        .kpi-bar:hover {
            filter: brightness(1.15);
        }

        .kpi-baseline-axis {
            width: 100%;
            height: 1px;
            background-color: #94a3b8;
            display: flex;
            justify-content: space-around;
            position: relative;
        }

        .kpi-axis-tick {
            width: 1px;
            height: 4px;
            background-color: #94a3b8;
        }

        .kpi-axis-labels {
            display: flex;
            justify-content: space-around;
            width: 100%;
            padding-top: 5px;
        }

        .kpi-label {
            font-size: 11px;
            color: var(--text-muted);
            text-align: center;
            width: 33.33%;
            font-weight: 500;
        }

        [data-theme="light"] .kpi-label {
            color: #475569;
        }

        /* Recommendations Card */
        .rec-card {
            background: linear-gradient(180deg, var(--surface) 0%, var(--surface-2) 100%);
            border: 1px solid var(--border);
            border-left: 4px solid var(--accent);
            border-radius: var(--radius);
            padding: 24px;
            margin-bottom: 32px;
            box-shadow: var(--shadow);
        }

        .rec-list {
            display: flex;
            flex-direction: column;
            gap: 12px;
            margin-top: 14px;
        }

        .rec-item {
            display: flex;
            align-items: flex-start;
            gap: 12px;
            font-size: 14px;
            line-height: 1.6;
            color: var(--text);
        }

        .rec-bullet {
            color: var(--accent);
            font-weight: 700;
            font-size: 18px;
            line-height: 1.2;
        }

        /* Bottom Action Bar */
        .bottom-action-bar {
            display: flex;
            justify-content: center;
            align-items: center;
            margin-top: 24px;
        }

        .btn-large-export {
            display: inline-flex;
            align-items: center;
            gap: 10px;
            padding: 14px 32px;
            font-size: 15px;
            font-weight: 700;
            background: #10b981;
            color: #ffffff;
            border: none;
            border-radius: var(--radius-sm);
            cursor: pointer;
            text-decoration: none;
            transition: var(--transition);
            box-shadow: 0 4px 16px rgba(16, 185, 129, 0.3);
        }

        .btn-large-export:hover {
            background: #059669;
            transform: translateY(-2px);
            box-shadow: 0 6px 20px rgba(16, 185, 129, 0.4);
        }
    </style>
</head>
<body>
    <div class="container">
        <!-- Header -->
        <div class="header">
            <div class="header-content">
                <h1>Takt vs Pitch Comparison</h1>
                <p>Parallel balancing passes: Method A (Strict Takt) vs Method B (IE Pitch ±15% Classification)</p>
            </div>
            <div class="header-actions">
                <nav class="nav-tabs">
                    <a href="/" class="nav-tab">Line Balancing</a>
                    <a href="/compare" class="nav-tab active">Takt vs Pitch Comparison</a>
                </nav>
                {% if session_id %}
                <a href="/api/export/compare/xlsx/{{ session_id }}" class="btn-export">
                    <span>Export Excel</span>
                </a>
                {% endif %}
                <button class="theme-toggle" onclick="toggleTheme()">🌙 Dark</button>
            </div>
        </div>

        <!-- 1. Input Config Card -->
        <form method="post" action="/compare" enctype="multipart/form-data" class="form-card">
            <h2>
                <span>Comparison Parameters</span>
            </h2>
            <div class="form-grid">
                <div class="field">
                    <label>Upload Excel/CSV Operations File</label>
                    <input type="file" name="file" accept=".csv,.xlsx,.xls" required>
                </div>
                <div class="field">
                    <label>Available Time</label>
                    <input type="number" name="shift_time" value="{% if result %}{{ result.shift_time_minutes }}{% else %}420{% endif %}" min="1" step="1" required>
                </div>
                <div class="field">
                    <label>Customer Demand</label>
                    <input type="number" name="production_target" value="{% if result %}{{ result.production_target }}{% else %}500{% endif %}" min="1" step="1" required>
                </div>
                <div class="field">
                    <button type="submit">Run Comparison</button>
                </div>
            </div>
        </form>

        {% if error %}
        <div class="error-box">
            <strong>Error:</strong> {{ error|safe }}
        </div>
        {% endif %}

        {% if result %}
        <!-- 2. Baseline "Before Balancing" KPI Card Row -->
        <div class="section-header">
            <div class="section-title">
                <span>Baseline Parameters & Control Limits</span>
            </div>
        </div>
        <div class="metrics-grid">
            <div class="metric-card">
                <div class="label">Customer Demand</div>
                <div class="value">{{ result.production_target }} <span class="unit">units</span></div>
            </div>
            <div class="metric-card">
                <div class="label">Available Time</div>
                <div class="value">{{ "%.1f"|format(result.shift_time_minutes) }} <span class="unit">min</span></div>
            </div>
            <div class="metric-card">
                <div class="label">SAM</div>
                <div class="value">{{ "%.1f"|format(result.total_sam / 60) }} <span class="unit">min</span></div>
            </div>
            <div class="metric-card">
                <div class="label">Input Operations</div>
                <div class="value">{{ result.before.num_operations }} <span class="unit">ops</span></div>
            </div>
            <div class="metric-card">
                <div class="label">Baseline Manpower</div>
                <div class="value">{{ result.before.total_manpower }} <span class="unit">operators</span></div>
            </div>
            <div class="metric-card">
                <div class="label">Baseline Efficiency</div>
                <div class="value">{{ "%.1f"|format(result.before.efficiency_balancing_rate) }} <span class="unit">%</span></div>
            </div>
            <div class="metric-card">
                <div class="label">Takt / IE Pitch Time</div>
                <div class="value" style="color: #60a5fa;">{{ "%.1f"|format(result.takt_time) }} / {{ "%.1f"|format(result.pitch_time) }} <span class="unit">sec</span></div>
            </div>
            <div class="metric-card">
                <div class="label">UCL / LCL (±15%)</div>
                <div class="value">{{ "%.1f"|format(result.ucl) }} / {{ "%.1f"|format(result.lcl) }} <span class="unit">sec</span></div>
            </div>
            {% if result.method_b.review_flag_count > 0 %}
            <div class="metric-card" style="border-color: rgba(245, 158, 11, 0.4);">
                <div class="label" style="color: var(--warning);">Method B Flags</div>
                <div class="value" style="color: var(--warning);">{{ result.method_b.review_flag_count }} <span class="status-badge status-warning">> UCL</span></div>
            </div>
            {% endif %}
        </div>

        <!-- ── Understanding Production Optimization Charts ── -->
        <div class="opt-overview" id="optOverviewSection">
            <h2 class="opt-overview__title">Understanding Production Optimization Charts</h2>
            <p class="opt-overview__subtitle">All details which are compulsory to understand the charts — consistent time scales across all three graphs for accurate visual comparison</p>

            <div class="opt-overview__body">
                <!-- Left: Key / Legend Panel -->
                <div class="opt-key-panel">
                    <h3 class="opt-key-panel__heading">Chart Reading Guide</h3>
                    <div class="opt-key-item">
                        <span class="opt-key-item__icon">📊</span>
                        <span class="opt-key-item__text"><strong>Legend:</strong> Bars represent operator stations / workstations and their assigned time</span>
                    </div>
                    <div class="opt-key-item">
                        <span class="opt-key-item__icon">📐</span>
                        <span class="opt-key-item__text"><strong>Axes:</strong> X-axis = Stations · Y-axis = Time (seconds) — <em>same scale across all charts</em></span>
                    </div>
                    <div class="opt-key-item">
                        <span class="opt-key-item__icon">⚙️</span>
                        <span class="opt-key-item__text"><strong>Takt Time:</strong> {{ "%.1f"|format(result.takt_time) }}s — Maximum time per station to meet customer demand</span>
                    </div>
                    <div class="opt-key-item">
                        <span class="opt-key-item__icon">📋</span>
                        <span class="opt-key-item__text"><strong>IE Pitch Time:</strong> {{ "%.1f"|format(result.pitch_time) }}s — Industrial Engineering standard target time</span>
                    </div>
                    <div class="opt-key-item">
                        <span class="opt-key-item__icon">📈</span>
                        <span class="opt-key-item__text"><strong>UCL (Upper Control Limit):</strong> {{ "%.1f"|format(result.ucl) }}s — Pitch + 15% tolerance ceiling for Method B</span>
                    </div>
                    <div class="opt-key-item">
                        <span class="opt-key-item__icon">🎯</span>
                        <span class="opt-key-item__text"><strong>Primary Goal:</strong> Bars closer to the reference line = better balanced line with less idle time</span>
                    </div>
                </div>

                <!-- Right: 3 Charts Grid -->
                <div class="opt-charts-grid">
                    <!-- Top: Before Balancing (full width) -->
                    <div class="opt-charts-grid__top-row">
                        <div class="opt-chart-box">
                            <h4 class="opt-chart-box__label">Before Balancing</h4>
                            <div class="opt-chart-box__canvas-wrap">
                                <canvas id="optChartBefore"></canvas>
                            </div>
                        </div>
                    </div>
                    <!-- Bottom: Method A & Method B side by side -->
                    <div class="opt-charts-grid__bottom-row">
                        <div class="opt-chart-box">
                            <h4 class="opt-chart-box__label">Method A — Takt Time Balancing</h4>
                            <div class="opt-chart-box__canvas-wrap">
                                <canvas id="optChartMethodA"></canvas>
                            </div>
                        </div>
                        <div class="opt-chart-box">
                            <h4 class="opt-chart-box__label">Method B — IE Pitch Time Balancing</h4>
                            <div class="opt-chart-box__canvas-wrap">
                                <canvas id="optChartMethodB"></canvas>
                            </div>
                        </div>
                    </div>
                </div>
            </div>
        </div>

        <!-- 3. Single Master Comparison Table -->
        <div class="comparison-section">
            <div class="section-header">
                <div class="section-title">
                    <span>Master 8-KPI Side-by-Side Comparison</span>
                </div>
                <div style="font-size: 12px; color: var(--text-muted);">
                    <span class="winner-pill" style="margin-right: 4px;">★</span> Highlighted cell indicates better-performing method per KPI
                </div>
            </div>
            <div class="comparison-table-wrapper">
                <table class="comparison-table">
                    <thead>
                        <tr>
                            <th style="width: 32%;">Metric Name</th>
                            <th style="width: 22%;">Before Balancing (Baseline)</th>
                            <th style="width: 23%; color: #22c55e;">Method A — Takt Time</th>
                            <th style="width: 23%; color: #fb923c;">Method B — IE Pitch</th>
                        </tr>
                    </thead>
                    <tbody>
                        {% for row in result.comparison %}
                        <tr>
                            <td>
                                <span class="metric-col-title">{{ row.metric }}</span>
                                <span class="metric-unit">({{ row.unit }})</span>
                            </td>
                            <td class="col-before">{{ row.formatted_before }}</td>
                            <td class="col-method-a {% if row.winner == 'method_a' or row.winner == 'tie' %}cell-winner{% endif %}">
                                {{ row.formatted_method_a }}
                                {% if row.winner == 'method_a' %}<span class="winner-pill">★</span>{% endif %}
                            </td>
                            <td class="col-method-b {% if row.winner == 'method_b' or row.winner == 'tie' %}cell-winner{% endif %}">
                                {{ row.formatted_method_b }}
                                {% if row.winner == 'method_b' %}<span class="winner-pill">★</span>{% endif %}
                            </td>
                        </tr>
                        {% endfor %}
                    </tbody>
                </table>
            </div>
        </div>

        <!-- 4. Grouped Bar Chart & Profile Curve (Visual Analysis) -->
        <div class="chart-card">
            <div class="section-header">
                <div class="section-title">
                    <span>Visual Analysis & Comparison Curves</span>
                </div>
                <div class="chart-tabs">
                    <button class="chart-tab-btn active" onclick="switchChartMode('grouped_bar', this)">Grouped KPI Comparison</button>
                    <button class="chart-tab-btn" onclick="switchChartMode('profile_lines', this)">Balancing Profile Curves</button>
                </div>
            </div>

            <!-- Grouped KPI Comparison 8-Card Section -->
            <div id="groupedKpiSection" class="kpi-visual-container">
                <div class="kpi-visual-header">
                    <div class="kpi-header-left">
                        <h2 class="kpi-main-title">All KPIs — Before vs. After, <span class="kpi-accent-glance">at a Glance</span></h2>
                        <p class="kpi-subtitle">Every KPI from the master table, charted side by side across the three scenarios.</p>
                    </div>
                    <div class="kpi-legend">
                        <div class="kpi-legend-item">
                            <span class="legend-box legend-before"></span>
                            <span class="legend-text">Before</span>
                        </div>
                        <div class="kpi-legend-item">
                            <span class="legend-box legend-takt"></span>
                            <span class="legend-text">After – Takt</span>
                        </div>
                        <div class="kpi-legend-item">
                            <span class="legend-box legend-pitch"></span>
                            <span class="legend-text">After – IE Pitch</span>
                        </div>
                    </div>
                </div>

                <div class="kpi-cards-grid">
                    {% set kpi_cards_data = [
                        {
                            'title': 'Operations (after merging)',
                            'b': result.before.num_operations,
                            'a': result.method_a.num_workstations,
                            'p': result.method_b.num_workstations,
                            'fmt': 'int'
                        },
                        {
                            'title': 'Number of Operators',
                            'b': result.before.total_manpower,
                            'a': result.method_a.total_manpower,
                            'p': result.method_b.total_manpower,
                            'fmt': 'int'
                        },
                        {
                            'title': 'Cycle Time (sec)',
                            'b': result.before.cycle_time,
                            'a': result.method_a.cycle_time,
                            'p': result.method_b.cycle_time,
                            'fmt': 'float'
                        },
                        {
                            'title': 'Achievable Output (pcs/day)',
                            'b': result.before.achievable_output,
                            'a': result.method_a.achievable_output,
                            'p': result.method_b.achievable_output,
                            'fmt': 'int'
                        },
                        {
                            'title': 'Efficiency = Balancing Rate (%)',
                            'b': result.before.efficiency_balancing_rate,
                            'a': result.method_a.efficiency_balancing_rate,
                            'p': result.method_b.efficiency_balancing_rate,
                            'fmt': 'float'
                        },
                        {
                            'title': 'Balancing Delay (%)',
                            'b': result.before.comparison_balance_delay,
                            'a': result.method_a.comparison_balance_delay,
                            'p': result.method_b.comparison_balance_delay,
                            'fmt': 'float'
                        },
                        {
                            'title': 'Smoothing Index (sec)',
                            'b': result.before.smoothing_index_seconds,
                            'a': result.method_a.smoothing_index_seconds,
                            'p': result.method_b.smoothing_index_seconds,
                            'fmt': 'float'
                        },
                        {
                            'title': 'Labour Productivity (pcs/op/shift)',
                            'b': result.before.comparison_labour_productivity,
                            'a': result.method_a.comparison_labour_productivity,
                            'p': result.method_b.comparison_labour_productivity,
                            'fmt': 'float'
                        }
                    ] %}

                    {% for card in kpi_cards_data %}
                    {% set vals = [card.b, card.a, card.p] %}
                    {% set v_min = vals|min %}
                    {% set v_max = vals|max %}
                    {% set v_diff = v_max - v_min %}
                    {% if v_diff <= 0.0001 %}
                        {% set h_b = 65.0 %}
                        {% set h_a = 65.0 %}
                        {% set h_p = 65.0 %}
                    {% else %}
                        {% set h_b = (28.0 + (((card.b - v_min) / v_diff) * 65.0))|round(1) %}
                        {% set h_a = (28.0 + (((card.a - v_min) / v_diff) * 65.0))|round(1) %}
                        {% set h_p = (28.0 + (((card.p - v_min) / v_diff) * 65.0))|round(1) %}
                    {% endif %}
                    <div class="kpi-mini-card">
                        <div class="kpi-mini-title">{{ card.title }}</div>
                        <div class="kpi-mini-chart">
                            <div class="kpi-bars-area">
                                <!-- Before Bar -->
                                <div class="kpi-bar-col">
                                    <span class="kpi-bar-val">{% if card.fmt == 'int' %}{{ "%.0f"|format(card.b) }}{% else %}{{ "%.1f"|format(card.b) }}{% endif %}</span>
                                    <div class="kpi-bar bar-before" style="height: {{ h_b }}%;"></div>
                                </div>
                                <!-- Takt Bar -->
                                <div class="kpi-bar-col">
                                    <span class="kpi-bar-val">{% if card.fmt == 'int' %}{{ "%.0f"|format(card.a) }}{% else %}{{ "%.1f"|format(card.a) }}{% endif %}</span>
                                    <div class="kpi-bar bar-takt" style="height: {{ h_a }}%;"></div>
                                </div>
                                <!-- Pitch Bar -->
                                <div class="kpi-bar-col">
                                    <span class="kpi-bar-val">{% if card.fmt == 'int' %}{{ "%.0f"|format(card.p) }}{% else %}{{ "%.1f"|format(card.p) }}{% endif %}</span>
                                    <div class="kpi-bar bar-pitch" style="height: {{ h_p }}%;"></div>
                                </div>
                            </div>
                            <div class="kpi-baseline-axis">
                                <div class="kpi-axis-tick"></div>
                                <div class="kpi-axis-tick"></div>
                                <div class="kpi-axis-tick"></div>
                            </div>
                            <div class="kpi-axis-labels">
                                <span class="kpi-label">Before</span>
                                <span class="kpi-label">Takt</span>
                                <span class="kpi-label">Pitch</span>
                            </div>
                        </div>
                    </div>
                    {% endfor %}
                </div>
            </div>

            <!-- Profile Line Curve Container -->
            <div id="profileLinesSection" class="chart-container" style="display: none;">
                <canvas id="comparisonChart"></canvas>
            </div>
        </div>

        <!-- 5. Balanced Workstation Layout (Single View with Infinity Toggle Button) -->
        <div class="layout-control-card">
            <div class="section-header">
                <div class="section-title">
                    <span>Workstation Layout</span>
                </div>
            </div>
            <!-- Method A View -->
            <div id="methodViewA" class="method-layout-container">
                <div class="method-col-header">
                    <div class="method-header-left">
                        <!-- Switch Button on Left in front of Flag -->
                        <span class="method-badge badge-takt">Method A: After Takt Time</span>
                        <button type="button" class="btn-method-switch" onclick="setMethodLayout('B')" title="Switch to Method B">
                            <span>Switch to Method B (IE Pitch)</span>
                        </button>
                    </div>
                    <div class="method-desc">Strict Takt = <strong>{{ "%.1f"|format(result.takt_time) }}s</strong> · Zero relaxation · Strict divide-and-increment</div>
                </div>

                <!-- Method A Mini KPI Row -->
                <div class="mini-kpi-grid" style="margin-bottom: 16px;">
                    <div class="mini-kpi">
                        <div class="k-lbl">Manpower</div>
                        <div class="k-val" style="color: #22c55e;">{{ result.method_a.total_manpower }} <span style="font-size: 11px; font-weight: normal;">ops</span></div>
                    </div>
                    <div class="mini-kpi">
                        <div class="k-lbl">Stations</div>
                        <div class="k-val">{{ result.method_a.num_workstations }}</div>
                    </div>
                    <div class="mini-kpi">
                        <div class="k-lbl">Efficiency</div>
                        <div class="k-val" style="color: #22c55e;">{{ "%.1f"|format(result.method_a.efficiency_balancing_rate) }}%</div>
                    </div>
                    <div class="mini-kpi">
                        <div class="k-lbl">Output</div>
                        <div class="k-val">{{ "%.0f"|format(result.method_a.achievable_output) }} <span style="font-size: 10px; font-weight: normal;">pcs/day</span></div>
                    </div>
                </div>

                <!-- Method A Workstation Table -->
                <div class="table-scroll">
                    <table class="data-table">
                        <thead>
                            <tr>
                                <th>WS #</th>
                                <th>Serial/Id</th>
                                <th>Operations</th>
                                <th>Machine</th>
                                <th>Predecessor</th>
                                <th>Basic Time</th>
                                <th>Combined SAM</th>
                                <th>Balancing SAM</th>
                                <th>M/P</th>
                                <th>Takt Time</th>
                                <th>Status</th>
                            </tr>
                        </thead>
                        <tbody>
                            {% for r in result.method_a.rows %}
                            <tr>
                                <td><strong>{{ r['Composite Operations'] }}</strong></td>
                                <td>{{ r['Serial/Id'] }}</td>
                                <td>{{ r['Operations'] }}</td>
                                <td>{{ r['Machine'] }}</td>
                                <td>{{ r['Predecessor'] }}</td>
                                <td>{{ r['Basic Time'] }}</td>
                                <td style="font-weight: 600;">{{ r['Combined SAM'] }}</td>
                                <td style="font-weight: 700; color: #22c55e;">{{ r['Balancing SAM'] }}</td>
                                <td><strong>{{ r['M/P'] }}</strong></td>
                                <td style="color: #22c55e; font-weight: 600;">{{ r['Takt Time'] }}</td>
                                <td><span class="status-badge status-ok">{{ r['Status'] }}</span></td>
                            </tr>
                            {% endfor %}
                        </tbody>
                    </table>
                </div>
            </div>

            <!-- Method B View -->
            <div id="methodViewB" class="method-layout-container" style="display: none;">
                <div class="method-col-header">
                    <div class="method-header-left">
                        <!-- Switch Button on Left in front of Flag -->
                        <span class="method-badge badge-pitch">Method B: After IE Pitch</span>
                        <button type="button" class="btn-method-switch" onclick="setMethodLayout('A')" title="Switch to Method A">
                            <span>Switch to Method A (Takt Time)</span>
                        </button>
                    </div>
                    <div class="method-desc">Takt Merge = <strong>{{ "%.1f"|format(result.takt_time) }}s</strong> · Pitch = {{ "%.1f"|format(result.pitch_time) }}s · UCL = {{ "%.1f"|format(result.ucl) }}s · LCL = {{ "%.1f"|format(result.lcl) }}s</div>
                </div>

                <!-- Method B Mini KPI Row -->
                <div class="mini-kpi-grid" style="margin-bottom: 16px;">
                    <div class="mini-kpi">
                        <div class="k-lbl">Manpower</div>
                        <div class="k-val" style="color: #fb923c;">{{ result.method_b.total_manpower }} <span style="font-size: 11px; font-weight: normal;">ops</span></div>
                    </div>
                    <div class="mini-kpi">
                        <div class="k-lbl">Stations</div>
                        <div class="k-val">{{ result.method_b.num_workstations }}</div>
                    </div>
                    <div class="mini-kpi">
                        <div class="k-lbl">Efficiency</div>
                        <div class="k-val" style="color: #fb923c;">{{ "%.1f"|format(result.method_b.efficiency_balancing_rate) }}%</div>
                    </div>
                    <div class="mini-kpi">
                        <div class="k-lbl">Output</div>
                        <div class="k-val">{{ "%.0f"|format(result.method_b.achievable_output) }} <span style="font-size: 10px; font-weight: normal;">pcs/day</span></div>
                    </div>
                </div>

                <!-- Method B Workstation Table -->
                <div class="table-scroll">
                    <table class="data-table">
                        <thead>
                            <tr>
                                <th>WS #</th>
                                <th>Serial/Id</th>
                                <th>Operations</th>
                                <th>Machine</th>
                                <th>Predecessor</th>
                                <th>Basic Time</th>
                                <th>Combined SAM</th>
                                <th>Balancing SAM</th>
                                <th>M/P</th>
                                <th>Pitch Time</th>
                                <th>LCL</th>
                                <th>UCL</th>
                                <th>Status</th>
                            </tr>
                        </thead>
                        <tbody>
                            {% for r in result.method_b.rows %}
                            {% set is_flagged = 'Above UCL' in r['Status'] or 'review' in r['Status'] %}
                            <tr class="{% if is_flagged %}row-flagged{% endif %}">
                                <td><strong>{{ r['Composite Operations'] }}</strong></td>
                                <td>{{ r['Serial/Id'] }}</td>
                                <td>{{ r['Operations'] }}</td>
                                <td>{{ r['Machine'] }}</td>
                                <td>{{ r['Predecessor'] }}</td>
                                <td>{{ r['Basic Time'] }}</td>
                                <td style="font-weight: 600;">{{ r['Combined SAM'] }}</td>
                                <td style="font-weight: 700; color: #fb923c;">{{ r['Balancing SAM'] }}</td>
                                <td><strong>{{ r['M/P'] }}</strong></td>
                                <td style="color: #fb923c; font-weight: 600;">{{ r['Pitch Time'] }}</td>
                                <td style="color: var(--text-muted);">{{ r['LCL'] }}</td>
                                <td style="color: var(--text-muted);">{{ r['UCL'] }}</td>
                                <td>
                                    {% if is_flagged %}
                                        <span class="status-badge status-warning">{{ r['Status'] }}</span>
                                    {% elif 'OK' in r['Status'] %}
                                        <span class="status-badge status-ok">{{ r['Status'] }}</span>
                                    {% else %}
                                        <span class="status-badge status-danger">{{ r['Status'] }}</span>
                                    {% endif %}
                                </td>
                            </tr>
                            {% endfor %}
                        </tbody>
                    </table>
                </div>
            </div>
        </div>

        <!-- 6. Text Callout Card (Recommendations) -->
        {% if result.recommendations %}
        <div class="rec-card">
            <div class="section-header" style="margin-bottom: 8px;">
                <div class="section-title">
                    <span>Balancing Analysis & Recommendations</span>
                </div>
            </div>
            <div class="rec-list">
                {% for rec in result.recommendations %}
                <div class="rec-item">
                    <span class="rec-bullet">•</span>
                    <span>{{ rec }}</span>
                </div>
                {% endfor %}
            </div>
        </div>
        {% endif %}

        <!-- 7. Bottom Action Bar -->
        {% if session_id %}
        <div class="bottom-action-bar">
            <a href="/api/export/compare/xlsx/{{ session_id }}" class="btn-large-export">
                <span>Download Full Comparison Report (Excel .xlsx with Charts)</span>
            </a>
        </div>
        {% endif %}

        {% endif %}
    </div>

    <script>
        // Theme Toggle
        function toggleTheme() {
            const html = document.documentElement;
            const currentTheme = html.getAttribute('data-theme') || 'dark';
            const newTheme = currentTheme === 'dark' ? 'light' : 'dark';
            html.setAttribute('data-theme', newTheme);
            localStorage.setItem('theme', newTheme);
            document.querySelectorAll('.theme-toggle').forEach(button => {
                button.textContent = newTheme === 'dark' ? '☀️ Light' : '🌙 Dark';
            });
            if (window.currentChartData && chartMode === 'profile_lines') {
                renderProfileChart();
            }
            if (window.currentChartData) {
                renderOptOverviewCharts();
            }
        }

        // Workstation Layout Switcher
        let activeMethodLayout = 'A';

        function toggleMethodLayout() {
            if (activeMethodLayout === 'A') {
                setMethodLayout('B');
            } else {
                setMethodLayout('A');
            }
        }

        function setMethodLayout(method) {
            activeMethodLayout = method;
            const viewA = document.getElementById('methodViewA');
            const viewB = document.getElementById('methodViewB');
            const tabA = document.getElementById('tabBtnMethodA');
            const tabB = document.getElementById('tabBtnMethodB');
            const shiftText = document.getElementById('toggleShiftText');

            if (method === 'A') {
                if (viewA) viewA.style.display = 'block';
                if (viewB) viewB.style.display = 'none';
                if (tabA) tabA.classList.add('active');
                if (tabB) tabB.classList.remove('active');
                if (shiftText) shiftText.textContent = 'Switch to Method B (IE Pitch)';
            } else {
                if (viewA) viewA.style.display = 'none';
                if (viewB) viewB.style.display = 'block';
                if (tabB) tabB.classList.add('active');
                if (tabA) tabA.classList.remove('active');
                if (shiftText) shiftText.textContent = 'Switch to Method A (Takt Time)';
            }
        }

        let chartMode = 'grouped_bar';

        function switchChartMode(mode, btn) {
            chartMode = mode;
            document.querySelectorAll('.chart-tab-btn').forEach(b => b.classList.remove('active'));
            if (btn) btn.classList.add('active');

            const groupedSection = document.getElementById('groupedKpiSection');
            const profileSection = document.getElementById('profileLinesSection');

            if (mode === 'grouped_bar') {
                if (groupedSection) groupedSection.style.display = 'block';
                if (profileSection) profileSection.style.display = 'none';
            } else {
                if (groupedSection) groupedSection.style.display = 'none';
                if (profileSection) profileSection.style.display = 'block';
                renderProfileChart();
            }
        }

        function renderProfileChart() {
            if (!window.currentChartData) return;
            const data = window.currentChartData;
            const ctx = document.getElementById('comparisonChart');
            if (!ctx) return;

            const isDark = (document.documentElement.getAttribute('data-theme') || 'dark') === 'dark';
            const textColor = isDark ? '#8b9cb3' : '#64748b';
            const gridColor = isDark ? 'rgba(255, 255, 255, 0.08)' : 'rgba(0, 0, 0, 0.06)';

            if (window.compChart) {
                window.compChart.destroy();
            }

            // Balancing Line Profile Curves
            window.compChart = new Chart(ctx, {
                type: 'line',
                data: {
                    labels: data.labels,
                    datasets: [
                        {
                            label: 'Method A: Takt Time SAM',
                            data: data.method_a_times,
                            borderColor: '#22c55e',
                            backgroundColor: 'rgba(34, 197, 94, 0.15)',
                            tension: 0.2,
                            borderWidth: 2.5,
                            pointRadius: 4,
                            pointHoverRadius: 6,
                        },
                        {
                            label: 'Method B: IE Pitch SAM',
                            data: data.method_b_times,
                            borderColor: '#fb923c',
                            backgroundColor: 'rgba(251, 146, 60, 0.15)',
                            tension: 0.2,
                            borderWidth: 2.5,
                            pointRadius: 4,
                            pointHoverRadius: 6,
                        },
                        {
                            label: 'Takt Time Ceiling (' + data.takt_time.toFixed(1) + 's)',
                            data: Array(data.labels.length).fill(data.takt_time),
                            borderColor: '#ef4444',
                            borderWidth: 2,
                            borderDash: [6, 4],
                            pointRadius: 0,
                            fill: false,
                        },
                        {
                            label: 'Pitch UCL (' + data.ucl.toFixed(1) + 's)',
                            data: Array(data.labels.length).fill(data.ucl),
                            borderColor: '#f59e0b',
                            borderWidth: 1.5,
                            borderDash: [4, 4],
                            pointRadius: 0,
                            fill: false,
                        },
                        {
                            label: 'Pitch LCL (' + data.lcl.toFixed(1) + 's)',
                            data: Array(data.labels.length).fill(data.lcl),
                            borderColor: '#8b5cf6',
                            borderWidth: 1.5,
                            borderDash: [4, 4],
                            pointRadius: 0,
                            fill: false,
                        }
                    ]
                },
                options: {
                    responsive: true,
                    maintainAspectRatio: false,
                    interaction: { mode: 'index', intersect: false },
                    plugins: {
                        legend: {
                            labels: { color: textColor, font: { size: 12, weight: 600 } }
                        },
                        tooltip: {
                            callbacks: {
                                label: function(context) {
                                    return context.dataset.label + ': ' + context.parsed.y.toFixed(1) + 's';
                                }
                            }
                        }
                    },
                    scales: {
                        x: {
                            title: { display: true, text: 'Workstation Index', color: textColor },
                            ticks: { color: textColor },
                            grid: { color: gridColor }
                        },
                        y: {
                            title: { display: true, text: 'Balancing SAM (seconds)', color: textColor },
                            ticks: { color: textColor },
                            grid: { color: gridColor }
                        }
                    }
                }
            });
        }

        // ── Understanding Production Optimization Charts ──
        let optChartBefore = null, optChartA = null, optChartB = null;

        function renderOptOverviewCharts() {
            if (!window.currentChartData) return;
            const data = window.currentChartData;

            const isDark = (document.documentElement.getAttribute('data-theme') || 'dark') === 'dark';
            const textColor = isDark ? '#8b9cb3' : '#64748b';
            const gridColor = isDark ? 'rgba(255, 255, 255, 0.06)' : 'rgba(0, 0, 0, 0.06)';

            // Compute a shared Y max across all three charts
            const allVals = [
                ...(data.before_times || []),
                ...(data.method_a_times || []),
                ...(data.method_b_times || []),
                data.takt_time || 0,
                data.pitch_time || 0,
                data.ucl || 0,
            ];
            const globalMax = Math.max(...allVals);
            const yMax = Math.ceil(globalMax * 1.15);  // +15% headroom

            // Destroy old instances
            if (optChartBefore) { optChartBefore.destroy(); optChartBefore = null; }
            if (optChartA) { optChartA.destroy(); optChartA = null; }
            if (optChartB) { optChartB.destroy(); optChartB = null; }

            // Common bar dataset factory
            function makeBarDataset(label, values, color, bgColor) {
                return {
                    label: label,
                    data: values,
                    backgroundColor: bgColor,
                    borderColor: color,
                    borderWidth: 1,
                    borderRadius: 3,
                    barPercentage: 0.7,
                    categoryPercentage: 0.8,
                };
            }

            // Common options factory
            function makeOpts(xLabel) {
                return {
                    responsive: true,
                    maintainAspectRatio: false,
                    interaction: { mode: 'index', intersect: false },
                    plugins: {
                        legend: {
                            display: true,
                            position: 'top',
                            align: 'end',
                            labels: { color: textColor, font: { size: 11, weight: 600 }, boxWidth: 14, padding: 10 }
                        },
                        tooltip: {
                            callbacks: {
                                label: function(ctx) { return ctx.dataset.label + ': ' + ctx.parsed.y.toFixed(1) + 's'; }
                            }
                        }
                    },
                    scales: {
                        x: {
                            title: { display: true, text: xLabel, color: textColor, font: { size: 11 } },
                            ticks: { color: textColor, font: { size: 10 } },
                            grid: { display: false }
                        },
                        y: {
                            title: { display: true, text: 'Time (seconds)', color: textColor, font: { size: 11 } },
                            min: 0,
                            max: yMax,
                            ticks: { color: textColor, font: { size: 10 } },
                            grid: { color: gridColor }
                        }
                    }
                };
            }

            // Annotation line helper
            function lineAnnotation(value, color, label, dashArr) {
                return {
                    type: 'line',
                    yMin: value,
                    yMax: value,
                    borderColor: color,
                    borderWidth: 2,
                    borderDash: dashArr || [],
                    label: {
                        display: true,
                        content: label,
                        position: 'end',
                        backgroundColor: color,
                        color: '#fff',
                        font: { size: 10, weight: 700 },
                        padding: { x: 6, y: 3 },
                        borderRadius: 3,
                    }
                };
            }

            // ─── 1. Before Balancing Chart ───
            const ctxBefore = document.getElementById('optChartBefore');
            if (ctxBefore) {
                const beforeLabels = data.before_labels || data.before_times.map((_, i) => 'Op ' + (i + 1));
                const optsBefore = makeOpts('Operations (Before Balancing)');
                optsBefore.plugins.annotation = {
                    annotations: {
                        taktLine: lineAnnotation(data.takt_time, '#ef4444', 'Takt Time ' + data.takt_time.toFixed(1) + 's', [6, 4]),
                        pitchLine: lineAnnotation(data.pitch_time, '#3b82f6', 'IE Pitch ' + data.pitch_time.toFixed(1) + 's', [4, 4]),
                    }
                };
                optChartBefore = new Chart(ctxBefore, {
                    type: 'bar',
                    data: {
                        labels: beforeLabels,
                        datasets: [
                            makeBarDataset('Operation Time', data.before_times, '#3882bd', 'rgba(56, 130, 189, 0.75)'),
                        ]
                    },
                    options: optsBefore
                });
            }

            // ─── 2. Method A — Takt Time Chart ───
            const ctxA = document.getElementById('optChartMethodA');
            if (ctxA) {
                const labelsA = data.labels.slice(0, data.method_a_times.length);
                const optsA = makeOpts('Workstations (Method A)');
                optsA.plugins.annotation = {
                    annotations: {
                        taktLine: lineAnnotation(data.takt_time, '#ef4444', 'Takt Time ' + data.takt_time.toFixed(1) + 's', [6, 4]),
                    }
                };
                optChartA = new Chart(ctxA, {
                    type: 'bar',
                    data: {
                        labels: labelsA,
                        datasets: [
                            makeBarDataset('Balancing SAM', data.method_a_times, '#22c55e', 'rgba(34, 197, 94, 0.75)'),
                        ]
                    },
                    options: optsA
                });
            }

            // ─── 3. Method B — IE Pitch Chart ───
            const ctxB = document.getElementById('optChartMethodB');
            if (ctxB) {
                const labelsB = data.labels.slice(0, data.method_b_times.length);
                const optsB = makeOpts('Workstations (Method B)');
                optsB.plugins.annotation = {
                    annotations: {
                        taktLine: lineAnnotation(data.takt_time, '#ef4444', 'Takt Time ' + data.takt_time.toFixed(1) + 's', [6, 4]),
                        uclLine: lineAnnotation(data.ucl, '#f59e0b', 'UCL ' + data.ucl.toFixed(1) + 's', [4, 4]),
                    }
                };
                optChartB = new Chart(ctxB, {
                    type: 'bar',
                    data: {
                        labels: labelsB,
                        datasets: [
                            makeBarDataset('Balancing SAM', data.method_b_times, '#fb923c', 'rgba(251, 146, 60, 0.75)'),
                        ]
                    },
                    options: optsB
                });
            }
        }

        // Restore theme on load
        window.addEventListener('DOMContentLoaded', function() {
            const savedTheme = localStorage.getItem('theme') || 'dark';
            document.documentElement.setAttribute('data-theme', savedTheme);
            document.querySelectorAll('.theme-toggle').forEach(button => {
                button.textContent = savedTheme === 'dark' ? '☀️ Light' : '🌙 Dark';
            });

            {% if session_id %}
            // Load Chart Data
            fetch('/api/takt-vs-pitch-chart-data/{{ session_id }}')
                .then(res => res.json())
                .then(data => {
                    if (data.error) return;
                    window.currentChartData = data;
                    // Render the overview optimization charts
                    renderOptOverviewCharts();
                    if (chartMode === 'profile_lines') {
                        renderProfileChart();
                    }
                })
                .catch(err => console.error('Error fetching chart data:', err));
            {% endif %}
        });
    </script>
</body>
</html>
"""

if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0", port=5000)
