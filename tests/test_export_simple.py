#!/usr/bin/env python
"""
Simple test to verify the export endpoints work and have " + " separator.
"""
import requests
import json
from pathlib import Path

BASE_URL = "http://127.0.0.1:5000"
SAMPLE_FILE = Path("data/sample_operations.csv")

print("Testing export endpoints...\n")

# Step 1: Upload file
print("1. Uploading sample file...")
with open(SAMPLE_FILE, 'rb') as f:
    files = {'file': f}
    data = {
        'total_ops': '24',
        'tolerance': '0.15'
    }
    
    response = requests.post(f"{BASE_URL}/", files=files, data=data)

print(f"   Response status: {response.status_code}")
html_response = response.text

# Look for session_id in various places
import re

# Try to find session_id in a script tag or variable
session_id = None

# Look for: var session_id = "..."
match = re.search(r'var\s+session_id\s*=\s*["\']([^"\']+)["\']', html_response)
if match:
    session_id = match.group(1)
    print(f"   Found session_id in script: {session_id}")

# Look for: sessionId = "..."
if not session_id:
    match = re.search(r'sessionId\s*=\s*["\']([^"\']+)["\']', html_response)
    if match:
        session_id = match.group(1)
        print(f"   Found sessionId: {session_id}")

# Look for session_id in export button
if not session_id:
    match = re.search(r"exportFile\(['\"]csv['\"]\s*,\s*['\"]([^'\"]+)['\"]\)", html_response)
    if match:
        session_id = match.group(1)
        print(f"   Found session_id in exportFile call: {session_id}")

# Look for hidden form input
if not session_id:
    match = re.search(r'<input[^>]*name=["\']session_id["\'][^>]*value=["\']([^"\']+)["\']', html_response)
    if match:
        session_id = match.group(1)
        print(f"   Found session_id in hidden input: {session_id}")

# Check if " + " is in the HTML
if " + " in html_response:
    print("   ✅ Found \" + \" separator in HTML response")
else:
    print("   ❌ No \" + \" separator found in HTML response")

# If we found session_id, test CSV export
if session_id:
    print(f"\n2. Testing CSV export with session_id: {session_id}")
    csv_response = requests.get(f"{BASE_URL}/api/export/csv/{session_id}")
    print(f"   Response status: {csv_response.status_code}")
    
    if csv_response.status_code == 200:
        if " + " in csv_response.text:
            print("   ✅ CSV contains \" + \" separator")
            # Show first few lines
            lines = csv_response.text.split('\n')[:3]
            for line in lines:
                print(f"      {line}")
        else:
            print("   ❌ CSV missing \" + \" separator")
    else:
        print(f"   ❌ CSV export failed: {csv_response.text[:200]}")
    
    print(f"\n3. Testing Excel export with session_id: {session_id}")
    xlsx_response = requests.get(f"{BASE_URL}/api/export/xlsx/{session_id}")
    print(f"   Response status: {xlsx_response.status_code}")
    
    if xlsx_response.status_code == 200:
        print("   ✅ Excel export successful")
        
        # Try to read with openpyxl
        try:
            import tempfile
            from openpyxl import load_workbook
            
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
                tmp.write(xlsx_response.content)
                tmp_path = tmp.name
            
            wb = load_workbook(tmp_path)
            ws = wb.active
            
            # Check for " + " in cells
            found_plus = False
            for row in ws.iter_rows(max_row=3):
                for cell in row:
                    if cell.value and " + " in str(cell.value):
                        found_plus = True
                        print(f"   ✅ Excel contains \" + \" separator: {str(cell.value)[:60]}...")
                        break
                if found_plus:
                    break
            
            if not found_plus:
                print("   ❌ No \" + \" separator found in Excel")
                
            Path(tmp_path).unlink()
        except Exception as e:
            print(f"   ⚠️  Could not read Excel: {e}")
    else:
        print(f"   ❌ Excel export failed: {xlsx_response.status_code}")
else:
    print("\n❌ Could not extract session_id from response")
    print("\nSearching for any potential session identifiers in HTML...")
    
    # Try to find any 8-character hex-like strings
    matches = re.findall(r'\b[a-f0-9]{8}\b', html_response)
    if matches:
        print(f"   Found potential session IDs: {matches[:3]}")
