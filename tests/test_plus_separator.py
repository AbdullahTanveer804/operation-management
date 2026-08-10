#!/usr/bin/env python
"""
Test script to verify " + " separator is working in Flask app output and exports.
"""
import requests
import sys
from pathlib import Path
import tempfile

# Test the Flask app with sample data
BASE_URL = "http://127.0.0.1:5000"
SAMPLE_FILE = Path("data/sample_operations.csv")

def test_flask_ui():
    """Test Flask web UI returns " + " separator in table"""
    print("Testing Flask UI with " + " separator...")
    
    with open(SAMPLE_FILE, 'rb') as f:
        files = {'file': f}
        data = {
            'total_ops': '24',
            'tolerance': '0.15'
        }
        
        response = requests.post(f"{BASE_URL}/", files=files, data=data)
        
    if response.status_code == 200:
        # Check for " + " in the HTML response
        if " + " in response.text:
            print("✅ Flask UI correctly shows \" + \" separator in HTML")
            
            # Extract a sample line to show user
            if "1 + 2" in response.text:
                print("✅ Found example: '1 + 2' in HTML output")
            return True
        else:
            print("❌ Flask UI missing \" + \" separator")
            return False
    else:
        print(f"❌ Flask returned error: {response.status_code}")
        return False

def test_csv_export():
    """Test CSV export contains " + " separator"""
    print("\nTesting CSV export with \" + \" separator...")
    
    # First, do a calculation to get session_id
    with open(SAMPLE_FILE, 'rb') as f:
        files = {'file': f}
        data = {
            'total_ops': '24',
            'tolerance': '0.15'
        }
        
        response = requests.post(f"{BASE_URL}/", files=files, data=data)
    
    if response.status_code == 200 and "session_id" in response.text:
        # Extract session_id from response
        import re
        match = re.search(r'session_id\s*=\s*["\']([^"\']+)["\']', response.text)
        if match:
            session_id = match.group(1)
            print(f"  Session ID: {session_id}")
            
            # Download CSV
            csv_response = requests.get(f"{BASE_URL}/api/export/csv/{session_id}")
            
            if csv_response.status_code == 200:
                csv_content = csv_response.text
                if " + " in csv_content:
                    print("✅ CSV export correctly contains \" + \" separator")
                    
                    # Show sample
                    lines = csv_content.split('\n')
                    for line in lines[:5]:
                        if " + " in line:
                            print(f"   Sample: {line[:80]}...")
                            break
                    return True
                else:
                    print("❌ CSV export missing \" + \" separator")
                    return False
            else:
                print(f"❌ CSV export failed: {csv_response.status_code}")
                return False
        else:
            print("❌ Could not extract session_id from response")
            return False
    else:
        print("❌ Failed to get session for CSV export")
        return False

def test_excel_export():
    """Test Excel export contains " + " separator"""
    print("\nTesting Excel export with \" + \" separator...")
    
    # First, do a calculation to get session_id
    with open(SAMPLE_FILE, 'rb') as f:
        files = {'file': f}
        data = {
            'total_ops': '24',
            'tolerance': '0.15'
        }
        
        response = requests.post(f"{BASE_URL}/", files=files, data=data)
    
    if response.status_code == 200 and "session_id" in response.text:
        # Extract session_id from response
        import re
        match = re.search(r'session_id\s*=\s*["\']([^"\']+)["\']', response.text)
        if match:
            session_id = match.group(1)
            
            # Download Excel
            xlsx_response = requests.get(f"{BASE_URL}/api/export/xlsx/{session_id}")
            
            if xlsx_response.status_code == 200:
                # Save to temp file and read with openpyxl
                with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
                    tmp.write(xlsx_response.content)
                    tmp_path = tmp.name
                
                try:
                    from openpyxl import load_workbook
                    wb = load_workbook(tmp_path)
                    ws = wb.active
                    
                    # Check for " + " in cells
                    found_plus = False
                    for row in ws.iter_rows(max_row=5):
                        for cell in row:
                            if cell.value and " + " in str(cell.value):
                                found_plus = True
                                print(f"✅ Excel contains \" + \" separator")
                                print(f"   Sample: {str(cell.value)[:80]}...")
                                break
                        if found_plus:
                            break
                    
                    if not found_plus:
                        print("❌ Excel export missing \" + \" separator")
                        return False
                    
                    return True
                finally:
                    Path(tmp_path).unlink()
            else:
                print(f"❌ Excel export failed: {xlsx_response.status_code}")
                return False
        else:
            print("❌ Could not extract session_id from response")
            return False
    else:
        print("❌ Failed to get session for Excel export")
        return False

if __name__ == "__main__":
    print("=" * 60)
    print("Testing \" + \" Separator Implementation")
    print("=" * 60)
    
    results = []
    
    try:
        results.append(("Flask UI", test_flask_ui()))
        results.append(("CSV Export", test_csv_export()))
        results.append(("Excel Export", test_excel_export()))
    except Exception as e:
        print(f"\n❌ Error during testing: {e}")
        sys.exit(1)
    
    print("\n" + "=" * 60)
    print("Test Results Summary")
    print("=" * 60)
    
    all_passed = True
    for test_name, passed in results:
        status = "✅ PASS" if passed else "❌ FAIL"
        print(f"{status}: {test_name}")
        if not passed:
            all_passed = False
    
    print("=" * 60)
    
    sys.exit(0 if all_passed else 1)
