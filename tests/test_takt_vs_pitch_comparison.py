"""
Unit and Integration Tests for Takt vs Pitch Comparison Mode
"""

import io
import math
import pytest
from flask import Flask

from src.line_balancer.models import Operation, Workstation
from src.line_balancer.takt_pitch_comparison import (
    balance_method_a_takt,
    balance_method_b_pitch,
    calculate_smoothing_index_seconds,
    calculate_takt_vs_pitch_comparison,
)
from app import app


@pytest.fixture
def sample_operations():
    """Sample operation dataset with various times and machine types."""
    return [
        Operation(op_id=1, name="Op 1", predecessors=[], machine_type="SNLS", basic_time=20.0),
        Operation(op_id=2, name="Op 2", predecessors=[1], machine_type="SNLS", basic_time=25.0),
        Operation(op_id=3, name="Op 3", predecessors=[2], machine_type="Overlock", basic_time=50.0),
        Operation(op_id=4, name="Op 4", predecessors=[3], machine_type="By Hand", basic_time=15.0),
        Operation(op_id=5, name="Op 5", predecessors=[4], machine_type="Overlock", basic_time=70.0),
    ]


def test_method_a_takt_balancing():
    """
    Test Method A:
    - Ceiling is Takt Time.
    - Merges compatible ops up to Takt Time ceiling.
    - Splits ops exceeding Takt Time with strict mode.
    """
    ops = [
        # Op 1 & 2 can merge: 20 + 25 = 45 <= 60
        Operation(op_id=1, name="Op 1", predecessors=[], machine_type="SNLS", basic_time=20.0),
        Operation(op_id=2, name="Op 2", predecessors=[1], machine_type="SNLS", basic_time=25.0),
        # Op 3 is 75s > 60s -> splits to M/P=2 (37.5s <= 60s)
        Operation(op_id=3, name="Op 3", predecessors=[2], machine_type="Overlock", basic_time=75.0),
        # Op 4 is 30s <= 60s
        Operation(op_id=4, name="Op 4", predecessors=[3], machine_type="By Hand", basic_time=30.0),
    ]
    takt_time = 60.0

    ws_list = balance_method_a_takt(ops, takt_time)

    assert len(ws_list) == 3
    # WS 1: Op 1 + Op 2 combined (45s, M/P=1)
    assert len(ws_list[0].operations) == 2
    assert ws_list[0].combined_basic_time == 45.0
    assert ws_list[0].manpower == 1
    assert ws_list[0].balancing_sam == 45.0

    # WS 2: Op 3 split (75s -> M/P=2, 37.5s)
    assert len(ws_list[1].operations) == 1
    assert ws_list[1].manpower == 2
    assert ws_list[1].balancing_sam == 37.5

    # WS 3: Op 4 standalone (30s, M/P=1)
    assert len(ws_list[2].operations) == 1
    assert ws_list[2].manpower == 1
    assert ws_list[2].balancing_sam == 30.0


def test_method_b_pitch_balancing_and_flagging():
    """
    Test Method B:
    - Merging ceiling is TAKT TIME (not UCL).
    - Ops whose combined total > UCL can still merge if <= Takt Time.
    - Post-merge classification flags 'Above UCL — review' without force-splitting.
    - Ops > Takt Time are manpower split.
    """
    ops = [
        # Op 1 & 2: 25 + 25 = 50s.
        # If pitch_time = 35s, UCL = 40.25s, Takt = 60s:
        # Combined 50s > UCL (40.25s), but <= Takt (60s).
        # MUST MERGE, with status "Above UCL — review" and M/P=1!
        Operation(op_id=1, name="Op 1", predecessors=[], machine_type="SNLS", basic_time=25.0),
        Operation(op_id=2, name="Op 2", predecessors=[1], machine_type="SNLS", basic_time=25.0),
        # Op 3: 35s <= UCL (40.25s) -> status "OK", M/P=1
        Operation(op_id=3, name="Op 3", predecessors=[2], machine_type="Overlock", basic_time=35.0),
        # Op 4: 90s > Takt (60s) -> split to M/P=2 (45s <= 60s). Status "Above UCL — review" because 45s > 40.25s.
        Operation(op_id=4, name="Op 4", predecessors=[3], machine_type="Press", basic_time=90.0),
    ]

    pitch_time = 35.0
    ucl = 40.25
    lcl = 29.75
    takt_time = 60.0

    ws_list, statuses = balance_method_b_pitch(ops, pitch_time, ucl, lcl, takt_time)

    assert len(ws_list) == 3
    # WS 1: Combined Op 1 + Op 2 (50s, M/P=1). Total > UCL but <= Takt.
    assert len(ws_list[0].operations) == 2
    assert ws_list[0].combined_basic_time == 50.0
    assert ws_list[0].manpower == 1
    assert ws_list[0].balancing_sam == 50.0
    assert statuses[0] == "Above UCL — review"

    # WS 2: Op 3 (35s <= UCL)
    assert len(ws_list[1].operations) == 1
    assert ws_list[1].balancing_sam == 35.0
    assert statuses[1] == "OK"

    # WS 3: Op 4 (90s > Takt -> split to M/P=2, 45s)
    assert len(ws_list[2].operations) == 1
    assert ws_list[2].manpower == 2
    assert ws_list[2].balancing_sam == 45.0
    assert statuses[2] == "Above UCL — review"


def test_smoothing_index_seconds_per_operator_position():
    """
    Test Smoothing Index in seconds:
    - √[Σ(Cmax − Ti)²]
    - Computed per operator position (not per workstation, not in minutes).
    """
    # 2 workstations:
    # WS 1: time=40s, manpower=2 (2 positions with 40s)
    # WS 2: time=50s, manpower=1 (1 position with 50s)
    # Cmax = 60s
    station_times_and_mp = [(40.0, 2), (50.0, 1)]
    c_max = 60.0

    # diff WS 1: (60 - 40)^2 * 2 = 400 * 2 = 800
    # diff WS 2: (60 - 50)^2 * 1 = 100 * 1 = 100
    # sum = 900
    # sqrt(900) = 30.0 seconds
    si = calculate_smoothing_index_seconds(station_times_and_mp, c_max)
    assert si == pytest.approx(30.0)


def test_full_takt_vs_pitch_comparison_calculation(sample_operations):
    """
    Test full comparison function:
    - Returns before, method_a, method_b, comparison.
    - All 6 new KPIs computed for all three.
    - Comparison has 8 headline KPIs formatted side-by-side.
    """
    shift_time_minutes = 480.0  # 8 hours = 28,800 seconds
    production_target = 600     # Takt = 28,800 / 600 = 48.0 seconds

    result = calculate_takt_vs_pitch_comparison(
        sample_operations, shift_time_minutes, production_target
    )

    assert "before" in result
    assert "method_a" in result
    assert "method_b" in result
    assert "comparison" in result

    # Check Takt Time
    assert result["takt_time"] == pytest.approx(48.0)

    # Check 6 New KPIs in Before
    before = result["before"]
    assert "cycle_time" in before
    assert "achievable_output" in before
    assert "efficiency_balancing_rate" in before
    assert "comparison_balance_delay" in before
    assert "smoothing_index_seconds" in before
    assert "comparison_labour_productivity" in before

    # Check 6 New KPIs in Method A
    method_a = result["method_a"]
    assert method_a["cycle_time"] == pytest.approx(48.0)  # Fixed at Takt Time
    assert method_a["achievable_output"] == pytest.approx(600.0)
    assert method_a["efficiency_balancing_rate"] + method_a["comparison_balance_delay"] == pytest.approx(100.0)

    # Check 6 New KPIs in Method B
    method_b = result["method_b"]
    max_ws_b = max(ws.balancing_sam for ws in method_b["workstations"])
    assert method_b["cycle_time"] == pytest.approx(max_ws_b)  # Bottleneck SAM
    assert method_b["efficiency_balancing_rate"] + method_b["comparison_balance_delay"] == pytest.approx(100.0)

    # Check distinct metric names coexist with existing metrics without collision
    assert "line_balancing_rate" in method_a
    assert "efficiency_balancing_rate" in method_a
    assert "line_efficiency" in method_a

    # Check Comparison table has 8 headline KPIs
    comparison = result["comparison"]
    assert len(comparison) == 8
    metric_keys = [c["key"] for c in comparison]
    assert "total_manpower" in metric_keys
    assert "num_workstations" in metric_keys
    assert "cycle_time" in metric_keys
    assert "achievable_output" in metric_keys
    assert "efficiency_balancing_rate" in metric_keys
    assert "comparison_balance_delay" in metric_keys
    assert "smoothing_index_seconds" in metric_keys
    assert "comparison_labour_productivity" in metric_keys


def test_takt_vs_pitch_flask_routes():
    """
    Test Flask endpoints for Takt vs Pitch Comparison:
    - GET /compare & GET /takt-vs-pitch
    - POST /compare
    - POST /api/takt-vs-pitch & POST /api/compare
    - GET /api/takt-vs-pitch-chart-data/<session_id>
    - GET /api/export/compare/xlsx/<session_id>
    """
    from openpyxl import load_workbook

    client = app.test_client()

    # 1. GET /compare & /takt-vs-pitch
    res_get = client.get("/compare")
    assert res_get.status_code == 200
    assert b"Takt vs Pitch Comparison" in res_get.data

    res_get_alt = client.get("/takt-vs-pitch")
    assert res_get_alt.status_code == 200

    # 2. POST /compare with CSV
    csv_content = """Serial No.,Operation Name,Predecessor,Machine Type,Basic Time
1,Op 1,-,SNLS,20.0
2,Op 2,1,SNLS,25.0
3,Op 3,2,Overlock,40.0
4,Op 4,3,By Hand,15.0
"""
    data = {
        "file": (io.BytesIO(csv_content.encode("utf-8")), "test_ops.csv"),
        "shift_time": "420",
        "production_target": "500",
    }
    res_post = client.post(
        "/compare", data=data, content_type="multipart/form-data"
    )
    assert res_post.status_code == 200
    assert b"Master 8-KPI Side-by-Side Comparison" in res_post.data
    assert b"Method A: After Takt Time" in res_post.data
    assert b"Method B: After IE Pitch" in res_post.data
    assert b"Visual Analysis &amp; Comparison Curves" in res_post.data or b"Visual Analysis & Comparison Curves" in res_post.data

    # 3. POST /api/compare
    data_api = {
        "file": (io.BytesIO(csv_content.encode("utf-8")), "test_ops.csv"),
        "shift_time": "420",
        "production_target": "500",
    }
    res_api = client.post(
        "/api/compare", data=data_api, content_type="multipart/form-data"
    )
    assert res_api.status_code == 200
    json_data = res_api.get_json()
    assert "session_id" in json_data
    assert "before" in json_data
    assert "method_a" in json_data
    assert "method_b" in json_data
    assert "comparison" in json_data
    assert len(json_data["comparison"]) == 8
    assert "recommendations" in json_data
    assert len(json_data["recommendations"]) >= 2

    session_id = json_data["session_id"]

    # 4. GET /api/takt-vs-pitch-chart-data/<session_id>
    res_chart = client.get(f"/api/takt-vs-pitch-chart-data/{session_id}")
    assert res_chart.status_code == 200
    chart_json = res_chart.get_json()
    assert "method_a_times" in chart_json
    assert "method_b_times" in chart_json
    assert "takt_time" in chart_json
    assert "kpi_labels" in chart_json
    assert len(chart_json["kpi_labels"]) == 8
    assert "kpi_before" in chart_json
    assert "kpi_method_a" in chart_json
    assert "kpi_method_b" in chart_json

    # 5. GET /api/export/compare/xlsx/<session_id>
    res_export = client.get(f"/api/export/compare/xlsx/{session_id}")
    assert res_export.status_code == 200
    assert res_export.mimetype == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    
    # Verify Excel workbook contents
    excel_bytes = io.BytesIO(res_export.data)
    wb = load_workbook(excel_bytes)
    sheet_names = wb.sheetnames
    assert "Comparison Summary" in sheet_names
    assert "Method A - Takt Time" in sheet_names
    assert "Method B - IE Pitch" in sheet_names
    assert "Before Balancing" in sheet_names



def test_recommendations_generation_method_a_wins():
    """Test recommendation text generation when Method A wins on operators and efficiency."""
    from src.line_balancer.comparison_recommendations import generate_takt_vs_pitch_recommendations

    mock_data = {
        "before": {"total_manpower": 20},
        "method_a": {
            "total_manpower": 12,
            "efficiency_balancing_rate": 88.5,
            "comparison_balance_delay": 11.5,
            "rows": [
                {"Composite Operations": 1, "Serial/Id": "1", "Operations": "Hem", "M/P": 2, "Status": "OK"},
                {"Composite Operations": 2, "Serial/Id": "2", "Operations": "Join", "M/P": 1, "Status": "OK"},
            ]
        },
        "method_b": {
            "total_manpower": 15,
            "efficiency_balancing_rate": 81.2,
            "comparison_balance_delay": 18.8,
            "rows": [
                {"Composite Operations": 1, "Serial/Id": "1", "Operations": "Hem", "M/P": 1, "Status": "Above UCL — review"},
                {"Composite Operations": 2, "Serial/Id": "2", "Operations": "Join", "M/P": 1, "Status": "OK"},
            ]
        }
    }

    sentences = generate_takt_vs_pitch_recommendations(mock_data)

    assert isinstance(sentences, list)
    assert len(sentences) >= 3
    # Sentence 1: Operator count
    assert "Method A requires 12 operators compared to 15 for Method B (3 fewer operators, saving 8 operators vs baseline (20))." in sentences[0]
    # Sentence 2: Efficiency
    assert "Method A achieves higher line efficiency at 88.5% (balancing delay of 11.5%) compared to Method B at 81.2% (balancing delay of 18.8%)." in sentences[1]
    # Sentence 3: Stations to flag
    assert "above UCL" in sentences[2] and "manpower splitting" in sentences[2]
    # Sentence 4: Overall
    assert "Method A is clearly preferable" in sentences[3]


def test_recommendations_generation_method_b_wins():
    """Test recommendation text generation when Method B wins on operators and efficiency."""
    from src.line_balancer.comparison_recommendations import generate_takt_vs_pitch_recommendations

    mock_data = {
        "before": {"total_manpower": 18},
        "method_a": {
            "total_manpower": 14,
            "efficiency_balancing_rate": 78.0,
            "comparison_balance_delay": 22.0,
            "rows": [
                {"Composite Operations": 1, "Serial/Id": "1", "Operations": "Sew", "M/P": 1, "Status": "OK"},
            ]
        },
        "method_b": {
            "total_manpower": 11,
            "efficiency_balancing_rate": 89.0,
            "comparison_balance_delay": 11.0,
            "rows": [
                {"Composite Operations": 1, "Serial/Id": "1", "Operations": "Sew", "M/P": 1, "Status": "OK"},
            ]
        }
    }

    sentences = generate_takt_vs_pitch_recommendations(mock_data)

    assert "Method B requires 11 operators compared to 14 for Method A (3 fewer operators, saving 7 operators vs baseline (18))." in sentences[0]
    assert "Method B achieves higher line efficiency at 89.0% (balancing delay of 11.0%) compared to Method A at 78.0% (balancing delay of 22.0%)." in sentences[1]
    assert "No operations required review in Method B, and no operations required manpower splitting in Method A." in sentences[2]
    assert "Method B is clearly preferable" in sentences[3]


def test_recommendations_generation_tie_and_tradeoff():
    """Test recommendation text generation on ties and trade-off situations."""
    from src.line_balancer.comparison_recommendations import generate_takt_vs_pitch_recommendations

    # Tie scenario
    tie_data = {
        "before": {"total_manpower": 16},
        "method_a": {
            "total_manpower": 10,
            "efficiency_balancing_rate": 85.0,
            "comparison_balance_delay": 15.0,
            "rows": []
        },
        "method_b": {
            "total_manpower": 10,
            "efficiency_balancing_rate": 85.0,
            "comparison_balance_delay": 15.0,
            "rows": []
        }
    }
    sentences_tie = generate_takt_vs_pitch_recommendations(tie_data)
    assert "Both Method A and Method B require an identical 10 operators, saving 6 operators vs baseline (16)." in sentences_tie[0]
    assert "Both methods deliver comparable line efficiency at 85.0% vs 85.0%" in sentences_tie[1]
    assert "closely matched results" in sentences_tie[3]

    # Trade-off scenario: Method A has fewer operators, Method B has higher efficiency
    tradeoff_data = {
        "before": {"total_manpower": 16},
        "method_a": {
            "total_manpower": 10,
            "efficiency_balancing_rate": 80.0,
            "comparison_balance_delay": 20.0,
            "rows": []
        },
        "method_b": {
            "total_manpower": 12,
            "efficiency_balancing_rate": 90.0,
            "comparison_balance_delay": 10.0,
            "rows": []
        }
    }
    sentences_tradeoff = generate_takt_vs_pitch_recommendations(tradeoff_data)
    assert "trade-off" in sentences_tradeoff[3]

