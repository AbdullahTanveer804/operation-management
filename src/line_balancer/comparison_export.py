"""
Excel Exporter for Takt vs Pitch Comparison Mode

Produces a professional multi-tab Excel workbook containing:
1. KPI Comparison & Summary (with 8-KPI comparison table, recommendations, and embedded matplotlib charts)
2. Method A (Takt Time) Workstation Table
3. Method B (IE Pitch) Workstation Table (with amber review flag highlights)
4. Before Balancing Operations Table
"""

import io
from typing import Dict, List, Optional
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from .models import Operation, Workstation


def style_range(ws, cell_range, border=None, alignment=None, fill=None):
    """Apply styling across an entire cell range (e.g. merged title banner cells)."""
    for row in ws[cell_range]:
        for cell in row:
            if border:
                cell.border = border
            if alignment:
                cell.alignment = alignment
            if fill:
                cell.fill = fill


def autofit_columns(ws, start_row=3, padding=3, min_width=6):
    """Auto-fit column widths like Alt+H+O+I in Excel."""
    for col in ws.columns:
        cells_to_check = [
            c for c in col if c.row >= start_row and c.value is not None
        ]
        if not cells_to_check:
            continue
        max_len = max(len(str(c.value)) for c in cells_to_check)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + padding,
                                                     min_width)


def generate_balancing_overview_chart(calc: Dict) -> io.BytesIO:
    """
    Generate 2x2 'Before vs After — Balancing Comparison Charts' layout:
    - Top Left: Chart Reading Guide with metadata & reference line keys
    - Top Right: Before Balancing (Operation Times + Takt + IE Pitch lines)
    - Bottom Left: Method A — Takt Time Balancing (Balancing SAM + Takt line)
    - Bottom Right: Method B — IE Pitch Time Balancing (Balancing SAM + Takt + UCL lines)
    All three bar charts share the exact same Y-axis scale.
    """
    fig = plt.figure(figsize=(15, 8.5), dpi=160, facecolor='#ffffff')
    gs = fig.add_gridspec(2, 2, hspace=0.35, wspace=0.22, left=0.06, right=0.96, top=0.88, bottom=0.08)

    takt_time = float(calc.get('takt_time', 0.0))
    pitch_time = float(calc.get('pitch_time', 0.0))
    ucl = float(calc.get('ucl', 0.0))

    ops = calc.get('sorted_operations', [])
    before_times = [float(getattr(op, 'basic_time', 0.0)) for op in ops]
    ws_a = calc.get('method_a', {}).get('workstations', [])
    times_a = [float(ws.balancing_sam) for ws in ws_a]
    ws_b = calc.get('method_b', {}).get('workstations', [])
    times_b = [float(ws.balancing_sam) for ws in ws_b]

    # Compute shared Y-axis max with clean step
    all_vals = before_times + times_a + times_b + [takt_time, pitch_time, ucl]
    global_max = max(all_vals) if all_vals else 10.0
    grid_step = 10 if global_max <= 100 else 20
    y_max = float(np.ceil((global_max * 1.15) / grid_step) * grid_step)

    # -------------------------------------------------------------------------
    # Cell 1 (Top-Left): Chart Reading Guide Panel
    # -------------------------------------------------------------------------
    ax_guide = fig.add_subplot(gs[0, 0])
    ax_guide.set_facecolor('#f8fafc')
    for spine in ax_guide.spines.values():
        spine.set_color('#e2e8f0')
        spine.set_linewidth(1.2)
    ax_guide.set_xticks([])
    ax_guide.set_yticks([])

    ax_guide.text(0.06, 0.90, "Chart Reading Guide", fontsize=12, fontweight='bold', color='#1e293b', va='top')
    
    guide_items = [
        ("Legend:", "Bars represent operator stations / workstations and their assigned time"),
        ("Axes:", "X-axis = Stations · Y-axis = Time (seconds) — same scale across all charts"),
        ("Takt Time (red line):", f"{takt_time:.1f}sec — Maximum time per station to meet customer demand. Shown on all three charts"),
        ("IE Pitch Time (yellow line):", f"{pitch_time:.1f}sec — Industrial Engineering standard target time. Shown on Before Balancing chart only"),
        ("UCL (blue line):", f"{ucl:.1f}sec — Pitch + 15% upper tolerance. Shown on Method B chart only"),
        ("Primary Goal:", "Bars closer to the reference line = better balanced line with less idle time")
    ]
    
    y_pos = 0.74
    for title, desc in guide_items:
        ax_guide.text(0.06, y_pos, f"• {title} ", fontsize=8.5, fontweight='bold', color='#334155', va='top')
        ax_guide.text(0.06, y_pos - 0.052, f"   {desc}", fontsize=7.8, color='#64748b', va='top')
        y_pos -= 0.122

    # -------------------------------------------------------------------------
    # Cell 2 (Top-Right): BEFORE BALANCING
    # -------------------------------------------------------------------------
    ax_before = fig.add_subplot(gs[0, 1])
    ax_before.set_facecolor('#ffffff')
    x_b = np.arange(len(before_times))
    labels_b = [f"Op {i+1}" for i in range(len(before_times))]
    ax_before.bar(x_b, before_times, color='#3882bd', width=0.65, label='Operation Time', zorder=2)
    ax_before.axhline(y=takt_time, color='#ef4444', linestyle='-', linewidth=2.0, label=f'Takt Time ({takt_time:.1f}s)', zorder=3)
    ax_before.axhline(y=pitch_time, color='#eab308', linestyle='-', linewidth=2.0, label=f'IE Pitch ({pitch_time:.1f}s)', zorder=3)
    
    ax_before.set_title("BEFORE BALANCING", fontsize=11, fontweight='bold', loc='left', color='#0f172a', pad=10)
    ax_before.set_xlabel("Operations (Before Balancing)", fontsize=9, color='#64748b', labelpad=5)
    ax_before.set_ylabel("Time (seconds)", fontsize=9, color='#64748b')
    ax_before.set_ylim(0, y_max)
    ax_before.set_xticks(x_b)
    ax_before.set_xticklabels(labels_b, fontsize=7.5, rotation=45, ha='right')
    ax_before.tick_params(axis='both', which='both', labelsize=8, colors='#64748b')
    ax_before.grid(True, linestyle='--', alpha=0.3, axis='y', zorder=1)
    ax_before.legend(loc='upper right', fontsize=8, frameon=False, ncol=3, handlelength=1.2)
    for s in ['top', 'right']:
        ax_before.spines[s].set_visible(False)
    for s in ['left', 'bottom']:
        ax_before.spines[s].set_color('#cbd5e1')

    # -------------------------------------------------------------------------
    # Cell 3 (Bottom-Left): METHOD A — TAKT TIME BALANCING
    # -------------------------------------------------------------------------
    ax_a = fig.add_subplot(gs[1, 0])
    ax_a.set_facecolor('#ffffff')
    x_a = np.arange(len(times_a))
    labels_a = [f"WS {i+1}" for i in range(len(times_a))]
    ax_a.bar(x_a, times_a, color='#22c55e', width=0.65, label='Balancing SAM', zorder=2)
    ax_a.axhline(y=takt_time, color='#ef4444', linestyle='-', linewidth=2.0, label=f'Takt Time ({takt_time:.1f}s)', zorder=3)
    
    ax_a.set_title("METHOD A — TAKT TIME BALANCING", fontsize=11, fontweight='bold', loc='left', color='#0f172a', pad=10)
    ax_a.set_xlabel("Workstations (Method A)", fontsize=9, color='#64748b', labelpad=5)
    ax_a.set_ylabel("Time (seconds)", fontsize=9, color='#64748b')
    ax_a.set_ylim(0, y_max)
    ax_a.set_xticks(x_a)
    ax_a.set_xticklabels(labels_a, fontsize=7.5, rotation=45, ha='right')
    ax_a.tick_params(axis='both', which='both', labelsize=8, colors='#64748b')
    ax_a.grid(True, linestyle='--', alpha=0.3, axis='y', zorder=1)
    ax_a.legend(loc='upper right', fontsize=8, frameon=False, ncol=2, handlelength=1.2)
    for s in ['top', 'right']:
        ax_a.spines[s].set_visible(False)
    for s in ['left', 'bottom']:
        ax_a.spines[s].set_color('#cbd5e1')

    # -------------------------------------------------------------------------
    # Cell 4 (Bottom-Right): METHOD B — IE PITCH TIME BALANCING
    # -------------------------------------------------------------------------
    ax_b = fig.add_subplot(gs[1, 1])
    ax_b.set_facecolor('#ffffff')
    x_b_arr = np.arange(len(times_b))
    labels_b_arr = [f"WS {i+1}" for i in range(len(times_b))]
    ax_b.bar(x_b_arr, times_b, color='#fb923c', width=0.65, label='Balancing SAM', zorder=2)
    ax_b.axhline(y=takt_time, color='#ef4444', linestyle='-', linewidth=2.0, label=f'Takt Time ({takt_time:.1f}s)', zorder=3)
    ax_b.axhline(y=ucl, color='#3b82f6', linestyle='-', linewidth=2.0, label=f'UCL ({ucl:.1f}s)', zorder=3)
    
    ax_b.set_title("METHOD B — IE PITCH TIME BALANCING", fontsize=11, fontweight='bold', loc='left', color='#0f172a', pad=10)
    ax_b.set_xlabel("Workstations (Method B)", fontsize=9, color='#64748b', labelpad=5)
    ax_b.set_ylabel("Time (seconds)", fontsize=9, color='#64748b')
    ax_b.set_ylim(0, y_max)
    ax_b.set_xticks(x_b_arr)
    ax_b.set_xticklabels(labels_b_arr, fontsize=7.5, rotation=45, ha='right')
    ax_b.tick_params(axis='both', which='both', labelsize=8, colors='#64748b')
    ax_b.grid(True, linestyle='--', alpha=0.3, axis='y', zorder=1)
    ax_b.legend(loc='upper right', fontsize=8, frameon=False, ncol=3, handlelength=1.2)
    for s in ['top', 'right']:
        ax_b.spines[s].set_visible(False)
    for s in ['left', 'bottom']:
        ax_b.spines[s].set_color('#cbd5e1')

    # Suptitle
    fig.suptitle("Before vs After — Balancing Comparison Charts", fontsize=14, fontweight='bold', color='#0f172a', x=0.06, y=0.96, ha='left')
    fig.text(0.06, 0.925, "Side-by-side time distribution across Before Balancing, Method A (Takt Time) and Method B (IE Pitch) — all charts share the same Y-axis scale for accurate visual comparison", fontsize=9, color='#64748b', ha='left')

    buf = io.BytesIO()
    plt.savefig(buf, format='png', bbox_inches='tight', facecolor=fig.get_facecolor(), dpi=160)
    plt.close(fig)
    buf.seek(0)
    return buf


def generate_all_kpis_card_grid_chart(calc: Dict) -> io.BytesIO:
    """
    Generate 2x4 grid containing all 8 KPI comparison mini-bar charts:
    1. Operations (after merging)
    2. Number of Operators
    3. Cycle Time (sec)
    4. Achievable Output (pcs/available time)
    5. Efficiency = Balancing Rate (%)
    6. Balancing Delay (%)
    7. Smoothing Index (sec)
    8. Labour Productivity (pcs/optr/shift)
    Each card displays values directly above each bar (Before: Red, Takt: Green, Pitch: Orange).
    """
    fig = plt.figure(figsize=(15, 7.5), dpi=160, facecolor='#ffffff')
    gs = fig.add_gridspec(2, 4, hspace=0.38, wspace=0.25, left=0.04, right=0.97, top=0.86, bottom=0.08)

    before = calc.get('before', {})
    method_a = calc.get('method_a', {})
    method_b = calc.get('method_b', {})

    ops = calc.get('sorted_operations', [])
    ws_a = method_a.get('workstations', [])
    ws_b = method_b.get('workstations', [])

    kpi_definitions = [
        {
            "title": "Operations (after merging)",
            "vals": [
                before.get('num_operations', len(ops)),
                method_a.get('num_workstations', len(ws_a)),
                method_b.get('num_workstations', len(ws_b))
            ],
            "fmt": "{:.0f}"
        },
        {
            "title": "Number of Operators",
            "vals": [
                before.get('total_manpower', len(ops)),
                method_a.get('total_manpower', 0),
                method_b.get('total_manpower', 0)
            ],
            "fmt": "{:.0f}"
        },
        {
            "title": "Cycle Time (sec)",
            "vals": [
                before.get('cycle_time', 0.0),
                method_a.get('cycle_time', 0.0),
                method_b.get('cycle_time', 0.0)
            ],
            "fmt": "{:.1f}"
        },
        {
            "title": "Achievable Output (pcs/available time)",
            "vals": [
                before.get('achievable_output', 0.0),
                method_a.get('achievable_output', 0.0),
                method_b.get('achievable_output', 0.0)
            ],
            "fmt": "{:.0f}"
        },
        {
            "title": "Efficiency = Balancing Rate (%)",
            "vals": [
                before.get('efficiency_balancing_rate', 0.0),
                method_a.get('efficiency_balancing_rate', 0.0),
                method_b.get('efficiency_balancing_rate', 0.0)
            ],
            "fmt": "{:.1f}"
        },
        {
            "title": "Balancing Delay (%)",
            "vals": [
                before.get('comparison_balance_delay', 0.0),
                method_a.get('comparison_balance_delay', 0.0),
                method_b.get('comparison_balance_delay', 0.0)
            ],
            "fmt": "{:.1f}"
        },
        {
            "title": "Smoothing Index (sec)",
            "vals": [
                before.get('smoothing_index_seconds', 0.0),
                method_a.get('smoothing_index_seconds', 0.0),
                method_b.get('smoothing_index_seconds', 0.0)
            ],
            "fmt": "{:.1f}"
        },
        {
            "title": "Labour Productivity (pcs/optr/shift)",
            "vals": [
                before.get('comparison_labour_productivity', 0.0),
                method_a.get('comparison_labour_productivity', 0.0),
                method_b.get('comparison_labour_productivity', 0.0)
            ],
            "fmt": "{:.1f}"
        }
    ]

    colors = ['#ef4444', '#22c55e', '#fb923c']
    x_pos = [0, 1, 2]
    x_labels = ['Before', 'Takt', 'Pitch']

    for idx, kpi in enumerate(kpi_definitions):
        r, c = divmod(idx, 4)
        ax = fig.add_subplot(gs[r, c])
        ax.set_facecolor('#ffffff')

        for spine in ax.spines.values():
            spine.set_color('#e2e8f0')
            spine.set_linewidth(1.0)

        vals = [float(v) for v in kpi["vals"]]
        max_val = max(vals) if max(vals) > 0 else 1.0

        bars = ax.bar(x_pos, vals, color=colors, width=0.38, zorder=2)
        
        # Value labels above bars
        for bar, val in zip(bars, vals):
            label_text = kpi["fmt"].format(val)
            ax.text(
                bar.get_x() + bar.get_width() / 2.0,
                bar.get_height() + (max_val * 0.04),
                label_text,
                ha='center',
                va='bottom',
                fontsize=8.5,
                fontweight='bold',
                color='#1e293b'
            )

        ax.set_title(kpi["title"], fontsize=9.5, fontweight='bold', color='#1e40af', pad=8, loc='left')
        ax.set_xticks(x_pos)
        ax.set_xticklabels(x_labels, fontsize=8.0, color='#64748b')
        ax.set_ylim(0, max_val * 1.30)
        ax.set_yticks([])
        ax.tick_params(axis='both', which='both', length=0)
        ax.grid(False)

    # Suptitle & Legend
    fig.suptitle("All KPIs — Before vs. After, at a Glance", fontsize=14, fontweight='bold', color='#0f172a', x=0.04, y=0.96, ha='left')
    fig.text(0.04, 0.925, "Every KPI from the master table, charted side by side across the three scenarios.", fontsize=9, color='#64748b', ha='left')

    legend_handles = [
        plt.Rectangle((0, 0), 1, 1, color='#ef4444', label='Before'),
        plt.Rectangle((0, 0), 1, 1, color='#22c55e', label='After – Takt'),
        plt.Rectangle((0, 0), 1, 1, color='#fb923c', label='After – IE Pitch')
    ]
    fig.legend(handles=legend_handles, loc='upper right', bbox_to_anchor=(0.97, 0.96), ncol=3, fontsize=8.5, frameon=False, handlelength=1.2, handleheight=1.2)

    buf = io.BytesIO()
    plt.savefig(buf, format='png', bbox_inches='tight', facecolor=fig.get_facecolor(), dpi=160)
    plt.close(fig)
    buf.seek(0)
    return buf


def generate_comparison_excel(calc: Dict) -> io.BytesIO:
    """
    Build a multi-tab Excel workbook with complete side-by-side comparison tables,
    detailed workstation views for Method A and Method B, baseline data, and embedded charts.
    """
    wb = Workbook()

    # Define standard styles
    font_title = Font(size=15, bold=True, color="1E40AF")
    font_section = Font(size=12, bold=True, color="1E293B")
    font_bold = Font(size=10, bold=True)
    font_regular = Font(size=10)
    font_header = Font(size=10, bold=True, color="FFFFFF")
    font_winner = Font(size=10, bold=True, color="166534")

    fill_header = PatternFill(start_color="2563EB",
                              end_color="2563EB",
                              fill_type="solid")
    fill_sub_a = PatternFill(start_color="3B82F6",
                             end_color="3B82F6",
                             fill_type="solid")
    fill_sub_b = PatternFill(start_color="10B981",
                             end_color="10B981",
                             fill_type="solid")
    fill_zebra = PatternFill(start_color="F8FAFC",
                             end_color="F8FAFC",
                             fill_type="solid")
    fill_winner = PatternFill(start_color="DCFCE7",
                              end_color="DCFCE7",
                              fill_type="solid")
    fill_warning = PatternFill(start_color="FEF3C7",
                               end_color="FEF3C7",
                               fill_type="solid")

    # True solid black border matching Excel's Alt+H+B+A (All Borders)
    thin_border_side = Side(border_style="thin", color="000000")
    cell_border = Border(left=thin_border_side,
                         right=thin_border_side,
                         top=thin_border_side,
                         bottom=thin_border_side)

    # =========================================================================
    # TAB 1: KPI Comparison & Summary
    # =========================================================================
    ws1 = wb.active
    ws1.title = "Comparison Summary"
    ws1.views.sheetView[0].showGridLines = True

    # Title Banner (Alt+H+B+A & Alt+H+A+C)
    ws1['A1'] = "Takt vs Pitch Balancing Comparison Report"
    ws1['A1'].font = font_title
    ws1.merge_cells('A1:E2')
    ws1['A1'].alignment = Alignment(horizontal="center", vertical="center")
    style_range(ws1, 'A1:E2', border=cell_border)
    ws1.row_dimensions[1].height = 30

    # Summary Metadata
    curr_row = 3
    meta_items = [
        ("Customer Demand", f"{calc['production_target']} units"),
        ("Available Shift Time", f"{calc['shift_time_minutes']:.1f} minutes"),
        ("Total Basic Time (SAM)",
         f"{calc['total_sam'] / 60:.2f} min ({calc['total_sam']:.1f} s)"),
        ("Takt Time (Method A Ceiling)", f"{calc['takt_time']:.2f} seconds"),
        ("IE Pitch Time (Method B Base)", f"{calc['pitch_time']:.2f} seconds"),
        ("Method B UCL / LCL (±15%)",
         f"{calc['ucl']:.2f} s / {calc['lcl']:.2f} s"),
    ]
    for label, val in meta_items:
        c_a = ws1[f'A{curr_row}']
        c_a.value = label
        c_a.font = font_bold
        c_a.border = cell_border
        c_b = ws1[f'B{curr_row}']
        c_b.value = val
        c_b.font = font_regular
        c_b.border = cell_border
        curr_row += 1

    curr_row += 1

    # Master Headline Comparison Table
    ws1[f'A{curr_row}'] = "Headline 8-KPI Side-by-Side Comparison"
    ws1[f'A{curr_row}'].font = font_section
    curr_row += 1

    headers = [
        "Metric Name", "Unit", "Before Balancing (Baseline)",
        "Method A — Takt Time", "Method B — IE Pitch"
    ]
    for col_idx, h in enumerate(headers, 1):
        cell = ws1.cell(row=curr_row, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(
            horizontal="center" if col_idx > 1 else "left", vertical="center")
        cell.border = cell_border
    ws1.row_dimensions[curr_row].height = 24
    curr_row += 1

    for r_idx, comp_row in enumerate(calc["comparison"]):
        c1 = ws1.cell(row=curr_row, column=1, value=comp_row["metric"])
        c2 = ws1.cell(row=curr_row, column=2, value=comp_row["unit"])
        c3 = ws1.cell(row=curr_row,
                      column=3,
                      value=comp_row["formatted_before"])
        c4 = ws1.cell(row=curr_row,
                      column=4,
                      value=comp_row["formatted_method_a"])
        c5 = ws1.cell(row=curr_row,
                      column=5,
                      value=comp_row["formatted_method_b"])

        for c in [c1, c2, c3, c4, c5]:
            c.font = font_regular
            c.border = cell_border
            c.alignment = Alignment(
                horizontal="center" if c.column > 1 else "left",
                vertical="center")
            if r_idx % 2 == 1:
                c.fill = fill_zebra

        # Highlight winner cell
        winner = comp_row.get("winner", "none")
        if winner == "method_a":
            c4.fill = fill_winner
            c4.font = font_winner
        elif winner == "method_b":
            c5.fill = fill_winner
            c5.font = font_winner
        elif winner == "tie":
            c4.fill = fill_winner
            c5.fill = fill_winner

        ws1.row_dimensions[curr_row].height = 20
        curr_row += 1

    curr_row += 1

    # Recommendation Callout Section
    recs = calc.get("recommendations", [])
    if recs:
        ws1[f'A{curr_row}'] = "Balancing Analysis & Recommendations"
        ws1[f'A{curr_row}'].font = font_section
        curr_row += 1
        for rec in recs:
            ws1[f'A{curr_row}'] = f"• {rec}"
            ws1[f'A{curr_row}'].font = font_regular
            ws1.merge_cells(f'A{curr_row}:E{curr_row}')
            style_range(ws1, f'A{curr_row}:E{curr_row}', border=cell_border)
            ws1[f'A{curr_row}'].alignment = Alignment(wrap_text=True,
                                                      vertical="center")
            curr_row += 1
        curr_row += 1

    # Embed Visual Comparison Charts
    try:
        # Chart 1: Before vs After Overview Charts (2x2 Grid with Reading Guide)
        overview_buf = generate_balancing_overview_chart(calc)
        overview_img = OpenpyxlImage(overview_buf)
        overview_img.width = 920
        overview_img.height = 520
        ws1.add_image(overview_img, f'A{curr_row}')
        curr_row += 28

        # Chart 2: All 8 KPIs Grid (2x4 Grid with exact value labels)
        kpi_grid_buf = generate_all_kpis_card_grid_chart(calc)
        kpi_img = OpenpyxlImage(kpi_grid_buf)
        kpi_img.width = 920
        kpi_img.height = 460
        ws1.add_image(kpi_img, f'A{curr_row}')
        curr_row += 25
    except Exception as e:
        print(f"Warning embedding chart images in Excel: {e}")

    # Auto column width for Sheet 1 (Alt+H+O+I)
    for col in ws1.columns:
        cells_to_check = [
            c for c in col[2:] if c.value and not str(c.value).startswith("•")
            and "Balancing Analysis" not in str(c.value)
        ]
        max_len = max((len(str(cell.value or '')) for cell in cells_to_check),
                      default=10)
        col_letter = get_column_letter(col[0].column)
        ws1.column_dimensions[col_letter].width = max(max_len + 4, 14)

    # =========================================================================
    # TAB 3: Method A (Takt Time) Workstation Table
    # =========================================================================
    ws2 = wb.create_sheet(title="Method A - Takt Time")
    ws2.views.sheetView[0].showGridLines = True

    # Title Banner (Alt+H+B+A & Alt+H+A+C)
    ws2['A1'] = "Method A (Takt Time) — Balanced Workstations"
    ws2['A1'].font = font_title
    ws2.merge_cells('A1:K2')
    ws2['A1'].alignment = Alignment(horizontal="center", vertical="center")
    style_range(ws2, 'A1:K2', border=cell_border)

    df_a = calc["method_a"]["report_df"].rename(columns={"Basic Time": "SAM"})
    cols_a = [
        c for c in [
            "Composite Operations", "Serial/Id", "Operations", "Machine",
            "Predecessor", "SAM", "Combined SAM", "Balancing SAM", "M/P",
            "Takt Time", "Status"
        ] if c in df_a.columns
    ]
    df_a_export = df_a[cols_a] if not df_a.empty else df_a

    headers_a = list(df_a_export.columns)
    for col_idx, col_name in enumerate(headers_a, 1):
        cell = ws2.cell(row=3, column=col_idx, value=col_name)
        cell.font = font_header
        cell.fill = fill_sub_a
        cell.border = cell_border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, row_data in enumerate(df_a_export.values, 4):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            cell.font = font_regular
            cell.border = cell_border
            # Alt + H + A + C: Center all cells horizontally and vertically
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if row_idx % 2 == 1:
                cell.fill = fill_zebra

    # Auto column width for Method A (Alt+H+O+I)
    autofit_columns(ws2, start_row=3, padding=3, min_width=6)

    # =========================================================================
    # TAB 4: Method B (IE Pitch) Workstation Table
    # =========================================================================
    ws3 = wb.create_sheet(title="Method B - IE Pitch")
    ws3.views.sheetView[0].showGridLines = True

    # Title Banner (Alt+H+B+A & Alt+H+A+C)
    ws3['A1'] = "Method B (IE Pitch) — Balanced Workstations"
    ws3['A1'].font = font_title
    ws3.merge_cells('A1:L2')
    ws3['A1'].alignment = Alignment(horizontal="center", vertical="center")
    style_range(ws3, 'A1:L2', border=cell_border)

    df_b = calc["method_b"]["report_df"].rename(columns={"Basic Time": "SAM"})
    cols_b = [
        c for c in [
            "Composite Operations", "Serial/Id", "Operations", "Machine",
            "Predecessor", "SAM", "Combined SAM", "Balancing SAM", "M/P",
            "Pitch Time", "UCL", "Status"
        ] if c in df_b.columns
    ]
    df_b_export = df_b[cols_b] if not df_b.empty else df_b

    headers_b = list(df_b_export.columns)
    for col_idx, col_name in enumerate(headers_b, 1):
        cell = ws3.cell(row=3, column=col_idx, value=col_name)
        cell.font = font_header
        cell.fill = fill_sub_b
        cell.border = cell_border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, row_data in enumerate(df_b_export.values, 4):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws3.cell(row=row_idx, column=col_idx, value=val)
            cell.font = font_regular
            cell.border = cell_border
            # Alt + H + A + C: Center all cells horizontally and vertically
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if row_idx % 2 == 1:
                cell.fill = fill_zebra
            # Highlight review flags in amber
            if "Above UCL" in str(val) or "review" in str(val).lower():
                cell.fill = fill_warning
                cell.font = font_bold

    # Auto column width for Method B (Alt+H+O+I)
    autofit_columns(ws3, start_row=3, padding=3, min_width=6)

    # =========================================================================
    # TAB 2: Before Balancing (Baseline) Operations Table
    # =========================================================================
    ws4 = wb.create_sheet(title="Before Balancing", index=1)
    ws4.views.sheetView[0].showGridLines = True

    # Title Banner (Alt+H+B+A & Alt+H+A+C)
    ws4['A1'] = "Before Balancing — Input Operations Baseline"
    ws4['A1'].font = font_title
    ws4.merge_cells('A1:F2')
    ws4['A1'].alignment = Alignment(horizontal="center", vertical="center")
    style_range(ws4, 'A1:F2', border=cell_border)

    headers_before = [
        "Serial No.", "Operation Name", "Predecessor", "Machine Type", "SAM",
        "Manpower"
    ]
    for col_idx, h in enumerate(headers_before, 1):
        cell = ws4.cell(row=3, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.border = cell_border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, op in enumerate(calc.get("sorted_operations", []), 4):
        preds_str = ",".join(map(str, op.predecessors)) if getattr(
            op, "predecessors", None) else "-"

        # Round SAM to 1 decimal place (e.g. 7.929207 -> 7.9, 21.884868 -> 21.9)
        sam_val = round(float(getattr(op, "basic_time", 0.0) or 0.0), 1)

        vals = [
            getattr(op, "op_id", ""),
            getattr(op, "name", ""), preds_str,
            getattr(op, "machine_type", ""), sam_val, 1
        ]
        for col_idx, val in enumerate(vals, 1):
            cell = ws4.cell(row=row_idx, column=col_idx, value=val)
            cell.font = font_regular
            cell.border = cell_border
            # Alt + H + A + C: Center all cells horizontally and vertically
            cell.alignment = Alignment(horizontal="center", vertical="center")

            # Ensure Excel displays 1 decimal place even for whole numbers (e.g. 7.0)
            if col_idx == 5:
                cell.number_format = '0.0'

            if row_idx % 2 == 1:
                cell.fill = fill_zebra

    # Auto column width for Before Balancing (Alt+H+O+I)
    autofit_columns(ws4, start_row=3, padding=3, min_width=6)

    # Return as buffer
    excel_buf = io.BytesIO()
    wb.save(excel_buf)
    excel_buf.seek(0)
    return excel_buf
